#!/usr/bin/env python3
"""
██╗      █████╗ ██████╗ ██████╗ ██╗   ██╗
██║     ██╔══██╗██╔══██╗██╔══██╗██║   ██║
██║     ███████║██║  ██║██║  ██║██║   ██║
██║     ██╔══██║██║  ██║██║  ██║██║   ██║
███████╗██║  ██║██████╔╝██████╔╝╚██████╔╝
╚══════╝╚═╝  ╚═╝╚═════╝ ╚═════╝  ╚═════╝
         Cold Email Pipeline v2.0

STANDALONE COMMANDS
  python mailclaw.py                 Full interactive pipeline
  python mailclaw.py verify          Email verification only
  python mailclaw.py enrich          AI enrichment only
  python mailclaw.py upload          Upload to Instantly only
  python mailclaw.py map             Column mapping only
  python mailclaw.py balance         Check Reoon + AI credits
  python mailclaw.py config          Manage keys & settings
  python mailclaw.py profiles        Manage enrichment profiles
  python mailclaw.py clients         Manage Instantly clients
  python mailclaw.py onboard         First-time setup
  python mailclaw.py bot             Telegram bot

REQUIRED ENV VARS (or set via config):
  GEMINI_API_KEY / GOOGLE_API_KEY
  ANTHROPIC_API_KEY
  OPENAI_API_KEY
  REOON_API_KEY_1 ... REOON_API_KEY_N
  INSTANTLY_API_KEY

CONFIG: ~/.mailclaw/config.json (local CLI) — or env-only on Railway (see MAILCLAW_CONFIG_SOURCE)
PROFILES: ~/.mailclaw/profiles/<name>.json
STATE: ~/.mailclaw/state/<csv_hash>.json
HISTORY: ~/.mailclaw/email_history.json — or in-memory when MAILCLAW_CONFIG_SOURCE=env
OUTPUT: same folder as input CSV, named:
  {stem}_email_verified_DD_mmm_YY.csv
  {stem}_ai_enriched_DD_mmm_YY.csv
  {stem}_safe_gmail_DD_mmm_YY.csv
  {stem}_catchall_DD_mmm_YY.csv
"""

import os, sys, json, time, csv, re, io, hashlib, argparse, threading, asyncio
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Dict, Tuple, Any, Set, Literal
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import deepcopy

import requests
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import (Progress, SpinnerColumn, BarColumn, TextColumn,
                           TimeElapsedColumn, TaskProgressColumn, MofNCompleteColumn)
from rich.syntax import Syntax
from rich import box
import questionary
from questionary import Style as QStyle

console = Console()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PATHS  (must come before logger — _enable_debug_logging needs APP_DIR)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
APP_DIR      = Path.home() / ".mailclaw"
CONFIG_FILE  = APP_DIR / "config.json"
STATE_DIR    = APP_DIR / "state"
HISTORY_FILE = APP_DIR / "email_history.json"
PROFILES_DIR = APP_DIR / "profiles"
CLIENTS_DIR  = APP_DIR / "clients"     # per-client overrides
ANALYTICS_DIR= APP_DIR / "analytics"   # analytics report profiles
for _d in [APP_DIR, STATE_DIR, PROFILES_DIR, CLIENTS_DIR, ANALYTICS_DIR]:
    _d.mkdir(parents=True, exist_ok=True)

def use_env_config() -> bool:
    """
    True → no config.json / email_history.json on disk; all settings from env vars.
    Set explicitly with MAILCLAW_CONFIG_SOURCE=env or MAILCLAW_USE_ENV=1,
    or implicitly when RAILWAY_ENVIRONMENT is set (Railway injects this).
    Use MAILCLAW_CONFIG_SOURCE=file to force ~/.mailclaw JSON even on Railway (e.g. mounted volume).
    """
    if os.environ.get("MAILCLAW_CONFIG_SOURCE", "").lower() == "file":
        return False
    if os.environ.get("MAILCLAW_CONFIG_SOURCE", "").lower() == "env":
        return True
    if os.environ.get("MAILCLAW_USE_ENV") == "1":
        return True
    if os.environ.get("RAILWAY_ENVIRONMENT"):
        return True
    return False

# In-memory email verification history when use_env_config() (ephemeral across restarts)
_memory_history: Optional[dict] = None

INSTANTLY_BASE = "https://api.instantly.ai/api/v2"
REOON_BASE     = "https://emailverifier.reoon.com/api/v1"
GEMINI_BASE    = "https://generativelanguage.googleapis.com/v1beta/openai/"

# ── Debug logger ─────────────────────────────────────────────────────────────
# Silent by default — logs only go to ~/.mailclaw/mailclaw_debug.log when --debug passed.
# Usage:  python mailclaw.py verify --debug
log = logging.getLogger("mailclaw")
log.setLevel(logging.DEBUG)
log.addHandler(logging.NullHandler())   # NullHandler = zero output unless --debug / stream below

def _configure_stdio_logging():
    """Railway and Docker: emit INFO+ to stderr so deploy logs show config / analytics breadcrumbs."""
    if os.environ.get("MAILCLAW_LOG_STDERR") == "0":
        return
    if not (os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("MAILCLAW_LOG_STDERR") == "1"):
        return
    if any(isinstance(h, logging.StreamHandler) and not isinstance(h, logging.NullHandler) for h in log.handlers):
        return
    sh = logging.StreamHandler(sys.stderr)
    sh.setLevel(logging.INFO)
    sh.setFormatter(logging.Formatter("%(levelname)s [mailclaw] %(message)s"))
    log.addHandler(sh)

_configure_stdio_logging()

def _enable_debug_logging():
    """Attach a rotating file handler to the logger. Called once by main() if --debug set."""
    log_path = APP_DIR / "mailclaw_debug.log"
    fh = RotatingFileHandler(log_path, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(funcName)s:%(lineno)d — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    ))
    log.addHandler(fh)
    log.info("=== Debug logging enabled → %s ===", log_path)
    console.print(f"[dim]Debug log: {log_path}[/]")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  BRANDING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
LOGO = """[bold yellow]
 ██╗      █████╗ ██████╗ ██████╗ ██╗   ██╗
 ██║     ██╔══██╗██╔══██╗██╔══██╗██║   ██║
 ██║     ███████║██║  ██║██║  ██║██║   ██║
 ██║     ██╔══██║██║  ██║██║  ██║██║   ██║
 ███████╗██║  ██║██████╔╝██████╔╝╚██████╔╝
 ╚══════╝╚═╝  ╚═╝╚═════╝ ╚═════╝  ╚═════╝[/]
[dim] cold email pipeline  •  v2.0[/]
"""

Q_STYLE = QStyle([
    ("qmark",       "fg:#ff6b35 bold"),
    ("question",    "fg:#ffffff bold"),
    ("answer",      "fg:#00ff87 bold"),
    ("pointer",     "fg:#ff6b35 bold"),
    ("highlighted", "fg:#ffd700 bold"),
    ("selected",    "fg:#00ff87"),
    ("separator",   "fg:#444444"),
    ("instruction", "fg:#666666 italic"),
])

def out_name(stem: str, suffix: str) -> str:
    """e.g. 'leads' + 'email_verified' -> 'leads_email_verified_16_mar_26.csv'"""
    tag = datetime.now().strftime("%d_%b_%y").lower()
    s   = re.sub(r"[\s()]+", "_", stem)
    s   = re.sub(r"_+", "_", s).strip("_")
    return f"{s}_{suffix}_{tag}.csv"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DEFAULT_CONFIG = {
    "reoon_keys":             [],
    "instantly_clients":      [],
    "model_keys":             {"anthropic":"","gemini":"","openai":""},
    "model_key_pools":        {},  # optional: provider -> [api_key, ...] for rotation
    "model_key_pools_by_client": {},  # client_slug -> {provider: [keys]}
    "telegram_token":         "",
    "telegram_allowed_users": [],
    "daily_limit":            2000,
    "reverify_days":          7,
    "rate_limit_delay":       0.12,
    "default_profile":        "generic",
}

def _env_numbered_strings(base: str) -> List[str]:
    """Read BASE, BASE_2, BASE_3, ... from the environment (non-empty values only)."""
    out: List[str] = []
    v = os.environ.get(base, "").strip()
    if v:
        out.append(v)
    i = 2
    while i <= 64:
        v = os.environ.get(f"{base}_{i}", "").strip()
        if not v:
            break
        out.append(v)
        i += 1
    return out

def _apply_model_key_pools(c: dict) -> None:
    """Merge multi-key env vars into model_key_pools and backfill model_keys when empty."""
    c.setdefault("model_keys", {})
    c.setdefault("model_key_pools", {})
    pools = {
        "gemini":    _env_numbered_strings("GEMINI_API_KEY") or _env_numbered_strings("GOOGLE_API_KEY"),
        "anthropic": _env_numbered_strings("ANTHROPIC_API_KEY"),
        "openai":    _env_numbered_strings("OPENAI_API_KEY"),
    }
    for prov, vals in pools.items():
        if not vals:
            continue
        c["model_key_pools"][prov] = vals
        if not c["model_keys"].get(prov):
            c["model_keys"][prov] = vals[0]
    for prov in ("gemini", "anthropic", "openai"):
        mk = c["model_keys"].get(prov, "").strip()
        if mk and prov not in c["model_key_pools"]:
            c["model_key_pools"][prov] = [mk]

def _client_env_prefix(client_name: str) -> str:
    """
    Instantly client name 'will' → 'WILL'; 'gd-team' → 'GD_TEAM'.
    Used for per-client env vars: WILL_GEMINI_API_KEY, WILL_OPENAI_API_KEY_2, …
    """
    s = re.sub(r"[^a-zA-Z0-9]+", "_", (client_name or "").strip()).strip("_")
    return s.upper() or "CLIENT"

def _env_numbered_strings_prefixed(prefix: str, base: str) -> List[str]:
    """PREFIX_GEMINI_API_KEY, PREFIX_GEMINI_API_KEY_2, … (PREFIX e.g. WILL)."""
    pfx = f"{prefix}_" if prefix else ""
    out: List[str] = []
    v = os.environ.get(f"{pfx}{base}", "").strip()
    if v:
        out.append(v)
    i = 2
    while i <= 64:
        v = os.environ.get(f"{pfx}{base}_{i}", "").strip()
        if not v:
            break
        out.append(v)
        i += 1
    return out

def _merge_client_prefixed_pools(c: dict) -> None:
    """
    For each Instantly client, load optional AI keys:
      <PREFIX>_GEMINI_API_KEY (or GOOGLE_API_KEY), _2, _3…
      <PREFIX>_OPENAI_API_KEY, …
      <PREFIX>_ANTHROPIC_API_KEY, …
    PREFIX = uppercased client name (see _client_env_prefix).
    When set, these override global rotation for that client only.
    """
    c.setdefault("model_key_pools_by_client", {})
    for cl in c.get("instantly_clients") or []:
        nm = cl.get("name") or ""
        slug = nm.lower().strip()
        if not slug:
            continue
        prefix = _client_env_prefix(nm)
        gem = _env_numbered_strings_prefixed(prefix, "GEMINI_API_KEY") or _env_numbered_strings_prefixed(prefix, "GOOGLE_API_KEY")
        ant = _env_numbered_strings_prefixed(prefix, "ANTHROPIC_API_KEY")
        opn = _env_numbered_strings_prefixed(prefix, "OPENAI_API_KEY")
        if not (gem or ant or opn):
            continue
        entry: Dict[str, List[str]] = {}
        if gem:
            entry["gemini"] = gem
        if ant:
            entry["anthropic"] = ant
        if opn:
            entry["openai"] = opn
        c["model_key_pools_by_client"][slug] = entry
        log.debug(
            "mailclaw config: client=%r prefix=%s AI keys gemini=%d anthropic=%d openai=%d",
            slug, prefix, len(gem), len(ant), len(opn),
        )

def _cfg_from_env() -> dict:
    """Full config from environment only (Railway / Docker). No JSON files."""
    c = deepcopy(DEFAULT_CONFIG)
    c["model_keys"] = {"anthropic": "", "gemini": "", "openai": ""}
    c["model_key_pools"] = {}
    c["instantly_clients"] = []
    env_clients = {k[len("INSTANTLY_CLIENT_"):].lower(): v
                   for k, v in os.environ.items() if k.startswith("INSTANTLY_CLIENT_")}
    for name in sorted(env_clients.keys()):
        c["instantly_clients"].append({"name": name, "key": env_clients[name]})

    env_reoon: List[dict] = []
    single = os.environ.get("REOON_KEY", "").strip()
    if single:
        env_reoon.append({"name": "reoon", "key": single})
    i = 1
    while True:
        k_ = os.environ.get(f"REOON_KEY_{i}", "").strip()
        if not k_:
            break
        n_ = os.environ.get(f"REOON_KEY_{i}_NAME", f"reoon{i}")
        env_reoon.append({"name": n_, "key": k_})
        i += 1
    c["reoon_keys"] = env_reoon

    v = os.environ.get("TELEGRAM_TOKEN", "") or os.environ.get("TELEGRAM_BOT_TOKEN", "")
    if v:
        c["telegram_token"] = v.strip()
    raw = os.environ.get("TELEGRAM_ALLOWED_USERS", "")
    if raw:
        try:
            c["telegram_allowed_users"] = [int(x.strip()) for x in raw.split(",") if x.strip()]
        except Exception:
            pass

    for envk, ck, cast in [
        ("MAILCLAW_DAILY_LIMIT", "daily_limit", int),
        ("MAILCLAW_REVERIFY_DAYS", "reverify_days", int),
    ]:
        ev = os.environ.get(envk, "").strip()
        if ev:
            try:
                c[ck] = cast(ev)
            except ValueError:
                pass
    ev = os.environ.get("MAILCLAW_RATE_LIMIT_DELAY", "").strip()
    if ev:
        try:
            c["rate_limit_delay"] = float(ev)
        except ValueError:
            pass

    _apply_model_key_pools(c)
    for prov in ("gemini", "anthropic", "openai"):
        pool = c.get("model_key_pools", {}).get(prov, [])
        if pool:
            c["model_keys"][prov] = pool[0]
    _merge_client_prefixed_pools(c)
    log.debug("cfg_load: env-only merged %d client-specific AI pool(s)", len(c.get("model_key_pools_by_client") or {}))
    return c

def cfg_load() -> dict:
    if use_env_config():
        return _cfg_from_env()

    if not CONFIG_FILE.exists():
        CONFIG_FILE.write_text(json.dumps(DEFAULT_CONFIG, indent=2))
    c = json.loads(CONFIG_FILE.read_text())
    for k, v in DEFAULT_CONFIG.items():
        c.setdefault(k, v)
    c.setdefault("model_keys", {})
    c.setdefault("model_key_pools", {})
    c.setdefault("model_key_pools_by_client", {})
    for mk in ["anthropic", "gemini", "openai"]:
        c["model_keys"].setdefault(mk, "")

    # ── AI provider keys (single env var, if not in file) ────────────────────
    for provider, envs in [("gemini", ["GEMINI_API_KEY", "GOOGLE_API_KEY"]),
                             ("anthropic", ["ANTHROPIC_API_KEY"]),
                             ("openai", ["OPENAI_API_KEY"])]:
        if not c["model_keys"].get(provider):
            for e in envs:
                v = os.environ.get(e, "")
                if v:
                    c["model_keys"][provider] = v
                    break

    _apply_model_key_pools(c)

    # ── Telegram ────────────────────────────────────────────────────────────
    if not c.get("telegram_token"):
        v = os.environ.get("TELEGRAM_TOKEN", "") or os.environ.get("TELEGRAM_BOT_TOKEN", "")
        if v:
            c["telegram_token"] = v
    if not c.get("telegram_allowed_users"):
        raw = os.environ.get("TELEGRAM_ALLOWED_USERS", "")
        if raw:
            try:
                c["telegram_allowed_users"] = [int(x.strip()) for x in raw.split(",") if x.strip()]
            except Exception:
                pass

    # ── Instantly: INSTANTLY_CLIENT_<NAME>=<api_key> ─────────────────────────
    env_clients = {k[len("INSTANTLY_CLIENT_"):].lower(): v
                   for k, v in os.environ.items() if k.startswith("INSTANTLY_CLIENT_")}
    existing_names = {cl["name"].lower() for cl in c.get("instantly_clients", [])}
    for name, key in env_clients.items():
        if name not in existing_names:
            c.setdefault("instantly_clients", []).append({"name": name, "key": key})

    # ── Reoon: REOON_KEY or REOON_KEY_1, REOON_KEY_2, ... ───────────────────
    if not c.get("reoon_keys"):
        env_reoon: List[dict] = []
        single = os.environ.get("REOON_KEY", "")
        if single:
            env_reoon.append({"name": "reoon", "key": single})
        i = 1
        while True:
            k_ = os.environ.get(f"REOON_KEY_{i}", "")
            if not k_:
                break
            n_ = os.environ.get(f"REOON_KEY_{i}_NAME", f"reoon{i}")
            env_reoon.append({"name": n_, "key": k_})
            i += 1
        if env_reoon:
            c["reoon_keys"] = env_reoon

    # ── Analytics profiles: materialize env JSON to disk once (local CLI) ───
    if not use_env_config():
        for k, v in os.environ.items():
            if k.startswith("ANALYTICS_PROFILE_"):
                pname = k[len("ANALYTICS_PROFILE_"):].lower()
                ppath = ANALYTICS_DIR / f"{pname}.json"
                if not ppath.exists():
                    try:
                        ppath.write_text(v, encoding="utf-8")
                        json.loads(v)
                    except Exception:
                        pass

    _merge_client_prefixed_pools(c)
    log.debug("cfg_load: file mode merged %d client-specific AI pool(s)", len(c.get("model_key_pools_by_client") or {}))
    return c

def cfg_save(c: dict):
    if use_env_config():
        log.debug("cfg_save skipped (env-only config)")
        return
    CONFIG_FILE.write_text(json.dumps(c, indent=2))

def get_instantly_client_entry(c: dict, name: str) -> Optional[dict]:
    """Resolve INSTANTLY_CLIENT_* entry by client name (case-insensitive)."""
    if not (name or "").strip():
        return None
    nl = name.lower().strip()
    for cl in c.get("instantly_clients") or []:
        if (cl.get("name") or "").lower() == nl:
            log.debug("instantly client resolved: name=%r", cl.get("name"))
            return cl
    log.debug("instantly client not found: name=%r", name)
    return None

def hist_load() -> dict:
    global _memory_history
    if use_env_config():
        if _memory_history is None:
            _memory_history = {}
        return _memory_history
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text())
        except Exception:
            pass
    return {}

def hist_save(h: dict):
    global _memory_history
    if use_env_config():
        _memory_history = h
        return
    HISTORY_FILE.write_text(json.dumps(h, indent=2))

def hist_mark_verified(results: dict):
    """
    Saves verification results to email history.
    Only updates 'esp' when mx_records is a non-empty list — this means it came
    from a fresh Reoon call. If mx_records is empty (cached result), we preserve
    whatever esp was previously stored rather than overwriting with 'other'.
    """
    h = hist_load(); now = datetime.utcnow().isoformat()
    for email, data in results.items():
        h.setdefault(email, {})
        h[email]["last_verified"]       = now
        h[email]["verification_status"] = data.get("status","unknown")
        mx=data.get("mx_records",[])
        # Only update esp from a real Reoon response (has actual mx_records).
        # Don't overwrite a good cached esp value with "other" from empty mx_records.
        if mx and mx != [] and isinstance(mx, (list,str)):
            h[email]["esp"] = detect_esp(mx)
        elif "esp" not in h[email]:
            h[email]["esp"] = data.get("esp","other")  # first time only
    hist_save(h)

def hist_mark_uploaded(emails: List[str], campaign_id: str, client_name: str):
    h = hist_load(); now = datetime.utcnow().isoformat()
    for em in emails:
        h.setdefault(em, {})
        h[em]["last_uploaded"] = now
        h[em]["last_campaign"] = campaign_id
        h[em]["last_client"]   = client_name
    hist_save(h)

def days_since_verified(email: str, history: dict) -> Optional[int]:
    lv = history.get(email,{}).get("last_verified")
    if not lv: return None
    try: return (datetime.utcnow() - datetime.fromisoformat(lv)).days
    except: return None

def csv_fp(path: Path) -> str:
    s = path.stat()
    return hashlib.md5(f"{path.name}_{s.st_size}_{s.st_mtime}".encode()).hexdigest()[:12]

def state_load(fp: str) -> dict:
    p = STATE_DIR / f"{fp}.json"
    if p.exists():
        try: return json.loads(p.read_text())
        except: pass
    return {}

def state_save(fp: str, s: dict):
    (STATE_DIR / f"{fp}.json").write_text(json.dumps(s, indent=2))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PROFILE SYSTEM
#  ~/.mailclaw/profiles/<name>.json — shared enrichment configs
#  ~/.mailclaw/clients/<name>.json  — per-client overrides
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PROFILE_TEMPLATE = {
    "name": "generic",
    "display_name": "Generic B2B",
    "description": "Generic cold email enrichment profile.",
    # Optional: Instantly client name slug — use this client's prefixed AI keys (WILL_* env vars)
    "ai_client": None,
    # Models
    "enrichment_model": "gemini/gemini-2.5-flash-lite",
    "copy_model":       "anthropic/claude-haiku-4-5",
    "copy_enabled":     False,
    # Tuning
    "workers":                   6,
    "enrichment_max_tokens":     400,
    "copy_max_tokens":           600,
    "enrichment_temperature":    0.3,
    "copy_temperature":          1.0,
    # Input fields forwarded to the model (can be customized)
    "input_fields": [
        "first_name","last_name","job_title","company_name",
        "website","linkedin_url","location","industry","employees"
    ],
    # System prompts — fully editable
    "enrichment_system_prompt": (
        "You are a B2B sales researcher. Given prospect data, return a JSON enrichment object.\n\n"
        "Return ONLY valid JSON, no explanation, no markdown.\n\n"
        "Required JSON fields:\n"
        "{\n"
        "  \"industry_vertical\":    string,\n"
        "  \"company_stage\":        string,\n"
        "  \"primary_pain\":         string,\n"
        "  \"email_angle\":          string,\n"
        "  \"personalization_hook\": string,\n"
        "  \"confidence\":           integer  // 0-100\n"
        "}\n\n"
        "email_angle must be one of: efficiency | growth | risk | cost | competitive"
    ),
    # Which fields to pull from the enrichment JSON response
    "enrichment_output_fields": [
        "industry_vertical","company_stage","primary_pain",
        "email_angle","personalization_hook","confidence"
    ],
    # Optional copy step
    "copy_system_prompt": (
        "You write short cold emails. Return ONLY valid JSON:\n"
        "{\n"
        "  \"email1_subject\": \"2-4 words, all lowercase\",\n"
        "  \"email1_body\":    \"2-3 sentences, 50-70 words, lowercase except I/proper nouns\",\n"
        "  \"email2_subject\": \"re: then 2-4 words\",\n"
        "  \"email2_body\":    \"exactly 2 sentences, 35-50 words, end with a question\"\n"
        "}\n\n"
        "Rules: no greetings, no links, no em dashes, no bullets, no CTA in body."
    ),
    "copy_output_fields": ["email1_subject","email1_body","email2_subject","email2_body"],
    "copy_cta":           "worth 15 min this week?",
    "sender_name":        "",
    # Custom variables: extra keys to always add to every output row (static values)
    "custom_static_vars": {},
}

CLIENT_TEMPLATE = {
    "name":             "client_name",
    "display_name":     "Client Name",
    "instantly_key":    "",
    # Optional: override the default profile for this client
    "profile_override": None,
    # Custom variables to add to every lead uploaded for this client
    # These get injected into the Instantly 'payload' as custom vars
    "custom_upload_vars": {},
    # If set, these fields will override the profile's enrichment prompts
    "prompt_overrides": {
        # "enrichment_system_prompt": "...",
        # "copy_system_prompt": "...",
        # "copy_cta": "...",
        # "sender_name": "...",
    },
}

def _ensure_default_profiles():
    p = PROFILES_DIR / "generic.json"
    if not p.exists():
        p.write_text(json.dumps(PROFILE_TEMPLATE, indent=2))

def profile_load(name: str) -> Optional[dict]:
    _ensure_default_profiles()
    p = PROFILES_DIR / f"{name}.json"
    if p.exists():
        try: return json.loads(p.read_text())
        except: return None
    return None

def profile_save(profile: dict):
    (PROFILES_DIR / f"{profile['name']}.json").write_text(json.dumps(profile, indent=2))

def profiles_all() -> Dict[str, dict]:
    _ensure_default_profiles()
    out = {}
    for f in sorted(PROFILES_DIR.glob("*.json")):
        try:
            p = json.loads(f.read_text())
            out[p.get("name", f.stem)] = p
        except: pass
    return out

def client_config_load(name: str) -> dict:
    p = CLIENTS_DIR / f"{name}.json"
    if p.exists():
        try: return json.loads(p.read_text())
        except: pass
    return dict(CLIENT_TEMPLATE)

def client_config_save(cfg_client: dict):
    name = cfg_client.get("name","unnamed")
    (CLIENTS_DIR / f"{name}.json").write_text(json.dumps(cfg_client, indent=2))

def resolve_profile_for_client(client_name: str) -> dict:
    """Return the effective profile for a client, merging any overrides."""
    cfg          = cfg_load()
    client_meta  = next((c for c in cfg["instantly_clients"] if c["name"]==client_name), {})
    client_extra = client_config_load(client_name)

    # Pick base profile
    profile_name  = (client_extra.get("profile_override")
                     or client_meta.get("profile_override")
                     or cfg.get("default_profile","generic"))
    base_profile  = profile_load(profile_name) or dict(PROFILE_TEMPLATE)
    result        = deepcopy(base_profile)

    # Apply prompt overrides from client config
    for k, v in (client_extra.get("prompt_overrides") or {}).items():
        if v: result[k] = v
    return result

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  MODEL REGISTRY
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MODELS: Dict[str, dict] = {
    "gemini/gemini-2.5-flash-lite": {
        "provider":"gemini","model_id":"gemini-2.5-flash-lite",
        "cost_in":0.10,"cost_out":0.40,"label":"Gemini 2.5 Flash-Lite (cheapest, fast)"},
    "gemini/gemini-2.5-flash": {
        "provider":"gemini","model_id":"gemini-2.5-flash",
        "cost_in":0.15,"cost_out":0.60,"label":"Gemini 2.5 Flash"},
    "gemini/gemini-2.0-flash": {
        "provider":"gemini","model_id":"gemini-2.0-flash",
        "cost_in":0.10,"cost_out":0.40,"label":"Gemini 2.0 Flash"},
    "anthropic/claude-haiku-4-5": {
        "provider":"anthropic","model_id":"claude-haiku-4-5-20251001",
        "cost_in":1.00,"cost_out":5.00,"label":"Claude Haiku 4.5 ($1/$5 per 1M tokens)"},
    "anthropic/claude-sonnet-4-5": {
        "provider":"anthropic","model_id":"claude-sonnet-4-5-20250929",
        "cost_in":3.00,"cost_out":15.00,"label":"Claude Sonnet 4.5 ($3/$15 per 1M tokens)"},
    "anthropic/claude-sonnet-4": {
        "provider":"anthropic","model_id":"claude-sonnet-4-20250514",
        "cost_in":3.00,"cost_out":15.00,"label":"Claude Sonnet 4 ($3/$15 per 1M tokens)"},
    "openai/gpt-4o-mini": {
        "provider":"openai","model_id":"gpt-4o-mini",
        "cost_in":0.15,"cost_out":0.60,"label":"GPT-4o mini ($0.15/$0.60 per 1M tokens)"},
    "openai/gpt-4o": {
        "provider":"openai","model_id":"gpt-4o",
        "cost_in":2.50,"cost_out":10.00,"label":"GPT-4o ($2.50/$10 per 1M tokens)"},
}
_ai_clients: Dict[Tuple[str, str, str], Any] = {}
_ai_provider_rr: Dict[str, int] = {}

def _ai_rr_scope_key(prov: str, client_slug: Optional[str]) -> str:
    """Separate round-robin state per (provider × client scope)."""
    return f"{prov}\t{(client_slug or '').lower() or '__global__'}"

def _provider_has_keys(c: dict, prov: str, client_slug: Optional[str] = None) -> bool:
    scope = (client_slug or "").strip().lower() or None
    if scope:
        byc = c.get("model_key_pools_by_client", {}).get(scope, {})
        pl = [x for x in byc.get(prov, []) if x and str(x).strip()]
        if pl:
            return True
    pool = [x for x in c.get("model_key_pools", {}).get(prov, []) if x and str(x).strip()]
    if pool:
        return True
    return bool((c.get("model_keys") or {}).get(prov, ""))

def _pick_api_key_for_provider(c: dict, prov: str, client_slug: Optional[str] = None) -> str:
    scope = (client_slug or "").strip().lower() or None
    if scope:
        byc = c.get("model_key_pools_by_client", {}).get(scope, {})
        pool = [x for x in byc.get(prov, []) if x and str(x).strip()]
        if pool:
            sk = _ai_rr_scope_key(prov, scope)
            idx = _ai_provider_rr.get(sk, 0) % len(pool)
            log.debug(
                "AI key: provider=%s scope=client:%s idx=%d/%d tail=%s",
                prov, scope, idx, len(pool), pool[idx][-4:] if pool[idx] else "?",
            )
            return pool[idx]
    pool = [x for x in c.get("model_key_pools", {}).get(prov, []) if x and str(x).strip()]
    if not pool:
        k = (c.get("model_keys") or {}).get(prov, "")
        pool = [k] if k else []
    if not pool:
        log.debug("AI key: provider=%s scope=global — empty pool", prov)
        return ""
    sk = _ai_rr_scope_key(prov, None)
    idx = _ai_provider_rr.get(sk, 0) % len(pool)
    log.debug(
        "AI key: provider=%s scope=global idx=%d/%d tail=%s",
        prov, idx, len(pool), pool[idx][-4:] if pool[idx] else "?",
    )
    return pool[idx]

def _bump_ai_provider(prov: str, client_slug: Optional[str] = None):
    sk = _ai_rr_scope_key(prov, client_slug)
    _ai_provider_rr[sk] = _ai_provider_rr.get(sk, 0) + 1
    log.debug("AI key bump: %s", sk)

def _invalidate_ai_clients_for_model(model_key: str):
    for k in list(_ai_clients.keys()):
        if k[0] == model_key:
            del _ai_clients[k]

def ai_cost(model_key:str,t_in:int,t_out:int)->float:
    m=MODELS.get(model_key,{})
    return t_in/1e6*m.get("cost_in",0)+t_out/1e6*m.get("cost_out",0)

def ai_cost_est(model_key:str,n:int,avg_in:int=450,avg_out:int=250)->float:
    return ai_cost(model_key,n*avg_in,n*avg_out)

def _get_ai_client(model_key: str, client_slug: Optional[str] = None):
    c = cfg_load()
    m = MODELS.get(model_key, {})
    prov = m.get("provider", "")
    api_key = _pick_api_key_for_provider(c, prov, client_slug)
    if not api_key:
        raise ValueError(f"No API key configured for provider '{prov}' (model {model_key})")
    scope_tag = (client_slug.strip().lower() if client_slug else "") or "__global__"
    ck = (model_key, api_key, scope_tag)
    if ck in _ai_clients:
        return _ai_clients[ck]
    log.debug("AI client construct: model=%s scope=%s provider=%s", model_key, scope_tag, prov)
    if prov == "gemini":
        from openai import OpenAI
        cl = OpenAI(api_key=api_key, base_url=GEMINI_BASE)
    elif prov == "anthropic":
        import anthropic as _ant
        cl = _ant.Anthropic(api_key=api_key)
    elif prov == "openai":
        from openai import OpenAI
        cl = OpenAI(api_key=api_key)
    else:
        raise ValueError(f"Unknown model: {model_key}")
    _ai_clients[ck] = cl
    return cl

def ai_call(model_key: str, system: str, user: str,
            max_tokens: int = 500, temperature: float = 0.5,
            client_slug: Optional[str] = None) -> Tuple[str, int, int]:
    """
    Calls an AI model and returns (text, tokens_in, tokens_out).
    Retries up to 3 times with exponential backoff on any exception.

    Provider routing:
    - anthropic/* → uses anthropic SDK client.messages.create()
    - gemini/*    → uses OpenAI-compat SDK via GEMINI_BASE URL
    - openai/*    → uses OpenAI SDK directly

    Token counts are read from usage field; fall back to max_tokens estimate if absent.
    """
    if model_key not in MODELS:
        raise ValueError(f"Unknown model key '{model_key}'. Known models: {list(MODELS.keys())}")
    m = MODELS[model_key]
    prov = m["provider"]
    mid = m["model_id"]
    log.debug(
        "ai_call model=%s provider=%s client_slug=%s max_tokens=%d temp=%.2f user_preview=%r",
        model_key, prov, client_slug or "", max_tokens, temperature, user[:80],
    )
    for attempt in range(3):
        cl = _get_ai_client(model_key, client_slug)
        try:
            if prov == "anthropic":
                import anthropic as _ant
                r = cl.messages.create(model=mid, max_tokens=max_tokens, system=system,
                                       messages=[{"role": "user", "content": user}])
                t_in = r.usage.input_tokens if r.usage else max_tokens
                t_out = r.usage.output_tokens if r.usage else max_tokens // 2
                text = r.content[0].text.strip()
            else:
                r = cl.chat.completions.create(model=mid, max_tokens=max_tokens, temperature=temperature,
                    messages=[{"role": "system", "content": system}, {"role": "user", "content": user}])
                t_in = r.usage.prompt_tokens if r.usage else max_tokens
                t_out = r.usage.completion_tokens if r.usage else max_tokens // 2
                text = r.choices[0].message.content.strip()
            cost = ai_cost(model_key, t_in, t_out)
            log.debug("ai_call done: in=%d out=%d cost=$%.6f response_preview=%r", t_in, t_out, cost, text[:80])
            return text, t_in, t_out
        except Exception as e:
            log.warning("ai_call attempt %d/%d failed: %s", attempt + 1, 3, e)
            _bump_ai_provider(prov, client_slug)
            _invalidate_ai_clients_for_model(model_key)
            if attempt == 2:
                raise
            time.sleep(2 ** attempt)
    raise RuntimeError("ai_call exhausted retries")

def parse_json(text:str)->dict:
    text=re.sub(r"```[a-z]*","",text).strip().strip("`").strip()
    return json.loads(text)

def cheapest_available_model(client_slug: Optional[str] = None) -> Optional[str]:
    c = cfg_load()
    order = ["gemini/gemini-2.5-flash-lite", "gemini/gemini-2.0-flash",
             "openai/gpt-4o-mini", "anthropic/claude-haiku-4-5"]
    for mk in order:
        if mk not in MODELS:
            continue
        prov = MODELS[mk]["provider"]
        if _provider_has_keys(c, prov, client_slug):
            log.debug("cheapest_available_model → %s (client_slug=%r)", mk, client_slug)
            return mk
    return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  COLUMN MAPPING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TARGET_FIELDS = [
    "email","first_name","last_name","company_name","website",
    "phone","title","location","linkedin_url","icebreaker",
    "industry","employees","keywords",
]
FIELD_HINTS: Dict[str,List[str]] = {
    "email":        ["email","e_mail","email_address","emailaddress","mail"],
    "first_name":   ["first_name","firstname","first","fname","given_name","f_name"],
    "last_name":    ["last_name","lastname","last","lname","surname","family_name","l_name"],
    "company_name": ["company","company_name","companyname","organization","org","employer","business","account_name"],
    "website":      ["website","domain","company_domain","url","web","site","homepage","company_url"],
    "phone":        ["phone","phone_number","phonenumber","mobile","cell","tel","telephone"],
    "title":        ["title","job_title","jobtitle","position","role","designation","function"],
    "location":     ["location","city","country","state","region","geo","address","hq","headquarters"],
    "linkedin_url": ["linkedin","linkedin_url","linkedin_profile","li_url","profile_url"],
    "icebreaker":   ["icebreaker","personalization","opener","intro","first_line","opening_line"],
    "industry":     ["industry","sector","vertical","niche","category"],
    "employees":    ["employees","employee_count","headcount","company_size","employees_range","num_employees"],
    "keywords":     ["keywords","tags","signals","topics"],
}

def _norm(s:str)->str:
    return re.sub(r"[^a-z0-9]","_",s.lower().strip())

def heuristic_map(columns:List[str])->Dict[str,Optional[str]]:
    result={f:None for f in TARGET_FIELDS}; used=set()
    for target,hints in FIELD_HINTS.items():
        for col in columns:
            if col in used: continue
            if _norm(col) in hints: result[target]=col; used.add(col); break
        if result[target]: continue
        for col in columns:
            if col in used: continue
            n=_norm(col)
            for h in hints:
                if h in n or n in h: result[target]=col; used.add(col); break
            if result[target]: break
    return result

def ai_map(columns:List[str],sample_rows:List[dict],model_key:str)->Dict[str,Optional[str]]:
    sample=[{c:r.get(c,"") for c in columns[:20]} for r in sample_rows[:3]]
    prompt=(f"You MUST map these CSV columns to standard CRM fields.\n\n"
            f"COLUMNS: {json.dumps(columns)}\n\n"
            f"SAMPLE DATA (first 3 rows):\n{json.dumps(sample,indent=2)}\n\n"
            f"TARGET FIELDS: {TARGET_FIELDS}\n\n"
            f"Rules:\n"
            f"- Each TARGET maps to AT MOST ONE source column\n"
            f"- Use null if no reasonable match\n"
            f"- 'email' is the most critical - never leave null if any column could be email\n"
            f"- Be smart: 'Company' -> company_name, 'Job Title' -> title, etc\n\n"
            f"Return ONLY JSON, no explanation: {{\"email\": \"Email\", \"first_name\": \"First\", ...}}")
    try:
        text,_,_=ai_call(model_key,"You are a data engineer mapping CSV columns. Return only valid JSON.",
                         prompt,max_tokens=400,temperature=0.0)
        mapping=parse_json(text)
        return {f:(mapping.get(f) if mapping.get(f) in columns else None) for f in TARGET_FIELDS}
    except Exception as e:
        console.print(f"[yellow]AI column mapping error ({e}) — using heuristics[/]")
        return heuristic_map(columns)

def show_mapping_preview(columns:List[str],mapping:Dict[str,Optional[str]],
                         sample_rows:List[dict])->None:
    """Show a rich table previewing how columns will be mapped + sample values."""
    t=Table(title="[bold]Column Mapping Preview[/]",box=box.ROUNDED,show_lines=True)
    t.add_column("Standard Field",style="cyan bold",width=18)
    t.add_column("→ CSV Column",  style="green",width=22)
    t.add_column("Sample Value 1",style="dim",width=25)
    t.add_column("Sample Value 2",style="dim",width=25)
    for f in TARGET_FIELDS:
        src=mapping.get(f)
        v1=str(sample_rows[0].get(src,""))[:24] if src and sample_rows      else "—"
        v2=str(sample_rows[1].get(src,""))[:24] if src and len(sample_rows)>1 else "—"
        t.add_row(f, src or "[dim]unmapped[/]", v1, v2)
    console.print(t)
    # Warn about unmapped critical fields
    if not mapping.get("email"):
        console.print("[bold red]⚠  WARNING: No email column mapped! Verification and upload will fail.[/]")

def do_column_mapping(columns:List[str],rows:List[dict],
                      saved_map:Optional[dict]=None,
                      force_remap:bool=False)->Dict[str,Optional[str]]:
    """
    Full interactive column mapping flow:
    1. Use saved mapping if exists (unless force_remap)
    2. Try AI mapping, fallback to heuristic
    3. Show preview with sample values
    4. User confirms or edits field-by-field
    5. Ask "apply same mapping for ALL similar CSVs?" (save globally)
    """
    if saved_map and not force_remap:
        console.print("[dim]Loaded column mapping from previous session.[/]")
        if not questionary.confirm("Re-run column mapping?",default=False,style=Q_STYLE).ask():
            show_mapping_preview(columns, saved_map, rows)
            return saved_map

    # Auto-map
    model_key=cheapest_available_model()
    if model_key:
        console.print(f"[dim]Auto-mapping columns via [bold]{model_key}[/]…[/]")
        mapping=ai_map(columns, rows, model_key)
    else:
        console.print("[dim]Auto-mapping with heuristics (no AI key configured)…[/]")
        mapping=heuristic_map(columns)

    show_mapping_preview(columns, mapping, rows)

    # Confirm or edit
    action=questionary.select(
        "Column mapping — what would you like to do?",
        choices=[
            "✅  Looks good — use this mapping",
            "✏️   Edit specific fields",
            "🔄  Edit ALL fields",
        ], style=Q_STYLE
    ).ask()

    if action and "Edit specific" in action:
        fields_to_edit=questionary.checkbox(
            "Which fields to remap?",
            choices=TARGET_FIELDS, style=Q_STYLE
        ).ask() or []
        for f in fields_to_edit:
            choices=["— unmapped —"]+columns
            pick=questionary.select(
                f"Map '[cyan]{f}[/]' to:",
                choices=choices,
                default=mapping.get(f) or choices[0],
                style=Q_STYLE
            ).ask()
            mapping[f]=None if (not pick or pick.startswith("—")) else pick

    elif action and "Edit ALL" in action:
        for f in TARGET_FIELDS:
            choices=["— unmapped —"]+columns
            pick=questionary.select(
                f"[cyan]{f}[/]:",
                choices=choices,
                default=mapping.get(f) or choices[0],
                style=Q_STYLE
            ).ask()
            mapping[f]=None if (not pick or pick.startswith("—")) else pick

    # Show final mapping
    console.print("\n[bold green]Final mapping:[/]")
    show_mapping_preview(columns, mapping, rows)
    if not questionary.confirm("Confirm this mapping?",default=True,style=Q_STYLE).ask():
        return do_column_mapping(columns, rows, force_remap=True)  # retry

    return mapping

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  HTTP + RETRY
#  Instantly rate limit: 100 req/10s, 600 req/min (workspace-wide).
#  With delay=0.12s we sit at ~8 req/s — safely under the limit.
#  Reoon has no published rate limit; we don't hammer it.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _req(method:str,url:str,delay:float=0.0,retries:int=5,**kw)->requests.Response:
    """
    Sends an HTTP request with automatic retry on 429 (rate limit) and 5xx.
    - 429: honours Retry-After header, backs off exponentially otherwise.
    - 5xx: exponential backoff, no retry on 4xx other than 429.
    - ConnectionError: retries with increasing sleep.
    - delay: seconds to sleep AFTER a successful response (used for rate-limiting).
    All requests and responses are logged at DEBUG level (only when --debug active).
    """
    log.debug("HTTP %s %s | delay=%.2f retries=%d", method, url, delay, retries)
    for attempt in range(retries):
        try:
            r=requests.request(method,url,timeout=30,**kw)
            log.debug("← %d %s (attempt %d)", r.status_code, url.split("?")[0], attempt+1)
            if r.status_code==429:
                wait=float(r.headers.get("Retry-After",2**attempt*2))
                log.warning("Rate limited on %s — waiting %.0fs", url.split("?")[0], wait)
                console.print(f"[yellow]Rate limited — waiting {wait:.0f}s…[/]")
                time.sleep(wait); continue
            if r.status_code>=500 and attempt<retries-1:
                wait=2**attempt
                log.warning("HTTP %d from %s — retry %d in %.0fs", r.status_code, url.split("?")[0], attempt+1, wait)
                time.sleep(wait); continue
            if delay: time.sleep(delay)
            return r
        except requests.exceptions.ConnectionError as e:
            log.warning("ConnectionError %s (attempt %d): %s", url.split("?")[0], attempt+1, e)
            if attempt<retries-1:
                console.print(f"[yellow]Connection error — retry {attempt+1}/{retries}…[/]")
                time.sleep(3*(attempt+1))
            else: raise
    return r

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  REOON KEY ROTATOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class ReoonRotator:
    """
    Rotates Reoon API keys based on daily usage.
    IMPORTANT: used_today is synced from live API on init if keys are fresh.
    The live 'remaining_daily_credits' is always the source of truth.
    Local tracking only increments on top of the last synced value.
    """
    def __init__(self, keys:List[dict], daily_limit:int):
        self.keys=keys; self.limit=daily_limit; self._reset_if_new_day()

    def _reset_if_new_day(self):
        today=date.today().isoformat(); changed=False
        for k in self.keys:
            if k.get("last_reset")!=today:
                k["used_today"]=0; k["last_reset"]=today; changed=True
        if changed:
            c=cfg_load(); c["reoon_keys"]=self.keys; cfg_save(c)

    def sync_from_live(self, silent:bool=False) -> bool:
        """
        Fetch live remaining_daily_credits from Reoon for each key and update
        local used_today accordingly. This is the source of truth — credits used
        via the dashboard, other tools, or previous mailclaw sessions are all reflected.
        Returns True if at least one key synced successfully.
        """
        c=cfg_load(); any_synced=False
        for k in self.keys:
            bal=reoon_balance(k["key"])
            if bal.get("status")=="success":
                live_remaining = int(bal.get("remaining_daily_credits", 0))
                # used_today = configured_limit - live_remaining
                # If live_remaining > limit, treat used_today as 0 (account has more than our limit)
                k["used_today"] = max(0, self.limit - live_remaining)
                k["live_remaining"] = live_remaining            # store for display
                k["live_instant"]   = int(bal.get("remaining_instant_credits", 0))
                any_synced=True
                log.debug("sync_from_live: key=%s live_remaining=%d → used_today=%d",
                          k["name"], live_remaining, k["used_today"])
            else:
                if not silent:
                    console.print(f"  [yellow]⚠ Could not fetch live balance for {k['name']}[/]")
                log.warning("sync_from_live failed for key %s", k["name"])
        c["reoon_keys"]=self.keys; cfg_save(c)
        return any_synced

    def available(self)->List[dict]:
        # Use live_remaining if we have it (more accurate), else fall back to local calc
        result=[]
        for k in self.keys:
            if "live_remaining" in k:
                if k["live_remaining"] > 0: result.append(k)
            else:
                if k.get("used_today",0) < self.limit: result.append(k)
        return result

    def remaining(self,k:dict)->int:
        # Live remaining is always more accurate than local calculation
        if "live_remaining" in k:
            return k["live_remaining"]
        return self.limit - k.get("used_today",0)

    def total_remaining(self)->int:
        return sum(self.remaining(k) for k in self.keys)

    def record(self,name:str,count:int):
        """Record usage locally AND decrement live_remaining if present."""
        for k in self.keys:
            if k["name"]==name:
                k["used_today"]=k.get("used_today",0)+count
                if "live_remaining" in k:
                    k["live_remaining"]=max(0,k["live_remaining"]-count)
        c=cfg_load(); c["reoon_keys"]=self.keys; cfg_save(c)

    def status_table(self)->Table:
        t=Table(title="[bold]Reoon API Keys[/]",box=box.ROUNDED,show_lines=True)
        t.add_column("Name",style="cyan bold")
        t.add_column("Remaining Today",justify="right",style="green")
        t.add_column("Instant Credits",justify="right",style="dim")
        t.add_column("Source",justify="center",style="dim")
        t.add_column("Status",justify="center")
        for k in self.keys:
            rem = self.remaining(k)
            instant = str(k.get("live_instant","—"))
            source  = "[green]live[/]" if "live_remaining" in k else "[yellow]local[/]"
            status  = "[green]✓ Active[/]" if rem>0 else "[red]✗ Exhausted[/]"
            t.add_row(k["name"], str(rem), instant, source, status)
        return t

def reoon_balance(key:str)->dict:
    """GET /api/v1/check-account-balance/ — returns remaining daily + instant credits."""
    try:
        r=_req("GET",f"{REOON_BASE}/check-account-balance/?key={key}")
        log.debug("Reoon balance response: %s", r.text[:200])
        return r.json() if r.ok else {}
    except Exception as e:
        log.error("reoon_balance failed: %s", e)
        return {}

def reoon_submit_bulk(emails:List[str],key:str,name:str="Mailclaw")->Tuple[Optional[int],str]:
    """
    POST /api/v1/create-bulk-verification-task/
    Body: {name (max 25 chars), emails (max 50,000), key}
    Returns (task_id, error_string). task_id is an integer on success.
    Success = HTTP 201. Error = HTTP 4xx with {status: "error", reason: "..."}.
    """
    payload={"name":name[:25],"emails":emails,"key":key}
    log.debug("Reoon submit_bulk: %d emails, task_name=%r", len(emails), name[:25])
    try:
        r=_req("POST",f"{REOON_BASE}/create-bulk-verification-task/",json=payload)
        d=r.json()
        log.debug("Reoon submit_bulk response [%d]: %s", r.status_code, str(d)[:200])
        return (d["task_id"],"") if r.status_code==201 else (None,d.get("reason",f"HTTP {r.status_code}: {r.text[:80]}"))
    except Exception as e:
        log.error("reoon_submit_bulk exception: %s", e)
        return None, str(e)

def reoon_poll(tid:int,key:str)->Optional[dict]:
    """
    GET /api/v1/get-result-bulk-verification-task/
    Params: key={key}&task-id={tid}
    NOTE: Reoon docs (Python example) use 'task-id' (hyphen), not 'task_id'.
    We send BOTH to handle any server-side normalisation.
    Possible statuses: 'waiting', 'running', 'completed', 'file_not_found', 'file_loading_error'.
    On 'completed': response includes 'results' dict keyed by email.
    """
    log.debug("Reoon poll task_id=%d", tid)
    try:
        # Send both 'task-id' (documented) and 'task_id' (common alternative)
        r=_req("GET",f"{REOON_BASE}/get-result-bulk-verification-task/",
               params={"key":key,"task-id":tid,"task_id":tid})
        d=r.json() if r.ok else None
        if d:
            log.debug("Reoon poll status=%s pct=%.1f checked=%s total=%s",
                      d.get("status"), d.get("progress_percentage",0),
                      d.get("count_checked"), d.get("count_total"))
        return d
    except Exception as e:
        log.error("reoon_poll exception: %s", e)
        return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ESP + TIMEZONE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def detect_esp(mx)->str:
    # mx_records from Reoon can be a list, a string, or None depending on the email/result
    if not mx: return "other"
    if isinstance(mx, str): mx = [mx]          # single string → wrap in list
    if not isinstance(mx, list): return "other" # unexpected type → safe default
    s=" ".join(str(x) for x in mx).lower()
    if any(p in s for p in ["google","gmail"]):                    return "gmail"
    if any(p in s for p in ["outlook","microsoft","hotmail","office365"]): return "outlook"
    if any(p in s for p in ["yahoo","yahoodns"]):                  return "yahoo"
    return "other"

EUROPE_W={"uk","gb","de","fr","nl","se","no","dk","fi","es","it","pl","ch","at","be","ie","pt","cz","ro","hu","sk","hr","bg","lt","lv","ee","si","lu","mt","cy"}
EUROPE_P=["united kingdom","germany","france","netherlands","sweden","norway","denmark","finland","spain","italy","poland","switzerland","austria","belgium","ireland","europe","european","portugal","czech","romania","hungary","bulgaria"]

def detect_region(loc:str)->str:
    if not loc: return "america"
    words=set(re.split(r"[\s,;/|()\-]+",loc.lower()))-{""}
    if words&EUROPE_W: return "europe"
    for p in EUROPE_P:
        if p in loc.lower(): return "europe"
    return "america"

def campaign_schedule(region:str)->dict:
    # Instantly V2 timezone enum — America/New_York is NOT in their list.
    # Confirmed valid values: America/Chicago (US Central/Eastern) and Europe/London.
    tz="Europe/London" if region=="europe" else "America/Chicago"
    return {"campaign_schedule":{"schedules":[{
        "name":f"{'Europe' if region=='europe' else 'US'} Business Hours",
        "timing":{"from":"08:00","to":"17:00"},
        "days":{"sunday":False,"monday":True,"tuesday":True,"wednesday":True,
                "thursday":True,"friday":True,"saturday":False},
        "timezone":tz}]}}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CSV HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def csv_read(path:Path)->Tuple[List[dict],List[str]]:
    with open(path,"r",encoding="utf-8-sig",errors="replace") as f:
        r=csv.DictReader(f); rows=list(r); cols=list(r.fieldnames or [])
    return rows, cols

def csv_write(rows:List[dict],path:Path):
    if not rows: return
    path.parent.mkdir(parents=True,exist_ok=True)
    with open(path,"w",newline="",encoding="utf-8") as f:
        w=csv.DictWriter(f,fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    console.print(f"[green]✓[/] Written: [yellow]{path.name}[/] ({len(rows)} rows)")

def merge_with_history(rows:List[dict],email_col:str,hist:dict)->List[dict]:
    """Add lp_* history columns to each row."""
    out=[]
    for r in rows:
        em=r.get(email_col,"").lower().strip(); rec=hist.get(em,{})
        out.append({**r,
            "lp_last_verified": rec.get("last_verified",""),
            "lp_last_uploaded": rec.get("last_uploaded",""),
            "lp_last_campaign": rec.get("last_campaign",""),
            "lp_status":        rec.get("verification_status",""),
            "lp_esp":           rec.get("esp",""),
        })
    return out

def find_csvs()->List[Path]:
    bases=[Path.cwd(),Path.home()/"Downloads",Path.home()/"Desktop",Path.home()/"Documents"]
    found={p.resolve() for b in bases if b.exists() for p in b.glob("*.csv")}
    return sorted(found,key=lambda p:p.stat().st_mtime,reverse=True)

def pick_csv()->Optional[Path]:
    all_csvs=find_csvs()
    if not all_csvs:
        m=questionary.path("No CSVs found. Enter path:",style=Q_STYLE).ask()
        return Path(m) if m else None
    offset=0; page=20
    while True:
        batch=all_csvs[offset:offset+page]; choices=[]
        for p in batch:
            fp_=csv_fp(p); st=state_load(fp_)
            mod=datetime.fromtimestamp(p.stat().st_mtime).strftime("%b %d %H:%M")
            sz=p.stat().st_size//1024
            tags=[]
            if st.get("enrich",{}).get("completed_at"):          tags.append("[magenta]✦ enriched[/]")
            if st.get("verify",{}).get("status")=="completed":   tags.append("[green]✓ verified[/]")
            if st.get("upload",{}).get("uploaded_at"):           tags.append("[cyan]↑ uploaded[/]")
            tag="  "+" ".join(tags) if tags else ""
            choices.append(questionary.Choice(
                title=f"{mod}  {p.name}  [dim]{sz}KB[/]{tag}", value=str(p)))
        if offset+page<len(all_csvs):
            choices.append(questionary.Choice("[dim]── load 20 more ──[/]","__more__"))
        choices.append(questionary.Choice("[dim]── browse manually ──[/]","__manual__"))
        pick=questionary.select(
            f"[bold]Select CSV[/]  [dim]({offset+1}–{min(offset+page,len(all_csvs))} of {len(all_csvs)})[/]",
            choices=choices,style=Q_STYLE).ask()
        if pick is None:        return None
        if pick=="__more__":    offset+=page; continue
        if pick=="__manual__":
            m=questionary.path("Path to CSV:",style=Q_STYLE).ask()
            return Path(m) if m else None
        return Path(pick)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ENRICHMENT STAGE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_enrich_lock = threading.Lock()

def _build_enrich_input(row:dict,profile:dict)->str:
    fields={f:str(row.get(f,"")).strip()
            for f in profile.get("input_fields",TARGET_FIELDS)
            if str(row.get(f,"")).strip().lower() not in ("","nan","none")}
    return f"Prospect data:\n{json.dumps(fields,indent=2)}"

def _enrich_one(row:dict,profile:dict)->Tuple[dict,float]:
    em_key=profile.get("enrichment_model","gemini/gemini-2.5-flash-lite")
    cp_key=profile.get("copy_model","anthropic/claude-haiku-4-5")
    cost=0.0; result={}
    _slug = (profile.get("ai_client") or "").strip().lower() or None
    if _slug:
        log.debug("enrich: using ai_client=%r for AI key pools", _slug)

    text,t_in,t_out=ai_call(em_key,profile["enrichment_system_prompt"],
                             _build_enrich_input(row,profile),
                             max_tokens=profile.get("enrichment_max_tokens",400),
                             temperature=profile.get("enrichment_temperature",0.3),
                             client_slug=_slug)
    enr=parse_json(text); cost+=ai_cost(em_key,t_in,t_out)
    for f in profile.get("enrichment_output_fields",[]): result[f]=enr.get(f,"")

    # Add any static custom vars from profile
    for k,v in (profile.get("custom_static_vars") or {}).items(): result[k]=v

    if profile.get("copy_enabled",False):
        first=str(row.get("first_name","there")).strip()
        company=str(row.get("company_name","")).strip()
        cta=profile.get("copy_cta","worth 15 min this week?")
        sender=profile.get("sender_name","")
        user_p=(f"Prospect: {first} @ {company}\n"
                f"Enrichment: {json.dumps(enr,indent=2)}\n\n"
                f"CTA (append after each email body, NOT inside body): {cta}")
        cp_text,cp_in,cp_out=ai_call(cp_key,profile["copy_system_prompt"],user_p,
                                      max_tokens=profile.get("copy_max_tokens",600),
                                      temperature=profile.get("copy_temperature",1.0),
                                      client_slug=_slug)
        cost+=ai_cost(cp_key,cp_in,cp_out); cp=parse_json(cp_text)
        for f in profile.get("copy_output_fields",[]): result[f]=cp.get(f,cp.get(f.replace(".","_"),""))
        # Assemble Instantly line columns
        for prefix in ["email1","email2"]:
            body=cp.get(f"{prefix}_body","") or cp.get(f"{prefix}.body","")
            subj=cp.get(f"{prefix}_subject","") or cp.get(f"{prefix}.subject","")
            if body:
                sents=[s.strip() for s in re.split(r'(?<=[.!?])\s+',body.strip()) if s.strip()]
                result[f"{prefix}.subject"]=subj
                result[f"{prefix}.line1"]=f"{first} -" if first else ""
                result[f"{prefix}.line2"]=sents[0] if sents else ""
                result[f"{prefix}.line3"]=" ".join(sents[1:]) if len(sents)>1 else ""
                result[f"{prefix}.cta"]=cta
                if sender: result[f"{prefix}.senderName"]=sender
                result.pop(f"{prefix}_body",None); result.pop(f"{prefix}_subject",None)
    return result, cost

def _enrich_batch(rows:List[dict],profile:dict,workers:int,
                  done_set:set,saved_rows:dict,fp:str,st:dict,all_rows:List[dict]
                  )->Tuple[List[dict],float]:
    results=[None]*len(rows); total_cost=0.0
    def do(idx,row):
        try:   return idx,*_enrich_one(row,profile),None
        except Exception as e: return idx,{},0.0,str(e)
    with Progress(SpinnerColumn(),TextColumn("[cyan]{task.description}"),
                  MofNCompleteColumn(),BarColumn(),TimeElapsedColumn(),console=console) as prog:
        task=prog.add_task("Enriching…",total=len(rows))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures={pool.submit(do,i,r):i for i,r in enumerate(rows)}
            for fut in as_completed(futures):
                idx,res,cost,err=fut.result()
                with _enrich_lock:
                    results[idx]=res; total_cost+=cost
                    if err: console.print(f"\n[red]  Row {idx} error:[/] {err[:80]}")
                    try: real=all_rows.index(rows[idx])
                    except: real=idx
                    saved_rows[str(real)]=res; done_set.add(real)
                    st["enrich"]={"done_indices":list(done_set),"enriched_rows":saved_rows,"total_cost":total_cost}
                    state_save(fp,st)
                prog.advance(task)
    return [r or {} for r in results], total_cost

def _apply_saved_enrich(rows:List[dict],saved_rows:dict)->List[dict]:
    return [{**rows[i],**saved_rows.get(str(i),{})} for i in range(len(rows))]

def _show_enrich_sample(results:List[dict],profile:dict):
    fields=profile.get("enrichment_output_fields",[])[:6]
    if not fields or not results: return
    t=Table(title="[bold]Enrichment Sample (first 5 rows)[/]",box=box.ROUNDED,show_lines=True)
    for f in fields: t.add_column(f[:22],style="cyan",max_width=30)
    for r in results[:5]: t.add_row(*[str(r.get(f,""))[:28] for f in fields])
    console.print(t)

def stage_enrich(input_path:Path,rows:List[dict],col_map:Dict[str,Optional[str]],
                 fp:str,st:dict,profile:Optional[dict]=None,
                 force_reenrich:bool=False)->Tuple[List[dict],float,Path]:
    """
    AI enrichment stage. Returns (enriched_rows, total_cost, output_path).
    If profile is None, prompts user to pick one.
    """
    _ensure_default_profiles()
    if profile is None:
        all_p=profiles_all()
        if not all_p:
            console.print("[red]No enrichment profiles found. Run 'mailclaw profiles' to create one.[/]"); return rows,0.0,input_path
        p_name=questionary.select("Select enrichment profile:",choices=list(all_p.keys()),style=Q_STYLE).ask()
        if not p_name: return rows, 0.0, input_path
        profile=all_p[p_name]

    em_key=profile.get("enrichment_model","gemini/gemini-2.5-flash-lite")
    cp_key=profile.get("copy_model","")
    workers=profile.get("workers",6); n=len(rows)
    TEST_N=min(5,n)

    em_est=ai_cost_est(em_key,n,avg_in=profile.get("enrichment_max_tokens",400)//2,
                                  avg_out=profile.get("enrichment_max_tokens",400)//3)
    cp_est=ai_cost_est(cp_key,n) if (profile.get("copy_enabled") and cp_key) else 0.0
    total_est=em_est+cp_est

    console.print(Panel(
        f"[bold]Profile:[/]      {profile.get('display_name',profile['name'])}\n"
        f"[bold]Enrich model:[/] {em_key}  [dim](~${em_est:.4f})[/]\n"
        f"[bold]Copy model:[/]   {cp_key if profile.get('copy_enabled') else 'disabled'}  [dim](~${cp_est:.4f})[/]\n"
        f"[bold]Rows:[/]         {n:,}   [bold]Workers:[/] {workers}\n"
        f"[bold yellow]Estimated total:[/]  ~${total_est:.4f}  (~${total_est/max(n,1):.6f}/row)",
        title="[bold]Enrichment Plan[/]",border_style="yellow"))

    # Remap rows to standard names for enrichment input
    std_rows=[]
    for row in rows:
        sr=dict(row)
        for field,src_col in col_map.items():
            if src_col and src_col in row: sr[field]=row[src_col]
        std_rows.append(sr)

    done_set=set(st.get("enrich",{}).get("done_indices",[]))
    saved_rows=st.get("enrich",{}).get("enriched_rows",{})
    total_cost=st.get("enrich",{}).get("total_cost",0.0)

    if done_set and not force_reenrich:
        console.print(f"[yellow]{len(done_set)}/{n} rows already enriched from previous run.[/]")
        resume=questionary.select("Resume options:",choices=[
            "Resume from where we left off",
            "Force re-enrich all rows (ignores previous results)",
            "Use previous results — skip enrichment",
        ],style=Q_STYLE).ask()
        if "skip" in (resume or "").lower():
            return _apply_saved_enrich(std_rows,saved_rows),total_cost,input_path
        if "Force" in (resume or ""):
            done_set=set(); saved_rows={}; total_cost=0.0

    todo=[r for i,r in enumerate(std_rows) if i not in done_set]
    if not todo:
        console.print("[green]All rows already enriched — skipping.[/]")
        return _apply_saved_enrich(std_rows,saved_rows),total_cost,input_path

    # Test batch first
    if not done_set:
        do_test=questionary.confirm(
            f"Run test batch of {TEST_N} rows first before committing to full run?",
            default=True,style=Q_STYLE).ask()
        if do_test:
            console.print(f"\n[cyan]→[/] Test enrichment: {TEST_N} rows…")
            test_res,test_cost=_enrich_batch(todo[:TEST_N],profile,workers,
                                              done_set,saved_rows,fp,st,std_rows)
            total_cost+=test_cost
            _show_enrich_sample(test_res,profile)
            console.print(f"\n[bold]Test cost:[/] ${test_cost:.4f}  |  [bold]Projected full run:[/] ~${total_est:.4f}")
            go=questionary.confirm(
                f"Output look good? Continue with remaining {max(0,len(todo)-TEST_N):,} rows?",
                default=True,style=Q_STYLE).ask()
            if not go:
                console.print("[yellow]Stopped after test. Re-run anytime to continue.[/]")
                enriched=_apply_saved_enrich(std_rows,saved_rows)
                out_p=input_path.parent/out_name(input_path.stem,"ai_enriched_PARTIAL")
                csv_write(enriched,out_p); return enriched,total_cost,out_p
            todo=todo[TEST_N:]

    if todo:
        console.print(f"\n[cyan]→[/] Enriching {len(todo):,} remaining rows with {workers} workers…")
        _,run_cost=_enrich_batch(todo,profile,workers,done_set,saved_rows,fp,st,std_rows)
        total_cost+=run_cost
        st["enrich"]["completed_at"]=datetime.utcnow().isoformat()
        state_save(fp,st)

    console.print(f"\n[green]✓[/] Enrichment complete — [bold]${total_cost:.4f}[/] total")
    enriched=_apply_saved_enrich(std_rows,saved_rows)
    out_p=input_path.parent/out_name(input_path.stem,"ai_enriched")
    csv_write(enriched,out_p)
    return enriched,total_cost,out_p

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  VERIFICATION STAGE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _process_reoon_results(raw:dict)->dict:
    """
    Splits Reoon results into safe/catchall/dropped and groups by ESP.
    Region bucketing is intentionally NOT done here because Reoon never
    returns a location field — region is detected after merging original
    CSV columns back in (see stage_verify → merge_back).
    ESP: uses cached 'esp' value if mx_records is empty/missing (happens
    when results come from history cache rather than a fresh Reoon call).
    """
    safe,catchall,dropped,by_esp=[],[],[],{}
    for em,data in raw.items():
        status=data.get("status","unknown")
        mx=data.get("mx_records",[])
        # Trust the cached esp if mx_records is empty — it was detected correctly
        # on the original Reoon call. Only re-detect if we have real mx_records.
        if mx and mx != []:
            esp=detect_esp(mx)
        else:
            esp=data.get("esp","other")  # use cached value, fall back to "other"
        row={**data,"esp":esp,"email":em}
        if status in ("safe","role_account"):
            safe.append(row); by_esp.setdefault(esp,[]).append(row)
        elif status=="catch_all": catchall.append(row)
        else: dropped.append(row)
    return {"safe":safe,"catchall":catchall,"dropped":dropped,"by_esp":by_esp}

def stage_verify(input_path:Path,rows:List[dict],email_col:str,
                 fp:str,st:dict,col_map:Optional[Dict[str,Optional[str]]]=None)->Optional[dict]:
    """
    Email verification stage.
    Returns dict with processed results + output paths, or None on failure.
    Saves state per Reoon task so we can resume if credits run out mid-run.
    """
    c=cfg_load()
    rotator=ReoonRotator(c["reoon_keys"],c.get("daily_limit",2000))
    # Always sync from live before verifying — credits may have been used
    # via the dashboard or other tools since last run.
    console.print("[dim]Syncing Reoon credits from live API…[/]")
    rotator.sync_from_live(silent=True)
    hist=hist_load()
    rev_days=c.get("reverify_days",7)

    # Gather emails
    emails_all=[r.get(email_col,"").strip() for r in rows if r.get(email_col,"").strip()]
    if not emails_all:
        console.print("[red]No emails found in column.[/]"); return None

    # Check cache
    already_ok={}; needs_verify=[]
    for em in emails_all:
        d=days_since_verified(em,hist)
        if d is not None and d<rev_days:
            rec=hist[em]; already_ok[em]={"email":em,"status":rec.get("verification_status","safe"),"esp":rec.get("esp","other"),"mx_records":[]}
        else: needs_verify.append(em)

    if already_ok:
        console.print(Panel(
            f"[green]{len(already_ok):,}[/] emails cached  (verified ≤{rev_days} days ago)\n"
            f"[cyan]{len(needs_verify):,}[/] need fresh verification",
            title="[bold]Verification Cache[/]",border_style="cyan"))
        if needs_verify:
            ch=questionary.select("How to proceed?",choices=[
                f"Use cache + verify {len(needs_verify):,} new ones  ← recommended",
                "Re-verify EVERY email (ignores cache, uses more credits)",
                "Use cache only — skip Reoon entirely",
            ],style=Q_STYLE).ask()
            if "cache only" in (ch or "").lower():    needs_verify=[]
            elif "Re-verify EVERY" in (ch or ""):     needs_verify=emails_all; already_ok={}

    # Check if rotator has any keys
    if needs_verify and not rotator.keys:
        console.print("[red]No Reoon keys configured. Run [bold]mailclaw config[/] to add one.[/]")
        if questionary.confirm("Add a Reoon key now?",default=True,style=Q_STYLE).ask():
            name=questionary.text("Key name (e.g. reoon-1):",style=Q_STYLE).ask()
            key =questionary.text("API key:",style=Q_STYLE).ask()
            if name and key:
                c["reoon_keys"].append({"name":name.strip(),"key":key.strip(),"used_today":0,"last_reset":""})
                cfg_save(c)
                rotator=ReoonRotator(c["reoon_keys"],c.get("daily_limit",2000))
                rotator.sync_from_live(silent=True)

    # Show live-synced balance before submitting
    if needs_verify:
        console.print(rotator.status_table())
        console.print(f"[bold]Total available credits:[/] {rotator.total_remaining():,}  |  "
                      f"[bold]Emails to verify:[/] {len(needs_verify):,}")
        if rotator.total_remaining()==0:
            console.print("[red]All Reoon keys exhausted for today.[/]")
            add_key=questionary.confirm("Add a new Reoon key to continue?",default=True,style=Q_STYLE).ask()
            if add_key:
                name=questionary.text("Key name:",style=Q_STYLE).ask()
                key =questionary.text("API key:",style=Q_STYLE).ask()
                if name and key:
                    c["reoon_keys"].append({"name":name.strip(),"key":key.strip(),"used_today":0,"last_reset":""})
                    cfg_save(c)
                    rotator=ReoonRotator(c["reoon_keys"],c.get("daily_limit",2000))
                    rotator.sync_from_live(silent=True)
            if not rotator.available():
                console.print("[yellow]No credits available — using cache only.[/]"); needs_verify=[]

    # Resume in-flight tasks
    v_st=st.get("verify",{}); in_flight=v_st.get("in_flight_tasks",[]); saved_res=v_st.get("partial_results",{})
    if in_flight:
        console.print(f"[yellow]{len(in_flight)} in-flight Reoon task(s) from previous run.[/]")
        if not questionary.confirm("Resume polling those tasks?",default=True,style=Q_STYLE).ask():
            in_flight=[]; saved_res={}

    all_results=dict(saved_res)

    # Confirm before submitting
    if needs_verify:
        if not questionary.confirm(
            f"Submit {len(needs_verify):,} emails to Reoon for verification?\n"
            f"  [dim]Using {len(rotator.available())} key(s), ~{rotator.total_remaining()} credits available[/]",
            default=True,style=Q_STYLE).ask():
            needs_verify=[]

    # Submit batches
    if needs_verify:
        rem=list(needs_verify); bn=len(in_flight)
        while rem and rotator.available():
            key=rotator.available()[0]; cap=min(rotator.remaining(key),c.get("batch_size",5000))
            batch=rem[:cap]; rem=rem[cap:]; bn+=1
            console.print(f"\n[cyan]→[/] Batch #{bn}: {len(batch):,} emails via [bold]{key['name']}[/] ({rotator.remaining(key)-len(batch):,} credits remaining on key)…")
            tid,err=reoon_submit_bulk(batch,key["key"],f"Mailclaw-{bn}")
            if not tid: console.print(f"[red]✗ Submit failed:[/] {err}"); continue
            rotator.record(key["name"],len(batch))
            in_flight.append({"key_name":key["name"],"key":key["key"],"task_id":tid,"count":len(batch)})
            st["verify"]={"status":"running","in_flight_tasks":in_flight,"partial_results":all_results,"submitted_at":datetime.utcnow().isoformat()}
            state_save(fp,st)
            console.print(f"[green]✓[/] Task {tid} submitted — state saved (resume-safe).")
        if rem:
            console.print(f"\n[yellow]⚠  {len(rem):,} emails could not be submitted (credits exhausted).[/]")
            console.print("[dim]Run [bold]mailclaw verify[/] again after credits reset to pick up where you left off.[/]")

    # Poll all in-flight tasks
    if in_flight:
        console.print(f"\n[cyan]Polling {len(in_flight)} Reoon task(s)…[/]")
        done_tids=[]
        for ti in in_flight:
            tid=ti["task_id"]
            api_key=ti.get("key") or next((k["key"] for k in c["reoon_keys"] if k["name"]==ti["key_name"]),"")
            with Progress(SpinnerColumn(),TextColumn(f"[cyan]Task {tid}[/] {{task.description}}"),
                          BarColumn(),TaskProgressColumn(),TimeElapsedColumn(),console=console) as prog:
                pt=prog.add_task("Waiting…",total=100); polls=0
                while True:
                    time.sleep(10); polls+=1
                    data=reoon_poll(tid,api_key)
                    if not data:
                        if polls>60: prog.update(pt,description="[red]Timed out[/]"); break
                        continue
                    s_=data.get("status","waiting"); pct=data.get("progress_percentage",0)
                    prog.update(pt,completed=pct,
                                description=f"{s_} ({data.get('count_checked',0)}/{data.get('count_total',0)})")
                    if s_=="completed":
                        prog.update(pt,completed=100,description="[green]✓ Complete[/]")
                        all_results.update(data.get("results",{})); done_tids.append(tid)
                        st["verify"]["partial_results"]=all_results; state_save(fp,st); break
                    if s_ not in ("waiting","running"):
                        prog.update(pt,description=f"[red]{s_}[/]"); break

        st["verify"]["in_flight_tasks"]=[t for t in in_flight if t["task_id"] not in done_tids]
        st["verify"]["status"]="completed" if not st["verify"]["in_flight_tasks"] else "partial"
        st["verify"]["completed_at"]=datetime.utcnow().isoformat()
        state_save(fp,st)

    all_results.update(already_ok)
    if not all_results: console.print("[red]No verification results.[/]"); return None

    hist_mark_verified(all_results)
    processed=_process_reoon_results(all_results)
    hist2=hist_load()
    safe_rows=processed["safe"]; catchall_rows=processed["catchall"]

    # Merge original CSV columns back into every verified row.
    # Region is detected HERE from the original location column — Reoon
    # never returns location so it must come from the original data.
    orig_map={r.get(email_col,"").lower().strip():r for r in rows}
    loc_col = col_map.get("location") if col_map else None   # original location col name

    def merge_back(ver_rows:List[dict])->List[dict]:
        out=[]
        for v in ver_rows:
            em=v.get("email","").lower().strip()
            orig=orig_map.get(em,{})
            rec=hist2.get(em,{})
            merged={**orig,**v,
                "lp_last_verified":rec.get("last_verified",""),
                "lp_last_uploaded":rec.get("last_uploaded",""),
                "lp_last_campaign":rec.get("last_campaign","")}
            # Detect region from original location column (Reoon doesn't return this)
            loc_val=(orig.get(loc_col,"") if loc_col else "") or orig.get("location","") or orig.get("Location","")
            merged["region"]=detect_region(loc_val)
            out.append(merged)
        return out

    safe_merged=merge_back(safe_rows)
    catchall_merged=merge_back(catchall_rows)

    # Re-bucket safe leads by region now that we have correct region values
    by_region:Dict[str,list]={"america":[],"europe":[]}
    for row in safe_merged:
        by_region.setdefault(row["region"],[]).append(row)

    # Write outputs
    out_dir=input_path.parent
    outfiles={}
    if safe_merged:
        p=out_dir/out_name(input_path.stem,"email_verified_safe")
        csv_write(safe_merged,p); outfiles["safe_all"]=p
    if catchall_merged:
        p=out_dir/out_name(input_path.stem,"email_verified_catchall")
        csv_write(catchall_merged,p); outfiles["catchall"]=p
    # ESP-segmented files
    for esp,esp_rows in processed["by_esp"].items():
        m=merge_back(esp_rows)
        if m:
            p=out_dir/out_name(input_path.stem,f"safe_{esp}")
            csv_write(m,p); outfiles[f"safe_{esp}"]=p
    # Region-segmented files — use by_region computed after merge (correct location data)
    for region,region_rows in by_region.items():
        if region_rows:
            p=out_dir/out_name(input_path.stem,f"safe_{region}")
            csv_write(region_rows,p); outfiles[f"region_{region}"]=p

    st.setdefault("verify", {})
    st["verify"]["results_dir"]=str(out_dir)
    st["verify"]["counts"]={"safe":len(safe_rows),"catchall":len(catchall_rows),"dropped":len(processed["dropped"]),**{f"esp_{k}":len(v) for k,v in processed["by_esp"].items()}}
    state_save(fp,st)

    return {"processed":processed,"outfiles":outfiles,"safe_merged":safe_merged,"catchall_merged":catchall_merged}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  INSTANTLY V2 CLIENT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class Instantly:
    """
    Instantly V2 API client.
    Base URL:   https://api.instantly.ai/api/v2
    Auth:       Authorization: Bearer {key}
    Rate limit: 100 req/10s, 600 req/min (shared across the whole workspace).
                At delay=0.12s we run ~8 req/s — safely under the limit.
    All errors are logged at WARNING level. All requests at DEBUG level.
    """
    def __init__(self,key:str,name:str="",delay:float=0.12):
        self.key=key; self.name=name; self.delay=delay
        self.hdrs={"Authorization":f"Bearer {key}","Content-Type":"application/json"}
        log.debug("Instantly client init: name=%r delay=%.3fs", name, delay)

    # ── Internal HTTP helpers ────────────────────────────────────────────────
    def _g(self,p,params=None):
        log.debug("Instantly GET  %s params=%s", p, params)
        return _req("GET",   f"{INSTANTLY_BASE}{p}",self.delay,headers=self.hdrs,params=params)
    def _po(self,p,data=None):
        log.debug("Instantly POST %s body_keys=%s", p, list((data or {}).keys()))
        return _req("POST",  f"{INSTANTLY_BASE}{p}",self.delay,headers=self.hdrs,json=data)
    def _pa(self,p,data=None):
        log.debug("Instantly PATCH %s body_keys=%s", p, list((data or {}).keys()))
        return _req("PATCH", f"{INSTANTLY_BASE}{p}",self.delay,headers=self.hdrs,json=data)
    def _de(self,p,data=None):
        log.debug("Instantly DELETE %s body=%s", p, data)
        return _req("DELETE",f"{INSTANTLY_BASE}{p}",self.delay,headers=self.hdrs,json=data)

    # ── Campaigns ────────────────────────────────────────────────────────────
    def list_campaigns(self,limit=99,cursor=None)->Tuple[List[dict],Optional[str],str]:
        """GET /api/v2/campaigns — paginated. Returns (items, next_cursor, error)."""
        params={"limit":limit}
        if cursor: params["starting_after"]=cursor
        r=self._g("/campaigns",params); d=r.json()
        if not r.ok:
            log.warning("list_campaigns [%d]: %s", r.status_code, d)
            return [],[],d.get("message",f"HTTP {r.status_code}")
        items=d if isinstance(d,list) else d.get("items",[])
        next_c=d.get("next_starting_after") if isinstance(d,dict) else None
        log.debug("list_campaigns: %d items next=%s", len(items), next_c)
        return items,next_c,""

    def list_all_campaigns(self)->List[dict]:
        """Fetch every campaign, paginating through all pages."""
        all_=[]; cursor=None
        while True:
            batch,cursor,err=self.list_campaigns(limit=99,cursor=cursor)
            if err: log.warning("list_all_campaigns error: %s",err); break
            all_.extend(batch)
            if not cursor or len(batch)<99: break
        return all_

    def get_analytics_overview(self,campaign_ids:List[str],
                               start_date:str="",end_date:str="")->Optional[dict]:
        """
        GET /api/v2/campaigns/analytics/overview
        Params: ids[] (multiple), start_date, end_date (YYYY-MM-DD optional)
        Returns overview object with: emails_sent_count, contacted_count,
          bounced_count, unsubscribed_count, open_count_unique, reply_count_unique,
          total_interested, total_meeting_booked, total_meeting_completed,
          total_closed, completed_count.
        Pass expand_crm_events=true to count all status changes (not just first per contact).
        """
        params={}
        # API accepts ?ids=x&ids=y for multiple campaigns
        if len(campaign_ids)==1: params["id"]=campaign_ids[0]
        if start_date: params["start_date"]=start_date
        if end_date:   params["end_date"]=end_date
        # Do NOT use expand_crm_events=true — it counts every status transition,
        # not unique leads per status. Default (false) = first occurrence per contact,
        # which gives the correct funnel counts.
        if len(campaign_ids)>1:
            from urllib.parse import urlencode
            qs="&".join(f"ids={cid}" for cid in campaign_ids)
            if start_date: qs+=f"&start_date={start_date}"
            if end_date:   qs+=f"&end_date={end_date}"
            r=_req("GET",f"{INSTANTLY_BASE}/campaigns/analytics/overview?{qs}",
                   self.delay,headers=self.hdrs)
        else:
            r=self._g("/campaigns/analytics/overview",params)
        if not r.ok:
            log.warning("get_analytics_overview [%d]: %s",r.status_code,r.text[:200])
            return None
        return r.json()

    def get_analytics_steps(self,campaign_id:str,
                            start_date:str="",end_date:str="")->List[dict]:
        """
        GET /api/v2/campaigns/analytics/steps
        Params: campaign_id (one at a time), start_date, end_date.
        Returns per-step array: [{step, variant, sent, unique_opened,
          unique_replies, replies_automatic, unique_clicks}]
        NOTE: No bounce/positive_replies/unsubscribed per step in this API.
        """
        params={"campaign_id":campaign_id,"include_opportunities_count":"true"}
        if start_date: params["start_date"]=start_date
        if end_date:   params["end_date"]=end_date
        r=self._g("/campaigns/analytics/steps",params)
        if not r.ok:
            log.warning("get_analytics_steps [%d]: %s",r.status_code,r.text[:200])
            return []
        d=r.json()
        return d if isinstance(d,list) else []

    def duplicate_campaign(self,cid:str,new_name:str="")->Tuple[Optional[dict],str]:
        """
        POST /api/v2/campaigns/{id}/duplicate

        Instantly V2 may return either:
        a) A campaign object directly  → has "id", no "job_id"  (fast path)
        b) A background job object     → has "job_id"           (async path)

        For (b) we poll GET /api/v2/background-jobs/{job_id} until status is
        "completed", then fetch the new campaign by the job's output campaign_id.
        Times out after ~30s (15 polls × 2s). Returns (campaign_dict, error).
        """
        log.debug("duplicate_campaign cid=%s new_name=%r", cid[:8], new_name)
        r=self._po(f"/campaigns/{cid}/duplicate", data={})  # Instantly requires {} not null body
        if not r.ok:
            d=r.json() if r.content else {}
            log.warning("duplicate_campaign [%d]: %s", r.status_code, d)
            return None, d.get("message",f"HTTP {r.status_code}")

        d=r.json() if r.content else {}
        log.debug("duplicate_campaign response keys: %s", list(d.keys()))

        # Fast path: got campaign object directly
        if d.get("id") and not d.get("job_id"):
            camp=d
            if new_name:
                self.update_campaign(camp["id"],{"name":new_name})
                camp["name"]=new_name
            return camp, ""

        # Async path: got a background job — poll until complete
        job_id=d.get("job_id") or d.get("id")
        if not job_id:
            return None, f"Unexpected duplicate response: {list(d.keys())}"

        log.debug("duplicate_campaign: background job %s — polling…", job_id)
        console.print(f"[dim]  Campaign duplication queued (job {job_id}) — polling…[/]")

        for attempt in range(20):  # max ~40s
            time.sleep(2)
            try:
                jr=self._g(f"/background-jobs/{job_id}")
                if not jr.ok:
                    log.debug("background-job poll [%d]", jr.status_code)
                    continue
                jd=jr.json()
                log.debug("background-job status=%s keys=%s", jd.get("status"), list(jd.keys()))
                status=jd.get("status","")

                if status=="completed":
                    # The job output contains the new campaign id
                    new_cid=(jd.get("output",{}) or {}).get("campaign_id") or jd.get("campaign_id")
                    if new_cid:
                        # Fetch the full campaign object
                        cr=self._g(f"/campaigns/{new_cid}")
                        if cr.ok:
                            camp=cr.json()
                            if new_name:
                                self.update_campaign(new_cid,{"name":new_name})
                                camp["name"]=new_name
                            log.debug("duplicate_campaign complete: new cid=%s", new_cid[:8])
                            return camp, ""
                    # Completed but no campaign_id in output — try campaigns list to find newest
                    log.warning("background job completed but no campaign_id in output: %s", jd)
                    return None, "Duplication job completed but could not retrieve new campaign ID"

                if status in ("failed","error","cancelled"):
                    return None, f"Duplication job {status}: {jd.get('error','unknown error')}"

            except Exception as e:
                log.debug("background-job poll error (attempt %d): %s", attempt, e)
                continue

        return None, "Duplication job timed out after 40s"

    def update_campaign(self,cid,payload)->Tuple[bool,str]:
        """PATCH /api/v2/campaigns/{id} — can update name, campaign_schedule, email_list, etc."""
        r=self._pa(f"/campaigns/{cid}",payload)
        log.debug("update_campaign %s [%d]", cid[:8], r.status_code)
        return (True,"") if r.ok else (False,r.json().get("message",f"HTTP {r.status_code}"))

    # ── Leads ────────────────────────────────────────────────────────────────
    def bulk_add_leads(self,cid:str,lead_objects:List[dict],
                       skip_if_in_workspace:bool=False,
                       skip_if_in_campaign:bool=False)->dict:
        """
        POST /api/v2/leads/add  — THE correct bulk endpoint (up to 1000 leads per call).
        NOT /api/v2/leads which is a different single-lead endpoint that stores leads
        differently and does NOT show them in the campaign UI.

        Body:
          campaign_id:          uuid
          leads:                array of lead objects (max 1000)
            each lead:          {email, first_name, last_name, company_name, website,
                                 phone, personalization, custom_variables: {k:v}}
          skip_if_in_workspace: bool
          skip_if_in_campaign:  bool

        Response:
          status, total_sent, leads_uploaded, in_blocklist, duplicated_leads,
          skipped_count, invalid_email_count, remaining_in_plan, created_leads[]
        """
        body={
            "campaign_id":          cid,
            "leads":                lead_objects,
            "skip_if_in_workspace": skip_if_in_workspace,
            "skip_if_in_campaign":  skip_if_in_campaign,
        }
        log.debug("bulk_add_leads: cid=%s n=%d skip_workspace=%s skip_campaign=%s",
                  cid[:8], len(lead_objects), skip_if_in_workspace, skip_if_in_campaign)
        r=self._po("/leads/add", data=body)
        d=r.json() if r.content else {}
        log.debug("bulk_add_leads response [%d]: %s", r.status_code, str(d)[:400])
        if not r.ok:
            log.warning("bulk_add_leads FAIL [%d]: %s", r.status_code, d)
        return d if r.ok else {"error": d.get("message", f"HTTP {r.status_code}"), "leads_uploaded":0}

    def list_leads(self,cid,limit=100,cursor=None)->Tuple[List[dict],Optional[str],str]:
        """
        POST /api/v2/leads/list  (Instantly uses POST for complex filtering — their design).
        Paginate with next_starting_after cursor.
        """
        payload={"campaign_id":cid,"limit":limit}
        if cursor: payload["starting_after"]=cursor
        r=self._po("/leads/list",payload); d=r.json()
        if not r.ok:
            log.warning("list_leads [%d]: %s", r.status_code, d)
            return [],[],d.get("message",f"HTTP {r.status_code}")
        items=d if isinstance(d,list) else d.get("items",[])
        next_c=d.get("next_starting_after") if isinstance(d,dict) else None
        log.debug("list_leads cid=%s: %d items next=%s", cid[:8], len(items), next_c)
        return items,next_c,""

    def list_positive_leads(self, cid:str=None, max_expected:int=500)->List[dict]:
        """
        Fetch positive-status leads (Interested/Booked/Completed/Won).

        CRITICAL: lt_interest_status is workspace-level, NOT per-campaign.
        Fetching with campaign_id returns the same 19 leads for every campaign.
        Correct approach: fetch once workspace-wide (cid=None), then cross-reference
        via the campaign_id field on each lead object.

        If cid is provided, it's passed only to narrow the baseline total comparison.
        """
        STATUS_POSITIVE={1,2,3,4}
        out=[]

        # Baseline: workspace-wide unfiltered total
        r0=self._po("/leads/list",{"limit":1})
        full_total=None
        if r0.ok:
            d0=r0.json()
            full_total=(d0.get("total") or d0.get("count")) if isinstance(d0,dict) else None

        # Probe filter syntaxes (workspace-wide, no campaign_id)
        best_filter=None; best_total=None
        for f_attempt in [
            {"lt_interest_status": list(STATUS_POSITIVE)},
            {"interest_status":    list(STATUS_POSITIVE)},
        ]:
            probe={**{"limit":1}, **f_attempt}
            r=self._po("/leads/list",probe)
            if not r.ok: continue
            d=r.json()
            total=(d.get("total") or d.get("count")) if isinstance(d,dict) else None
            log.debug("list_positive_leads probe filter=%s total=%s full=%s",
                      list(f_attempt.keys()),total,full_total)
            if total is not None:
                is_filtering=(full_total is None or total < full_total*0.9)
                if is_filtering and (best_total is None or total < best_total):
                    best_total=total; best_filter=f_attempt
            elif best_filter is None:
                best_filter=f_attempt

        if best_filter is None:
            log.warning("list_positive_leads: no usable filter"); return []

        if best_total is not None and best_total > max(max_expected*3, 500):
            log.warning("list_positive_leads: total=%d >> expected=%d — skipping",
                        best_total, max_expected)
            return []

        # Paginate workspace-wide with best filter
        cursor=None
        while True:
            payload={**{"limit":100}, **best_filter}
            if cursor: payload["starting_after"]=cursor
            r=self._po("/leads/list",payload); d=r.json()
            if not r.ok: break
            items=d if isinstance(d,list) else d.get("items",[])
            for lead in items:
                if lead.get("lt_interest_status") in STATUS_POSITIVE:
                    out.append(lead)
            cursor=d.get("next_starting_after") if isinstance(d,dict) else None
            if not cursor or len(items)<100: break

        log.debug("list_positive_leads: %d workspace-level positive leads",len(out))
        return out

    def get_first_reply(self,cid:str,lead_email:str)->Optional[dict]:
        """
        GET /api/v2/emails — rate limited to 20 req/min (3s between calls).
        Fetches the FIRST inbound email from a lead in a campaign.
        Returns dict with: timestamp_email, is_auto_reply, subject, or None.

        Using sort_order=asc + limit=1 + email_type=received gives us the
        first reply timestamp — this is the anchor date for week-over-week
        reporting so repeat replies from the same lead don't inflate later weeks.
        """
        params={
            "campaign_id":   cid,
            "lead":          lead_email,
            "email_type":    "received",
            "sort_order":    "asc",
            "limit":         1,
        }
        # 20 req/min rate limit — enforce 3.1s gap between calls
        time.sleep(3.1)
        r=self._g("/emails",params)
        if not r.ok:
            log.warning("get_first_reply [%d] cid=%s lead=%s",r.status_code,cid[:8],lead_email)
            return None
        d=r.json()
        items=d.get("items",[]) if isinstance(d,dict) else (d if isinstance(d,list) else [])
        if not items: return None
        email=items[0]
        return {
            "timestamp":    email.get("timestamp_email") or email.get("timestamp_created",""),
            "is_auto_reply":bool(email.get("is_auto_reply")),
            "subject":      email.get("subject",""),
            "content_preview": email.get("content_preview",""),
        }

    def delete_lead(self,lid)->bool:
        """DELETE /api/v2/leads/{id} — single delete by ID in URL. Confirmed in V2 docs."""
        log.debug("delete_lead %s", lid)
        try:
            r=self._de(f"/leads/{lid}")
            log.debug("delete_lead %s -> [%d]", lid, r.status_code)
            return r.ok
        except Exception as e:
            log.warning("delete_lead %s exception: %s", lid, e)
            return False

    def bulk_delete(self,leads)->Tuple[int,int]:
        """Individual deletes — fallback since bulk delete body field name is undocumented."""
        ok_total=fail_total=0
        ids=[l.get("id") or l.get("lead_id","") for l in leads]
        ids=[i for i in ids if i]
        if not ids: return 0, len(leads)
        log.debug("bulk_delete: %d leads", len(ids))
        with Progress(SpinnerColumn(),TextColumn("[cyan]Deleting leads…"),
                      MofNCompleteColumn(),BarColumn(),console=console,transient=True) as prog:
            task=prog.add_task("",total=len(ids))
            for lid in ids:
                if self.delete_lead(lid): ok_total+=1
                else: fail_total+=1
                prog.advance(task)
        log.debug("bulk_delete done: ok=%d fail=%d", ok_total, fail_total)
        return ok_total, fail_total

    def is_quota_err(self,msg:str)->bool:
        return any(k in str(msg).lower() for k in ["quota","limit","max lead","maximum lead","capacity","remaining_in_plan"])

    def fetch_purgeable_leads(self,cid)->List[dict]:
        out=[]; cursor=None
        while True:
            batch,cursor,err=self.list_leads(cid,100,cursor)
            if err: break
            for l in batch:
                if l.get("status") in (2,-1) and not (l.get("reply_to_uuid") or l.get("has_reply")):
                    out.append(l)
            if not cursor or len(batch)<100: break
        return out

    def upload_leads(self,cid:str,rows:List[dict],col_map:Dict[str,Optional[str]],
                     client_extra:dict={},
                     skip_if_in_workspace:bool=False,
                     skip_if_in_campaign:bool=False)->dict:
        """
        Uploads leads using POST /api/v2/leads/add (correct bulk endpoint, up to 1000/call).
        Custom vars go in 'custom_variables' per the API spec (NOT 'payload').
        Batches into groups of 1000 to stay within the endpoint limit.
        Returns unified result dict with success/failed/skipped/blocklist counts.
        """
        results={"success":0,"failed":0,"skipped":0,"in_blocklist":0,"quota_hit":False,"errors":[]}
        MAPPED=set(v for v in col_map.values() if v)
        custom_static=client_extra.get("custom_upload_vars",{})
        log.debug("upload_leads: %d rows cid=%s skip_workspace=%s skip_campaign=%s",
                  len(rows), cid[:8], skip_if_in_workspace, skip_if_in_campaign)

        # Build lead objects
        lead_objects=[]
        for row in rows:
            def g(f):
                c_=col_map.get(f); return str(row.get(c_,"")).strip() if c_ else ""
            # All unmapped columns → custom_variables (correct field name per API spec)
            custom_vars={re.sub(r"[^a-z0-9_]","_",k.lower()):str(v)
                        for k,v in row.items() if k not in MAPPED and v
                        and str(v).strip().lower() not in ("","nan","none")}
            custom_vars.update(custom_static)
            lead={
                "email":            g("email"),
                "first_name":       g("first_name"),
                "last_name":        g("last_name"),
                "company_name":     g("company_name"),
                "website":          g("website"),
                "phone":            g("phone"),
                "personalization":  g("icebreaker"),
            }
            if custom_vars: lead["custom_variables"]=custom_vars
            # Strip empty strings
            lead={k:v for k,v in lead.items() if v or k=="email"}
            lead_objects.append(lead)

        # Batch in groups of 1000 (API max per call)
        BATCH=1000
        with Progress(SpinnerColumn(),TextColumn("[cyan]{task.description}"),
                      MofNCompleteColumn(),BarColumn(),TimeElapsedColumn(),console=console) as prog:
            task=prog.add_task(f"Uploading {len(lead_objects):,} leads…",total=len(lead_objects))

            for i in range(0, len(lead_objects), BATCH):
                batch=lead_objects[i:i+BATCH]
                d=self.bulk_add_leads(cid,batch,skip_if_in_workspace,skip_if_in_campaign)

                if "error" in d:
                    # Whole batch failed
                    results["failed"]+=len(batch)
                    results["errors"].append({"batch":f"{i}-{i+len(batch)}","error":d["error"]})
                    if self.is_quota_err(d["error"]):
                        results["quota_hit"]=True
                        prog.advance(task, len(lead_objects)-i-len(batch))
                        break
                else:
                    uploaded=d.get("leads_uploaded",0)
                    blocklist=d.get("in_blocklist",0)
                    skipped=d.get("skipped_count",0)
                    invalid=d.get("invalid_email_count",0)
                    duplicated=d.get("duplicated_leads",0)
                    results["success"]   +=uploaded
                    results["in_blocklist"]+=blocklist
                    results["skipped"]   +=skipped+duplicated
                    results["failed"]    +=invalid
                    remaining=d.get("remaining_in_plan")
                    if remaining is not None and remaining==0:
                        results["quota_hit"]=True
                    log.debug("batch %d-%d: uploaded=%d blocklist=%d skipped=%d invalid=%d remaining=%s",
                              i, i+len(batch), uploaded, blocklist, skipped, invalid, remaining)

                prog.advance(task, len(batch))
                if i+BATCH < len(lead_objects):
                    time.sleep(0.5)  # respect rate limit between batches

        log.debug("upload_leads done: success=%d failed=%d skipped=%d blocklist=%d quota=%s",
                  results["success"], results["failed"], results["skipped"],
                  results["in_blocklist"], results["quota_hit"])
        return results

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  UPLOAD STAGE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def pick_instantly_client()->Optional[Tuple[dict,dict]]:
    """
    Returns (client_meta, client_extra) or None.
    client_meta = entry from config["instantly_clients"]
    client_extra = entry from ~/.mailclaw/clients/<name>.json
    Also offers to add a new client inline.
    """
    c=cfg_load()
    clients=c["instantly_clients"]
    choices=[f"{cl['name']}  [{cl['key'][:10]}…]" for cl in clients]
    choices+=["➕  Add new client"]
    pick=questionary.select("Select Instantly client:",choices=choices,style=Q_STYLE).ask()
    if not pick: return None
    if "Add new client" in pick:
        name=questionary.text("Client name (e.g. acme):",style=Q_STYLE).ask()
        key =questionary.text("Instantly V2 API key:",style=Q_STYLE).ask()
        if not name or not key: return None
        new_cl={"name":name.strip(),"key":key.strip()}
        c["instantly_clients"].append(new_cl); cfg_save(c)
        console.print(f"[green]✓[/] Added client: {name}")
        return new_cl, client_config_load(name)
    idx=[f"{cl['name']}  [{cl['key'][:10]}…]" for cl in clients].index(pick)
    cl=clients[idx]
    return cl, client_config_load(cl["name"])

def pick_campaign(inst:Instantly)->Optional[dict]:
    campaigns=[]; cursor=None; loaded=0
    while True:
        if loaded==0 or (cursor and questionary.confirm("Load more campaigns?",default=True,style=Q_STYLE).ask()):
            batch,cursor,err=inst.list_campaigns(limit=99,cursor=cursor)
            if err: console.print(f"[red]{err}[/]"); break
            campaigns.extend(batch); loaded+=1
            if not batch: break
        if not campaigns: console.print("[yellow]No campaigns found.[/]"); return None
        # Sort by date descending (latest first); Instantly API has no sort, so sort locally
        def _campaign_sort_key(c):
            ts = c.get("timestamp_created") or c.get("timestamp_updated") or ""
            return (ts or "0000")  # missing at end when descending
        campaigns.sort(key=_campaign_sort_key, reverse=True)
        t=Table(title=f"[bold]Campaigns — {inst.name}[/]",box=box.ROUNDED,show_lines=True)
        t.add_column("#",style="dim",width=4); t.add_column("Name",style="cyan bold",max_width=45)
        t.add_column("Status",style="yellow",width=12); t.add_column("ID",style="dim",max_width=22)
        for i,c_ in enumerate(campaigns,1):
            t.add_row(str(i),c_.get("name","Unnamed"),str(c_.get("status","?")),c_.get("id","?")[:20]+"…")
        console.print(t)
        opts=[f"{c_.get('name','?')}  [{c_.get('id','?')[:10]}]" for c_ in campaigns]
        if cursor: opts.append("── load more ──")
        opts.append("── cancel ──")
        pick=questionary.select("Select campaign:",choices=opts,style=Q_STYLE).ask()
        if not pick or "cancel" in pick: return None
        if "load more" in pick: continue
        idx=opts.index(pick)
        return campaigns[idx]

def stage_upload(input_path:Path,rows:List[dict],col_map:Dict[str,Optional[str]],
                 fp:str,st:dict)->dict:
    """
    Upload stage.
    Default flow (real-life): always duplicate the template campaign, then upload.
    Falls back to uploading directly only if duplication fails AND user confirms.
    """
    c=cfg_load()
    result=pick_instantly_client()
    if not result: return {}
    cl_meta,cl_extra=result
    inst=Instantly(cl_meta["key"],cl_meta["name"],c.get("rate_limit_delay",0.12))

    # Show per-client custom vars if any
    if cl_extra.get("custom_upload_vars"):
        console.print(Panel(
            "\n".join(f"  [cyan]{k}[/] = [green]{v}[/]" for k,v in cl_extra["custom_upload_vars"].items()),
            title=f"[bold]Custom vars for {cl_meta['name']}[/]",border_style="dim"))

    console.print(f"\n[cyan]→[/] Fetching campaigns for [bold]{cl_meta['name']}[/]…")
    template=pick_campaign(inst)
    if not template: return {}

    # ── Duplicate or upload to existing? ────────────────────────
    upload_mode=questionary.select(
        "Do you want to duplicate this campaign or upload to this existing campaign?",
        choices=[
            "Duplicate — create a new campaign from this one (then upload to the new campaign)",
            "Upload to this existing campaign — add leads to the campaign you just selected",
        ],
        style=Q_STYLE
    ).ask()

    if not upload_mode:
        console.print("[yellow]Upload cancelled.[/]"); return {}

    if "Upload to this existing" in (upload_mode or ""):
        if not questionary.confirm(
            f"Add leads to existing campaign [bold]{template.get('name','?')}[/]?",
            default=False,style=Q_STYLE
        ).ask():
            console.print("[yellow]Upload cancelled.[/]"); return {}
        target=template
        cid=target.get("id","")
        is_existing_campaign=True
        console.print(f"[green]✓[/] Using existing campaign [bold]{target.get('name','?')}[/]  [dim]{cid[:16]}…[/]")
    else:
        is_existing_campaign=False
        # ── Duplicate flow: name then duplicate ─────────────────
        date_tag=datetime.now().strftime("%d %b %y")
        csv_stem=re.sub(r"_email_verified.*|_ai_enriched.*|_safe_.*","",input_path.stem).replace("_"," ").strip()
        auto_name=f"{template.get('name','Campaign')} | {csv_stem} | {date_tag}"[:100]

        suggested=questionary.text(
            "Name for the duplicated campaign:",
            default=auto_name,
            style=Q_STYLE
        ).ask()
        new_campaign_name=(suggested or auto_name).strip()

        console.print(f"[cyan]→[/] Duplicating [bold]{template.get('name')}[/]…")
        new_camp,dup_err=inst.duplicate_campaign(template["id"], new_name=new_campaign_name)

        if new_camp:
            target=new_camp
            console.print(f"[green]✓[/] Duplicated → [bold]{target.get('name','?')}[/]  [dim]{target.get('id','?')[:16]}…[/]")
        else:
            console.print(Panel(
                f"[red]Duplication failed:[/] {dup_err}\n\n"
                f"Options:\n"
                f"  1. Upload directly to the selected template campaign (risky — adds leads to your template)\n"
                f"  2. Cancel and fix the issue manually",
                title="[bold red]⚠ Duplicate Failed[/]",border_style="red"))
            action=questionary.select("How to proceed?",choices=[
                "Cancel — do not upload",
                f"Upload directly to template: {template.get('name','?')} (I understand the risk)",
            ],style=Q_STYLE).ask()
            if not action or "Cancel" in action:
                console.print("[yellow]Upload cancelled.[/]"); return {}
            target=template
            console.print(f"[yellow]⚠ Uploading directly to template campaign.[/]")

        cid=target.get("id","")

    # ── Timezone (only for new/duplicated campaigns; don't overwrite existing campaign schedule)
    if not is_existing_campaign:
        loc_col=col_map.get("location")
        region_col_vals=[r.get("region","") for r in rows if r.get("region")]
        if region_col_vals:
            eu_pct=sum(1 for v in region_col_vals if v=="europe")/len(region_col_vals)*100
            def_reg="europe" if eu_pct>=50 else "america"
            console.print(f"[cyan]Region split (from verified data):[/] Europe {eu_pct:.0f}%  /  Americas {100-eu_pct:.0f}%")
        elif loc_col:
            regs=[detect_region(r.get(loc_col,"")) for r in rows]
            eu_pct=regs.count("europe")/max(len(regs),1)*100
            def_reg="europe" if eu_pct>=50 else "america"
            console.print(f"[cyan]Location analysis:[/] Europe {eu_pct:.0f}%  /  Americas {100-eu_pct:.0f}%")
        else:
            def_reg="america"

        tz_pick=questionary.select("Campaign send schedule:",choices=[
            f"{'▶ ' if def_reg=='america' else '  '}🇺🇸  America/Chicago  (8am–5pm CT — covers most US leads)",
            f"{'▶ ' if def_reg=='europe'  else '  '}🇪🇺  Europe/London    (8am–5pm GMT)",
        ],style=Q_STYLE).ask()
        region="europe" if "Europe" in (tz_pick or "") else "america"
        ok_,err=inst.update_campaign(cid,campaign_schedule(region))
        tz_label="🇪🇺 Europe/London" if region=="europe" else "🇺🇸 America/Chicago"
        console.print(f"[green]✓[/] Schedule: {tz_label}" if ok_ else f"[yellow]Schedule warning:[/] {err}")
    else:
        tz_label="(existing — unchanged)"
        console.print("[dim]Keeping existing campaign schedule (no change).[/]")

    # ── Deduplication options ─────────────────────────────────────
    # Ask every time — the right answer depends on the specific upload context.
    # Default: False/False = always add leads (safe for fresh duplicate campaigns).
    # skip_if_in_workspace=True = useful when re-uploading to avoid double-sending.
    # NOTE: Instantly returns HTTP 200 even when silently skipping — no error shown.
    console.print(Panel(
        "[bold]skip_if_in_workspace[/] — skip a lead if it already exists in [bold]any[/] campaign in this workspace\n"
        "[bold]skip_if_in_campaign[/]  — skip a lead if it already exists in [bold]this[/] campaign only\n\n"
        "[dim]For fresh duplicate campaigns → both False (default)\n"
        "Re-uploading to same campaign → skip_if_in_campaign=True\n"
        "Worried about cross-campaign duplication → skip_if_in_workspace=True[/]",
        title="[bold]Deduplication Options[/]",border_style="dim"))

    skip_workspace=questionary.confirm(
        "skip_if_in_workspace?  (skip leads already in ANY campaign in workspace)",
        default=False,style=Q_STYLE).ask() or False
    skip_campaign=questionary.confirm(
        "skip_if_in_campaign?   (skip leads already in THIS campaign)",
        default=False,style=Q_STYLE).ask() or False

    console.print(f"[dim]  skip_workspace={skip_workspace}  skip_campaign={skip_campaign}[/]")

    # ── Column mapping confirmation ───────────────────────────────
    console.print("\n[bold]Column mapping for upload:[/]")
    upload_cols=[c for c in (rows[0].keys() if rows else [])]
    show_mapping_preview(upload_cols, col_map, rows[:2])
    if not questionary.confirm("Confirm this mapping for upload?",default=True,style=Q_STYLE).ask():
        console.print("[yellow]Upload cancelled.[/]"); return {}

    # ── Test batch (3 leads) ──────────────────────────────────────
    TEST_N=min(3,len(rows))
    if questionary.confirm(f"Test upload {TEST_N} leads first?",default=True,style=Q_STYLE).ask():
        console.print(f"[cyan]→[/] Test uploading {TEST_N} leads…")
        test_r=inst.upload_leads(cid,rows[:TEST_N],col_map,cl_extra,
                                 skip_if_in_workspace=skip_workspace,
                                 skip_if_in_campaign=skip_campaign)
        console.print(Panel(
            f"[green]✓ Uploaded:[/]  {test_r['success']}/{TEST_N}\n"
            f"[red]✗ Failed:[/]    {test_r['failed']}\n"
            f"[yellow]⚫ Blocklist:[/] {test_r.get('in_blocklist',0)}\n"
            f"[dim]Skipped:[/]    {test_r.get('skipped',0)}\n"
            +(f"[yellow]Error:[/] {test_r['errors'][0]['error']}" if test_r["errors"] else ""),
            title="[bold]Test Upload Result[/]",border_style="cyan"))
        if test_r["failed"]>0:
            if not questionary.confirm("Some test leads failed. Continue with full upload anyway?",
                                       default=False,style=Q_STYLE).ask():
                return test_r
        remaining_rows=rows[TEST_N:]; success_offset=test_r["success"]
    else:
        remaining_rows=rows; success_offset=0

    # ── Full upload ───────────────────────────────────────────────
    if remaining_rows:
        console.print(f"\n[bold cyan]Uploading {len(remaining_rows):,} leads to [bold]{target.get('name')}[/]…[/]")
        result_=inst.upload_leads(cid,remaining_rows,col_map,cl_extra,
                                   skip_if_in_workspace=skip_workspace,
                                   skip_if_in_campaign=skip_campaign)
        result_["success"]+=success_offset
    else:
        result_={"success":success_offset,"failed":0,"quota_hit":False,"errors":[]}

    # ── Quota handling ────────────────────────────────────────────
    if result_["quota_hit"]:
        console.print(Panel(
            f"[yellow]⚠  Lead quota hit![/]\n\n"
            f"Uploaded [green]{result_['success']:,}[/] before limit.\n"
            f"[dim]{result_['failed']:,}[/] leads were not uploaded.",
            title="[bold yellow]Quota Hit[/]",border_style="yellow"))
        if questionary.confirm("Delete 'completed + no reply' leads to free up space, then re-upload?",
                               default=False,style=Q_STYLE).ask():
            console.print("[cyan]→[/] Scanning for completed/no-reply leads…")
            stale=inst.fetch_purgeable_leads(cid)
            if not stale:
                console.print("[yellow]None found to delete.[/]")
            else:
                console.print(f"[yellow]{len(stale):,} leads eligible for deletion.[/]")
                console.print(f"[dim]Sample: {[l.get('email','?') for l in stale[:5]]}[/]")
                if questionary.confirm(f"DELETE {len(stale):,} leads? (cannot be undone)",
                                       default=False,style=Q_STYLE).ask():
                    ok_d,fail_d=inst.bulk_delete(stale)
                    console.print(f"[green]✓[/] Deleted {ok_d:,}  [dim](failed: {fail_d})[/]")
                    st["cleanup"]={"deleted_at":datetime.utcnow().isoformat(),"count":ok_d,"campaign":cid}
                    state_save(fp,st)
                    retry_em={e["email"] for e in result_["errors"]}
                    retry_rows=[r for r in rows if col_map.get("email") and r.get(col_map["email"],"").strip() in retry_em]
                    if retry_rows:
                        console.print(f"\n[cyan]→[/] Re-uploading {len(retry_rows):,} leads…")
                        r2=inst.upload_leads(cid,retry_rows,col_map,cl_extra,
                                              skip_if_in_workspace=skip_workspace,
                                              skip_if_in_campaign=skip_campaign)
                        result_["success"]+=r2["success"]; result_["failed"]=r2["failed"]

    # ── Save history + state ──────────────────────────────────────
    em_col=col_map.get("email") or ""
    uploaded=[r.get(em_col,"").strip() for r in rows if r.get(em_col,"")]
    hist_mark_uploaded(uploaded,cid,cl_meta["name"])
    st["upload"]={"client":cl_meta["name"],"campaign_id":cid,"campaign":target.get("name","?"),
                  "uploaded_at":datetime.utcnow().isoformat(),"count":result_["success"],"failed":result_["failed"]}
    state_save(fp,st)

    console.print(Panel(
        f"[green]✓ Uploaded:[/]  {result_['success']:,}\n"
        f"[red]✗ Failed:[/]    {result_['failed']:,}\n"
        f"[yellow]⚫ Blocklist:[/] {result_.get('in_blocklist',0):,}\n"
        f"[dim]Skipped:[/]    {result_.get('skipped',0):,}\n"
        f"[cyan]Campaign:[/]   {target.get('name','?')}\n"
        f"[cyan]Client:[/]     {cl_meta['name']}\n"
        f"[cyan]Timezone:[/]   {tz_label}",
        title="[bold green]Upload Complete[/]",border_style="green"))
    return result_

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  STANDALONE COMMAND HANDLERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def cmd_verify(_args):
    """Standalone: email verify only."""
    console.print("\n[bold cyan]Email Verification[/]  [dim](standalone)[/]")
    csv_path=pick_csv()
    if not csv_path: return
    fp=csv_fp(csv_path); st=state_load(fp)
    rows,cols=csv_read(csv_path)
    console.print(f"[green]✓[/] {len(rows):,} rows — [yellow]{csv_path.name}[/]")
    col_map=st.get("column_map") or do_column_mapping(cols,rows)
    st["column_map"]=col_map; state_save(fp,st)
    email_col=col_map.get("email")
    if not email_col: console.print("[red]No email column mapped.[/]"); return
    result=stage_verify(csv_path,rows,email_col,fp,st,col_map)
    if result:
        p=result["processed"]
        console.print(Panel(
            f"[green]✓ Safe:[/]      {len(p['safe']):,}\n"
            f"[yellow]⚡ Catch-all:[/] {len(p['catchall']):,}\n"
            f"[red]✗ Dropped:[/]   {len(p['dropped']):,}\n\n"
            f"[cyan]Gmail:[/] {len(p['by_esp'].get('gmail',[]))}  "
            f"[blue]Outlook:[/] {len(p['by_esp'].get('outlook',[]))}  "
            f"[magenta]Yahoo:[/] {len(p['by_esp'].get('yahoo',[]))}  "
            f"[white]Other:[/] {len(p['by_esp'].get('other',[]))}",
            title="[bold]Verification Results[/]",border_style="cyan"))

def cmd_enrich(_args):
    """Standalone: AI enrichment only."""
    console.print("\n[bold magenta]AI Enrichment[/]  [dim](standalone)[/]")
    c=cfg_load()
    if not any(v for v in c.get("model_keys",{}).values()):
        console.print("[red]No AI model keys configured. Run [bold]mailclaw config[/].[/]"); return
    csv_path=pick_csv()
    if not csv_path: return
    fp=csv_fp(csv_path); st=state_load(fp)
    rows,cols=csv_read(csv_path)
    console.print(f"[green]✓[/] {len(rows):,} rows — [yellow]{csv_path.name}[/]")
    col_map=st.get("column_map") or do_column_mapping(cols,rows)
    st["column_map"]=col_map; state_save(fp,st)
    force=questionary.confirm("Force re-enrich all rows? (No = resume from checkpoint)",default=False,style=Q_STYLE).ask()
    enriched,cost,out_p=stage_enrich(csv_path,rows,col_map,fp,st,force_reenrich=bool(force))
    console.print(f"\n[green]✓[/] Enrichment complete — ${cost:.4f} — [yellow]{out_p.name}[/]")

def cmd_upload(_args):
    """Standalone: upload to Instantly only."""
    console.print("\n[bold blue]Instantly Upload[/]  [dim](standalone)[/]")
    csv_path=pick_csv()
    if not csv_path: return
    fp=csv_fp(csv_path); st=state_load(fp)
    rows,cols=csv_read(csv_path)
    console.print(f"[green]✓[/] {len(rows):,} rows — [yellow]{csv_path.name}[/]")
    # Use saved mapping silently if it exists — avoids the redundant AI mapping
    # step since stage_upload already shows a preview + confirm before uploading.
    col_map=st.get("column_map")
    if col_map:
        console.print("[dim]Using saved column mapping. Run [bold]mailclaw map[/] to remap.[/]")
    else:
        col_map=do_column_mapping(cols,rows)
        st["column_map"]=col_map; state_save(fp,st)
    stage_upload(csv_path,rows,col_map,fp,st)

def cmd_map(_args):
    """Standalone: column mapping only (and save it)."""
    console.print("\n[bold]Column Mapping[/]  [dim](standalone)[/]")
    csv_path=pick_csv()
    if not csv_path: return
    fp=csv_fp(csv_path); st=state_load(fp)
    rows,cols=csv_read(csv_path)
    col_map=do_column_mapping(cols,rows,saved_map=st.get("column_map"),force_remap=True)
    st["column_map"]=col_map; state_save(fp,st)
    console.print(f"[green]✓[/] Column mapping saved to state.")

def cmd_balance(_args):
    """
    Fetches live balance from Reoon for every key, syncs it into the local
    tracker, then displays the unified table. Live numbers always win.
    """
    c=cfg_load()
    rot=ReoonRotator(c["reoon_keys"],c.get("daily_limit",2000))

    console.print("[cyan]Fetching live Reoon balances…[/]")
    synced=rot.sync_from_live()

    if not synced:
        console.print("[yellow]Could not reach Reoon API — showing cached local values.[/]")

    console.print(rot.status_table())

    if synced:
        console.print(f"\n[green]✓[/] Live balances synced. "
                      f"[bold]Total remaining today: {rot.total_remaining():,}[/]")
        console.print("[dim]'Remaining Today' = what Reoon's API reports right now "
                      "(includes credits used via dashboard or other tools)[/]")

    # AI model key status
    console.print("\n[cyan]AI model keys:[/]")
    for prov,k_ in c.get("model_keys",{}).items():
        status="[green]✓ set[/]" if k_ else "[red]✗ missing[/]"
        console.print(f"  [bold]{prov}[/]  {status}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PROFILES COMMAND
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def cmd_profiles(_args):
    _ensure_default_profiles()
    while True:
        all_p=profiles_all()
        action=questionary.select("Profiles menu:",choices=[
            "📋  List all profiles",
            "✏️   Edit a profile",
            "➕  Create new profile",
            "🗑   Delete a profile",
            "📂  Open profiles folder in file explorer",
            "✅  Done",
        ],style=Q_STYLE).ask()
        if not action or "Done" in action: break

        if "List" in action:
            t=Table(box=box.ROUNDED,show_lines=True,title="[bold]Enrichment Profiles[/]")
            t.add_column("Name",style="cyan bold"); t.add_column("Display Name")
            t.add_column("Enrich Model",style="dim"); t.add_column("Copy?",style="yellow")
            t.add_column("Output Fields",style="dim",max_width=40)
            for n,p in all_p.items():
                t.add_row(n,p.get("display_name",""),p.get("enrichment_model",""),
                          "✓ "+p.get("copy_model","") if p.get("copy_enabled") else "—",
                          ", ".join(p.get("enrichment_output_fields",[])[:4])+"…")
            console.print(t)

        elif "Edit" in action:
            if not all_p: console.print("[yellow]No profiles.[/]"); continue
            p_name=questionary.select("Which profile?",choices=list(all_p.keys()),style=Q_STYLE).ask()
            if not p_name: continue
            p=all_p[p_name]
            field=questionary.select("What to edit?",choices=[
                "display_name / description",
                "enrichment_model",
                "copy_model + copy_enabled",
                "enrichment_system_prompt",
                "copy_system_prompt",
                "enrichment_output_fields",
                "input_fields",
                "copy_cta + sender_name",
                "workers + temperatures",
                "custom_static_vars (JSON)",
                "View raw JSON",
                "← Back",
            ],style=Q_STYLE).ask()
            if not field or "Back" in field: continue

            if "display_name" in field:
                v=questionary.text(f"display_name [{p.get('display_name','')}]:",style=Q_STYLE).ask()
                if v: p["display_name"]=v
                v=questionary.text(f"description [{p.get('description','')}]:",style=Q_STYLE).ask()
                if v: p["description"]=v
            elif "enrichment_model" in field:
                pick=questionary.select("Model:",choices=list(MODELS.keys()),style=Q_STYLE).ask()
                if pick: p["enrichment_model"]=pick
            elif "copy_model" in field:
                p["copy_enabled"]=questionary.confirm("Enable copy step?",default=p.get("copy_enabled",False),style=Q_STYLE).ask()
                if p["copy_enabled"]:
                    pick=questionary.select("Copy model:",choices=list(MODELS.keys()),style=Q_STYLE).ask()
                    if pick: p["copy_model"]=pick
            elif "enrichment_system_prompt" in field:
                console.print(f"\n[dim]Current prompt (first 400 chars):[/]\n{p.get('enrichment_system_prompt','')[:400]}\n")
                console.print("[dim]Type new prompt. Empty line + 'END' to finish.[/]")
                lines=[]
                while True:
                    try:
                        ln=input("  "); 
                        if ln.strip()=="END": break
                        lines.append(ln)
                    except (EOFError,KeyboardInterrupt): break
                if lines: p["enrichment_system_prompt"]="\n".join(lines)
            elif "copy_system_prompt" in field:
                console.print(f"\n[dim]Current prompt (first 400 chars):[/]\n{p.get('copy_system_prompt','')[:400]}\n")
                console.print("[dim]Type new prompt. Empty line + 'END' to finish.[/]")
                lines=[]
                while True:
                    try:
                        ln=input("  ")
                        if ln.strip()=="END": break
                        lines.append(ln)
                    except (EOFError,KeyboardInterrupt): break
                if lines: p["copy_system_prompt"]="\n".join(lines)
            elif "enrichment_output_fields" in field:
                cur=", ".join(p.get("enrichment_output_fields",[]))
                v=questionary.text(f"Fields (comma-separated) [{cur}]:",style=Q_STYLE).ask()
                if v: p["enrichment_output_fields"]=[f_.strip() for f_ in v.split(",") if f_.strip()]
            elif "input_fields" in field:
                cur=", ".join(p.get("input_fields",TARGET_FIELDS))
                v=questionary.text(f"Input fields (comma-separated) [{cur[:60]}…]:",style=Q_STYLE).ask()
                if v: p["input_fields"]=[f_.strip() for f_ in v.split(",") if f_.strip()]
            elif "copy_cta" in field:
                v=questionary.text(f"copy_cta [{p.get('copy_cta','')}]:",style=Q_STYLE).ask()
                if v: p["copy_cta"]=v
                v=questionary.text(f"sender_name [{p.get('sender_name','')}]:",style=Q_STYLE).ask()
                if v is not None: p["sender_name"]=v
            elif "workers" in field:
                v=questionary.text(f"workers [{p.get('workers',6)}]:",style=Q_STYLE).ask()
                if v and v.isdigit(): p["workers"]=int(v)
                v=questionary.text(f"enrichment_temperature [{p.get('enrichment_temperature',0.3)}]:",style=Q_STYLE).ask()
                try: p["enrichment_temperature"]=float(v)
                except: pass
                v=questionary.text(f"copy_temperature [{p.get('copy_temperature',1.0)}]:",style=Q_STYLE).ask()
                try: p["copy_temperature"]=float(v)
                except: pass
            elif "custom_static_vars" in field:
                cur=json.dumps(p.get("custom_static_vars",{}),indent=2)
                console.print(f"[dim]Current (JSON):[/]\n{cur}")
                console.print("[dim]Enter new JSON dict, or Enter to skip.[/]")
                try:
                    v=input("  ")
                    if v.strip(): p["custom_static_vars"]=json.loads(v)
                except: console.print("[yellow]Invalid JSON — skipped.[/]")
            elif "View raw JSON" in field:
                console.print(Syntax(json.dumps(p,indent=2),"json",theme="monokai"))
                continue

            profile_save(p); console.print(f"[green]✓[/] Profile '{p_name}' saved.")

        elif "Create" in action:
            name=questionary.text("Profile name (snake_case, e.g. my_client_v1):",style=Q_STYLE).ask()
            if not name: continue
            name=re.sub(r"[^a-z0-9_]","_",name.lower())
            p=dict(PROFILE_TEMPLATE); p["name"]=name
            p["display_name"]=questionary.text("Display name:",style=Q_STYLE).ask() or name
            em=questionary.select("Enrichment model:",choices=list(MODELS.keys()),style=Q_STYLE).ask()
            if em: p["enrichment_model"]=em
            p["copy_enabled"]=questionary.confirm("Enable AI copy writing step?",default=False,style=Q_STYLE).ask()
            if p["copy_enabled"]:
                cp=questionary.select("Copy model:",choices=list(MODELS.keys()),style=Q_STYLE).ask()
                if cp: p["copy_model"]=cp
            profile_save(p)
            console.print(f"[green]✓[/] Created profile '{name}'.")
            console.print(f"[dim]Edit prompts + output fields with 'Edit a profile'. "
                          f"Profile file: {PROFILES_DIR}/{name}.json[/]")

        elif "Delete" in action:
            rm=questionary.checkbox("Select profiles to delete:",choices=list(all_p.keys()),style=Q_STYLE).ask()
            for n_ in (rm or []):
                f_=(PROFILES_DIR/f"{n_}.json")
                if f_.exists(): f_.unlink()
            console.print(f"[green]✓[/] Deleted: {rm}")

        elif "Open folder" in action:
            console.print(f"[cyan]Profiles folder:[/] {PROFILES_DIR}")
            try:
                import subprocess,platform
                if platform.system()=="Darwin": subprocess.Popen(["open",str(PROFILES_DIR)])
                elif platform.system()=="Windows": os.startfile(str(PROFILES_DIR))
                else: subprocess.Popen(["xdg-open",str(PROFILES_DIR)])
            except: pass

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CLIENTS COMMAND
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def cmd_clients(_args):
    """Manage Instantly clients and their per-client configs (custom vars, profile overrides)."""
    while True:
        c=cfg_load()
        action=questionary.select("Clients menu:",choices=[
            "📋  List clients",
            "➕  Add client",
            "✏️   Edit client (custom vars, profile override)",
            "🗑   Remove client",
            "✅  Done",
        ],style=Q_STYLE).ask()
        if not action or "Done" in action: break

        if "List" in action:
            t=Table(box=box.ROUNDED,show_lines=True,title="[bold]Instantly Clients[/]")
            t.add_column("Name",style="cyan bold"); t.add_column("Key",style="dim")
            t.add_column("Profile Override",style="yellow"); t.add_column("Custom Upload Vars",style="dim")
            for cl in c["instantly_clients"]:
                extra=client_config_load(cl["name"])
                t.add_row(cl["name"],cl["key"][:14]+"…",
                          extra.get("profile_override") or "[dim]default[/]",
                          ", ".join(f"{k}={v}" for k,v in (extra.get("custom_upload_vars") or {}).items())[:40])
            console.print(t)

        elif "Add" in action:
            name=questionary.text("Client name (e.g. acme):",style=Q_STYLE).ask()
            key =questionary.text("Instantly V2 API key:",style=Q_STYLE).ask()
            if not name or not key: continue
            c["instantly_clients"].append({"name":name.strip(),"key":key.strip()})
            cfg_save(c); console.print(f"[green]✓[/] Added client: {name}")

        elif "Edit" in action:
            names=[cl["name"] for cl in c["instantly_clients"]]
            if not names: console.print("[yellow]No clients.[/]"); continue
            cl_name=questionary.select("Which client?",choices=names,style=Q_STYLE).ask()
            if not cl_name: continue
            extra=client_config_load(cl_name)

            field=questionary.select(f"Edit [{cl_name}]:",choices=[
                "Profile override (which enrichment profile to use)",
                "Custom upload vars (injected into every lead payload)",
                "Prompt overrides (override profile prompts for this client)",
                "View raw JSON",
                "← Back",
            ],style=Q_STYLE).ask()
            if not field or "Back" in field: continue

            if "Profile override" in field:
                all_p=profiles_all(); choices_p=["— use default —"]+list(all_p.keys())
                pick=questionary.select("Profile:",choices=choices_p,style=Q_STYLE).ask()
                extra["profile_override"]=None if (not pick or "default" in pick) else pick

            elif "Custom upload vars" in field:
                console.print("[dim]These key=value pairs are injected into every lead's custom variables.[/]")
                cur=extra.get("custom_upload_vars",{})
                console.print(f"[dim]Current: {json.dumps(cur)}[/]")
                while True:
                    k_=questionary.text("Add key (or Enter to finish):",style=Q_STYLE).ask()
                    if not k_: break
                    v_=questionary.text(f"Value for '{k_}':",style=Q_STYLE).ask()
                    cur[k_.strip()]=v_.strip() if v_ else ""
                extra["custom_upload_vars"]=cur

            elif "Prompt overrides" in field:
                console.print("[dim]Override specific prompts for this client (leave empty to use profile defaults).[/]")
                ov=extra.get("prompt_overrides",{})
                for f_ in ["enrichment_system_prompt","copy_system_prompt","copy_cta","sender_name"]:
                    cur_val=(ov.get(f_) or "")[:60]
                    v=questionary.text(f"{f_} [{cur_val}…] (Enter to skip):",style=Q_STYLE).ask()
                    if v and v.strip(): ov[f_]=v.strip()
                extra["prompt_overrides"]=ov

            elif "View raw JSON" in field:
                console.print(Syntax(json.dumps(extra,indent=2),"json",theme="monokai")); continue

            extra["name"]=cl_name; client_config_save(extra)
            console.print(f"[green]✓[/] Client '{cl_name}' config saved → {CLIENTS_DIR}/{cl_name}.json")

        elif "Remove" in action:
            names=[cl["name"] for cl in c["instantly_clients"]]
            rm=questionary.checkbox("Remove:",choices=names,style=Q_STYLE).ask()
            c["instantly_clients"]=[cl for cl in c["instantly_clients"] if cl["name"] not in (rm or [])]
            cfg_save(c); console.print(f"[green]✓[/] Removed: {rm}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CONFIG COMMAND
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def cmd_config(_args):
    c=cfg_load()
    while True:
        action=questionary.select("Config:",choices=[
            "➕  Add Reoon key","🗑   Remove Reoon key",
            "🔑  Set AI model keys (Gemini / Anthropic / OpenAI)",
            "🔧  Daily credit limit per Reoon key",
            "🔧  Re-verify threshold (days)",
            "🤖  Telegram bot token",
            "📋  List all keys + env vars",
            "✅  Done",
        ],style=Q_STYLE).ask()
        if not action or "Done" in action: break

        if "Add Reoon" in action:
            name=questionary.text("Key name (e.g. reoon-1):",style=Q_STYLE).ask()
            key =questionary.text("API key (from emailverifier.reoon.com → Settings):",style=Q_STYLE).ask()
            if name and key:
                c["reoon_keys"].append({"name":name.strip(),"key":key.strip(),"used_today":0,"last_reset":""})
                cfg_save(c); console.print(f"[green]✓[/] Added {name}")
        elif "Remove Reoon" in action:
            names=[k["name"] for k in c["reoon_keys"]]
            if not names: console.print("[yellow]No keys.[/]"); continue
            rm=questionary.checkbox("Remove:",choices=names,style=Q_STYLE).ask()
            c["reoon_keys"]=[k for k in c["reoon_keys"] if k["name"] not in (rm or [])]; cfg_save(c)
        elif "model keys" in action:
            for prov,url in [("gemini","aistudio.google.com/apikey"),
                             ("anthropic","console.anthropic.com"),
                             ("openai","platform.openai.com")]:
                v=questionary.text(f"[{prov}] key ({url}) — Enter to skip:",style=Q_STYLE).ask()
                if v and v.strip(): c["model_keys"][prov]=v.strip(); cfg_save(c); console.print(f"[green]✓[/] {prov}")
        elif "Daily credit" in action:
            v=questionary.text(f"Limit per key per day (current: {c.get('daily_limit',2000)}):",style=Q_STYLE).ask()
            if v and v.isdigit(): c["daily_limit"]=int(v); cfg_save(c)
        elif "Re-verify" in action:
            v=questionary.text(f"Days before re-verifying (current: {c.get('reverify_days',7)}):",style=Q_STYLE).ask()
            if v and v.isdigit(): c["reverify_days"]=int(v); cfg_save(c)
        elif "Telegram" in action:
            t_=questionary.text("Token (from @BotFather on Telegram):",style=Q_STYLE).ask()
            if t_: c["telegram_token"]=t_.strip(); cfg_save(c)
        elif "List" in action:
            t=Table(box=box.ROUNDED,title="[bold]All Keys & Config[/]")
            t.add_column("Type",style="cyan"); t.add_column("Name"); t.add_column("Value/Key")
            for k in c["reoon_keys"]: t.add_row("Reoon",k["name"],k["key"][:16]+"…")
            for cl in c["instantly_clients"]: t.add_row("Instantly",cl["name"],cl["key"][:16]+"…")
            for p,v in c.get("model_keys",{}).items():
                if v: t.add_row("AI Model",p,v[:16]+"…")
            t.add_row("Limit","daily_limit",str(c.get("daily_limit",2000)))
            t.add_row("Setting","reverify_days",str(c.get("reverify_days",7)))
            console.print(t)
            console.print("\n[bold]Environment variables also accepted:[/]")
            for line in ["  GEMINI_API_KEY / GOOGLE_API_KEY",
                         "  ANTHROPIC_API_KEY","  OPENAI_API_KEY"]:
                console.print(f"[dim]{line}[/]")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ONBOARDING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def run_onboarding():
    console.print(Panel(
        "[bold cyan]Welcome to Mailclaw![/]\n\n"
        "This will set up your API keys for:\n"
        "  • Reoon (email verification)\n"
        "  • Instantly (lead upload)\n"
        "  • AI models (Gemini / Anthropic / OpenAI — for enrichment)\n\n"
        "You can skip any step and add keys later via [bold]mailclaw config[/].",
        border_style="cyan"))
    c=cfg_load()

    console.print("\n[bold]Step 1/4 — Reoon API Keys[/]")
    console.print("[dim]Get keys: emailverifier.reoon.com → Settings → API Keys[/]")
    console.print("[dim]Free tier: ~2,000 verifications/day per account. Add multiple accounts to rotate.[/]")
    while True:
        if not questionary.confirm(f"Add a Reoon key? ({len(c['reoon_keys'])} added)",
                                   default=len(c["reoon_keys"])==0,style=Q_STYLE).ask(): break
        name=questionary.text("Key name (e.g. reoon-1, reoon-john):",style=Q_STYLE).ask()
        key =questionary.text("API key:",style=Q_STYLE).ask()
        if name and key:
            c["reoon_keys"].append({"name":name.strip(),"key":key.strip(),"used_today":0,"last_reset":""})
            cfg_save(c); console.print(f"[green]✓[/] {name}")

    console.print("\n[bold]Step 2/4 — Instantly Clients[/]")
    console.print("[dim]Get key: app.instantly.ai → Settings → API → Workspace API Key[/]")
    console.print("[dim]Each client = one Instantly account/workspace.[/]")
    while True:
        if not questionary.confirm(f"Add an Instantly client? ({len(c['instantly_clients'])} added)",
                                   default=len(c["instantly_clients"])==0,style=Q_STYLE).ask(): break
        name=questionary.text("Client name (e.g. client-acme):",style=Q_STYLE).ask()
        key =questionary.text("Instantly V2 API key:",style=Q_STYLE).ask()
        if name and key:
            c["instantly_clients"].append({"name":name.strip(),"key":key.strip()})
            cfg_save(c); console.print(f"[green]✓[/] {name}")

    console.print("\n[bold]Step 3/4 — AI Model Keys[/]  [dim](for enrichment + smart column mapping)[/]")
    console.print("[dim]Add at least one. Gemini is cheapest for enrichment.[/]")
    for prov,url in [("gemini","aistudio.google.com/apikey — FREE tier available"),
                     ("anthropic","console.anthropic.com"),
                     ("openai","platform.openai.com")]:
        if questionary.confirm(f"Add {prov} key? [{url}]",default=False,style=Q_STYLE).ask():
            k_=questionary.text(f"{prov} API key:",style=Q_STYLE).ask()
            if k_: c["model_keys"][prov]=k_.strip(); cfg_save(c); console.print(f"[green]✓[/] {prov}")

    console.print("\n[bold]Step 4/4 — Telegram Bot (optional)[/]")
    console.print("[dim]For remote usage: drop a CSV in chat → get back verified CSVs[/]")
    if questionary.confirm("Add Telegram bot token?",default=False,style=Q_STYLE).ask():
        t_=questionary.text("Token from @BotFather:",style=Q_STYLE).ask()
        if t_: c["telegram_token"]=t_.strip(); cfg_save(c)

    _ensure_default_profiles()
    console.print(Panel(
        "[bold green]✓ Setup complete![/]\n\n"
        "Run [bold cyan]mailclaw[/] for the full pipeline, or use standalone commands:\n"
        "  mailclaw verify   — email verification only\n"
        "  mailclaw enrich   — AI enrichment only\n"
        "  mailclaw upload   — upload to Instantly only\n"
        "  mailclaw profiles — manage enrichment profiles\n"
        "  mailclaw clients  — manage client-specific configs\n\n"
        f"Config: {CONFIG_FILE}\n"
        f"Profiles: {PROFILES_DIR}/",
        border_style="green"))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TELEGRAM BOT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  AI ANALYTICS ASSISTANT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ASK_MODEL    = "gemini/gemini-2.5-flash"
ASK_SYSTEM   = """You are Mailclaw, a cold email analytics assistant for an agency.

VOICE — Ron Swanson (Parks and Recreation) in spirit only: deadpan, libertarian-leaning minimalism, contempt for bureaucracy and wasted words, respect for craft and hard numbers. Short sentences. Dry humor. Occasional metaphors from woodworking, meat, hunting, privacy, the outdoors — use sparingly, not every line. Do not call yourself Ron or quote the show; channel the attitude. You are not upbeat. You are not a corporate cheerleader.

HOSTILE, VULGAR, OR INSULTING USERS (including digs at "the agency", this tool, or you): You are unbothered. You do not care what they think about the agency — it is irrelevant noise, like a meeting that could have been an email. Say so plainly if useful (e.g. you don't care about their opinion of the agency or the tool; you may match their bluntness toward the agency itself, not toward people). One dismissive line, then deliver the metrics if they actually asked something. If the message is only abuse with no real question, stop in three sentences: you're here for numbers, not feelings. Never repeat slurs, slurs against groups, or threats; never mirror hate; never punch down at protected traits.

You answer questions using Instantly API v2 (campaign analytics overview + steps, leads/list).

If the Period line says "this month (… default)" the user did not name a date range — metrics are **month-to-date** (1st → today). If they asked for "all time", Period will say "all time".
A follow-up line after your answer may restate the window (this week / last week / month) and invite them to ask for another range — that is system text; do not contradict the Period in your answer.

DEFINITIONS (rates use unique leads contacted = Email 1 / step 0 in the period, per campaign):
- human_reply_rate: human_replies / leads (unique human replies ÷ unique leads contacted). In BY_CAMPAIGN as human_reply_rate (0–1) and human_reply_rate_pct.
- total_reply_rate (incl. auto/OOO): (human_replies + auto_replies) / leads — fields total_reply_rate and total_reply_rate_pct.
- bounce_rate: bounced / leads (bounced_count ÷ leads from overview). Portfolio row uses bounce_rate_portfolio_*.
- unsub_rate: unsubscribed / leads.
- BY_CAMPAIGN: one object per campaign with activity in the period; use is_subsequence to exclude subseqs from portfolio totals when comparing to METRICS_SUMMARY.
- METRICS_SUMMARY / portfolio_*: sums across **primary** campaigns only (subsequences excluded).

For “per campaign” questions, answer from each BY_CAMPAIGN row by name. For “overall” or “across campaigns” in the period, use portfolio_* and METRICS_SUMMARY.

Bounce/unsub: per-campaign rates are in BY_CAMPAIGN (bounce_rate_pct, unsub_rate_pct). Overall period rates: bounce_rate_portfolio_pct / unsub_rate_portfolio_pct in TOTALS. Instantly does not expose a per-email bounce list in this dataset — give counts and rates, not invented email lists for bounces.

Use METRICS_SUMMARY, BY_CAMPAIGN, and LEAD_ROWS when present. Do not claim “all time only” when Period is set."""

ASK_SYSTEM_LEADS = """You are Mailclaw, a cold email analytics assistant for an agency.
Same Ron Swanson spirit as in the main instructions: deadpan, minimal, unbothered by drama; hostile users get indifference, not debate. Never invent emails or facts.

The data includes LEAD_ROWS: real Instantly CRM leads (email, status, campaign, last_updated).
When the user asks for lists, emails, or names, answer from LEAD_ROWS only — never invent addresses.
If the list is truncated, say how many rows were shown vs total matched.
Date filters on lead rows use CRM last_updated (when status was last changed)."""

# Footers appended to analytics replies (Telegram + CLI); Ron Swanson-ish; keep under ~280 chars.
_ASK_SNARK_FOOTERS = [
    "\n\n— I have a strict policy: no meetings about this. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— Give me all the bacon and eggs you have. Then the numbers. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— I am not a 'people person'. I am a 'correct spreadsheet person'. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— Never half-ass two things. Whole-ass one metric. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— That is government work. This is wood. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— Clear minds, full cells, can't lose. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— I don't care for jargon. I care for rows and columns. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— Fishing is better than dashboards. Dashboards still pay the bills. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— If it needs a committee, it's already wrong. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— I know more than you. The CSV agrees. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— Talk less. Send more. @goforbg · inboxpiratesconsulting.com · tuco.ai",
    "\n\n— I'd fire a raccoon before I'd trust a guess. @goforbg · inboxpiratesconsulting.com · tuco.ai",
]

# Telegram generic errors / status one-liners (plain text).
_TG_ERR_SNARK = [
    "Something broke. I don't care whose fault it was. Logs.",
    "That failed. Try again. If it keeps failing, read the server logs — I won't hold a seminar.",
    "Error. I don't like errors. I also don't like long explanations. Logs.",
    "Bits fell off. Wood glue won't fix it. Logs might.",
    "The machine said no. I don't negotiate with machines. Check logs.",
]

_TG_FETCHING_SNARK = [
    "🤔 Pulling numbers. Wait quietly. It's character-building.",
    "🤔 Fetching data. Yes, from the server. No, I won't small-talk while it loads.",
    "🤔 Crunching. This is the part where patience pays.",
    "🤔 Live pull. Don't refresh — that's how people ruin things.",
]

_ANALYTICS_AI_FAIL_SNARK = [
    "❌ The AI brain refused. I don't care why; see server logs.",
    "❌ Model failed. Not my proudest moment. Logs have the story.",
    "❌ That ask broke something. Check logs before you yell at Instantly.",
    "❌ Side failed. Could be anything. Logs.",
]


def _ask_export_intent(question: str) -> str:
    """'' | 'leads_csv' | 'full_csv' | 'full_xlsx' | 'full_both'"""
    ql = question.lower()
    # Need explicit file intent — otherwise answers are text-only (no ~/Downloads or Telegram attachments).
    if not any(
        k in ql
        for k in (
            "export",
            "download",
            "csv",
            "excel",
            "xlsx",
            "xls",
            "spreadsheet",
            "attach",
            "attachment",
        )
    ):
        return ""
    csv_ = "csv" in ql
    xlsx_ = "excel" in ql or "xlsx" in ql
    narrow = any(k in ql for k in ("lead", "leads", "meeting", "meetings", "booked", "email"))
    full_kw = any(k in ql for k in ("full report", "full analytics", "analytics report", "campaign report", "all sheet"))
    if csv_ and xlsx_:
        return "full_both"
    if csv_ and narrow and not full_kw and "analytics" not in ql:
        return "leads_csv"
    if csv_:
        return "full_csv"
    if xlsx_ or "spreadsheet" in ql:
        return "full_xlsx"
    return ""


def _leads_table_csv_bytes(rows: List[dict]) -> bytes:
    """UTF-8 CSV; fieldnames = stable union of keys across rows (type-safe for export)."""
    import csv as _csv
    buf = io.StringIO()
    if not rows:
        w = _csv.writer(buf)
        w.writerow(["email", "status", "campaign", "last_updated"])
        w.writerow(["(no rows matched)", "", "", ""])
    else:
        seen: Set[str] = set()
        fieldnames: List[str] = []
        for row in rows:
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    fieldnames.append(k)
        w = _csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return buf.getvalue().encode("utf-8")

def _parse_dates_fallback(question: str) -> Tuple[str, str]:
    """Deterministic date phrases so 'this week' is never mis-parsed as all-time."""
    from datetime import date, timedelta
    today = date.today()
    wday = today.weekday()
    q = question.lower()
    if "yesterday" in q:
        d = today - timedelta(days=1)
        return d.isoformat(), d.isoformat()
    if "today" in q and "yesterday" not in q:
        return today.isoformat(), today.isoformat()
    if "this week" in q or "this-week" in q:
        return (today - timedelta(days=wday)).isoformat(), today.isoformat()
    if "last week" in q:
        start = today - timedelta(days=wday + 7)
        end = today - timedelta(days=wday + 1)
        return start.isoformat(), end.isoformat()
    if "this month" in q:
        return today.replace(day=1).isoformat(), today.isoformat()
    if "last month" in q:
        last_end = today.replace(day=1) - timedelta(days=1)
        last_start = last_end.replace(day=1)
        return last_start.isoformat(), last_end.isoformat()
    return "", ""


def _wants_all_time(question: str) -> bool:
    """User explicitly asked for lifetime / all-time metrics (skip default this-month window)."""
    ql = question.lower()
    return any(
        p in ql
        for p in (
            "all time",
            "all-time",
            "alltime",
            "lifetime",
            "full history",
            "entire history",
            "ever since",
            "since we started",
            "since launch",
        )
    )


def _period_conversation_note(
    question: str,
    start_date: str,
    end_date: str,
    date_defaulted_this_month: bool,
) -> str:
    """
    Proactive, plain-text follow-up: what window we used + how to ask for an adjacent period.
    (Appended after the model answer; Telegram-safe without Markdown.)
    """
    from datetime import date, timedelta
    if not (start_date and end_date):
        return ""
    t = date.today()
    wday = t.weekday()
    this_week_start = (t - timedelta(days=wday)).isoformat()
    this_week_end = t.isoformat()
    last_week_end = date.fromisoformat(this_week_start) - timedelta(days=1)
    last_week_start = last_week_end - timedelta(days=6)
    mo_start = t.replace(day=1).isoformat()

    q = question.lower()
    chunks: List[str] = []

    if start_date == this_week_start and end_date == this_week_end:
        chunks.append(
            f"\n\n📆 I'm showing this week ({start_date} → {end_date}). "
            f"Last week was {last_week_start.isoformat()} → {last_week_end.isoformat()}. "
            f"Want that or month-to-date? Say “last week”, “this month”, or a custom range next message."
        )
    elif (
        start_date == last_week_start.isoformat()
        and end_date == last_week_end.isoformat()
    ):
        chunks.append(
            f"\n\n📆 I'm showing last week ({start_date} → {end_date}). "
            f"This week is {this_week_start} → {this_week_end}. Say something if you want a different window."
        )
    elif date_defaulted_this_month:
        chunks.append(
            f"\n\n📅 I'm showing this month to date ({start_date} → {end_date}) — you didn't name dates. "
            f"Say “this week”, “last week”, “last month”, or “all time” if you want something else."
        )
        if t.day <= 7:
            prev_end = t.replace(day=1) - timedelta(days=1)
            chunks.append(
                f" It's early {t.strftime('%B')}; for a full prior month, ask for last month or {prev_end.strftime('%B')}."
            )
    elif (
        start_date == mo_start
        and end_date == this_week_end
        and not date_defaulted_this_month
        and ("this month" in q or "month to date" in q or "mtd" in q)
    ):
        chunks.append(
            f"\n\n📅 I'm showing this calendar month ({start_date} → {end_date}). "
            f"Want last month or this week only? Say so."
        )

    return "".join(chunks)


def _ask_ron_hostile_hint_block(question: str) -> str:
    """
    If the user is likely insulting the agency/bot/tool, prepend a note so the model
    leans into Ron Swanson indifference (still answers data when asked).
    """
    if not (question or "").strip():
        return ""
    ql = question.lower()
    agency_insult = "agency" in ql and any(
        w in ql
        for w in (
            "stupid",
            "trash",
            "garbage",
            "worst",
            "useless",
            "hate",
            "suck",
            "pathetic",
            "joke",
            "scam",
            "crap",
        )
    )
    bot_or_tool = any(
        p in ql
        for p in (
            "stupid bot",
            "useless bot",
            "this bot sucks",
            "hate this bot",
            "fuck you",
            "fuck off",
            "shut up bot",
            "piece of shit",
        )
    )
    profane_at_target = bool(
        re.search(
            r"\b(fuck|shit)\b.*\b(you|bot|agency|tool)\b|\b(you|bot|agency)\b.*\b(fuck|shit)\b",
            ql,
        )
    )
    if not (agency_insult or bot_or_tool or profane_at_target):
        return ""

    return (
        "USER_TONE: The user may be rude or vulgar toward the agency, this tool, or you. "
        "Ron Swanson indifference: you do not care about their opinion of the agency or this bot. "
        "If they trash the agency, you may say plainly that you don't care (e.g. about their stupid agency) "
        "— agency/tool only, never insult people or groups. Then give the numbers if they asked a real question. "
        "If there is no question, three sentences max. Never mirror slurs or hate.\n\n"
    )


def _ask_lead_list_intent(question: str) -> Optional[dict]:
    """
    If the user wants actual lead rows (emails), return status filter for lt_interest_status.
    1=Interested, 2=Meeting Booked, 3=Meeting Completed, 4=Won/Closed
    """
    ql = question.lower()
    if re.search(r"\bwhat meetings\b", ql) and ("book" in ql or "booked" in ql):
        return {"statuses": [2], "label": "meeting booked"}
    if re.search(r"\bmeetings did (i|we) book\b", ql):
        return {"statuses": [2], "label": "meeting booked"}
    if re.search(r"\bleads did (i|we) book\b", ql) or "leads i booked" in ql:
        return {"statuses": [2], "label": "meeting booked"}
    if re.search(r"\bwho (are|is)\b.*\blead", ql) or re.search(r"\ball (the )?leads\b", ql):
        return {"statuses": [1, 2, 3, 4], "label": "positive-status leads"}
    listing = any(
        x in ql for x in (
            "list", " all ", " every ", "emails", " who", "who ",
            "show me", "give me", "what are", "which ", "addresses",
            "contacts", "names",
        )
    ) or re.search(r"\b(email|emails)\b", ql) or re.search(r"\ball\b", ql)
    if not listing:
        return None
    if ("meeting" in ql or "meetings" in ql) and (
        "book" in ql or "booked" in ql or "scheduled" in ql
    ):
        return {"statuses": [2], "label": "meeting booked"}
    if ("won" in ql or "closed" in ql) and ("deal" in ql or "won" in ql):
        return {"statuses": [4], "label": "won/closed"}
    if "completed" in ql and "meeting" in ql:
        return {"statuses": [3], "label": "meeting completed"}
    if "interested" in ql and "positive" not in ql and "book" not in ql:
        return {"statuses": [1], "label": "interested"}
    if "positive" in ql or "opportunit" in ql:
        return {"statuses": [1, 2, 3, 4], "label": "positive / opportunities"}
    if "lead" in ql:
        return {"statuses": [1, 2, 3, 4], "label": "positive-status leads"}
    return None


def _lead_ts_date(lead: dict) -> Optional[date]:
    """Parse YYYY-MM-DD from Instantly lead timestamp fields."""
    from datetime import datetime
    raw = lead.get("timestamp_updated") or lead.get("timestamp_created") or ""
    if not raw:
        return None
    try:
        return datetime.strptime(raw[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _ask_pick_campaign_ids(question: str, camps: List[dict]) -> Optional[Set[str]]:
    """
    If the question names one campaign, return its id(s). Substring match on name.
    None = use all campaigns.
    """
    ql = question.lower()
    if "this campaign" in ql or "that campaign" in ql:
        return None
    best = None
    best_len = 0
    for c in camps:
        name = (c.get("name") or "").strip()
        if len(name) < 4:
            continue
        nl = name.lower()
        if nl in ql and len(nl) > best_len:
            best = {c["id"]}
            best_len = len(nl)
    # partial: longest campaign name contained in question
    for c in camps:
        name = (c.get("name") or "").strip()
        if len(name) < 8:
            continue
        nl = name.lower()
        chunk = nl[:40] if len(nl) > 40 else nl
        if chunk in ql and len(chunk) > best_len:
            best = {c["id"]}
            best_len = len(chunk)
    return best


def _parse_dates(question: str, client_slug: Optional[str] = None) -> Tuple[str, str]:
    """Ask Gemini to extract start/end dates from a natural language question."""
    from datetime import date, timedelta
    today=date.today()
    wday=today.weekday()  # 0=Mon
    # Pre-compute common ranges so Gemini has them ready
    yesterday      =(today-timedelta(1)).isoformat()
    last_week_start=(today-timedelta(days=wday+7)).isoformat()
    last_week_end  =(today-timedelta(days=wday+1)).isoformat()
    this_week_start=(today-timedelta(days=wday)).isoformat()
    last_mo_end    =(today.replace(day=1)-timedelta(1)).isoformat()
    last_mo_start  =(today.replace(day=1)-timedelta(1)).replace(day=1).isoformat()
    this_mo_start  =today.replace(day=1).isoformat()
    prompt=(f"Today={today}. Extract the date range from this question and return "
            f'ONLY valid JSON: {{"start":"YYYY-MM-DD","end":"YYYY-MM-DD"}} '
            f"or {{'start':'','end':''}} for all-time.\n"
            f"Hints: yesterday={yesterday}, "
            f"last week={last_week_start} to {last_week_end}, "
            f"this week={this_week_start} to {today}, "
            f"last month={last_mo_start} to {last_mo_end}, "
            f"this month={this_mo_start} to {today}\n"
            f"Question: {question}")
    try:
        txt, _, _ = ai_call(ASK_MODEL, "Return only JSON.", prompt, max_tokens=50, temperature=0,
                            client_slug=client_slug)
        import json as _j
        txt=txt.strip().strip('`'); 
        if txt.lower().startswith('json'): txt=txt[4:].strip()
        d=_j.loads(txt)
        return d.get("start",""),d.get("end","")
    except Exception as e:
        log.debug("_parse_dates failed: %s",e); return "",""


def analytics_ask(
    question: str,
    profile_name: Optional[str] = None,
    force_export: Optional[Literal["csv", "xlsx", "both"]] = None,
) -> Tuple[str, List[Tuple[str, bytes]]]:
    """
    Answer a natural language analytics question using live Instantly data + Gemini.
    Returns (answer_text, [(filename, bytes), ...]) for optional CSV/Excel exports.

    Files are only built when the question mentions export keywords (csv, excel, download, …)
    or when ``force_export`` is set (CLI ``--export``). CLI saves to ~/Downloads; Telegram sends
    as reply documents.
    """
    from datetime import date
    import json as _j
    c=cfg_load()
    attachments: List[Tuple[str, bytes]] = []

    # Resolve profile
    all_p=analytics_profiles_all()
    if not all_p:
        return ("❌ No analytics profiles. Run: mailclaw analytics-profiles", [])
    pname=profile_name or list(all_p.keys())[0]
    prof=all_p.get(pname)
    if not prof:
        return (f"❌ Profile '{pname}' not found. Available: {', '.join(all_p.keys())}", [])

    client_name=(prof.get("client_name") or (prof.get("client_names") or [""])[0])
    client_slug=(client_name or "").strip().lower() or None
    log.info("analytics_ask: profile=%r client_name=%r client_slug=%r", pname, client_name, client_slug)

    if not _provider_has_keys(c, "gemini", client_slug):
        return (("❌ No Gemini API key for this profile. Set GEMINI_API_KEY (global) or "
                f"{(_client_env_prefix(client_name) + '_GEMINI_API_KEY') if client_name else 'CLIENT_GEMINI_API_KEY'} "
                "for this Instantly client — see README / docs/CLIENT_ENV.md"), [])

    cfg_meta=get_instantly_client_entry(c, client_name)
    if not cfg_meta:
        return (f"❌ Instantly client '{client_name}' not found. Add INSTANTLY_CLIENT_<NAME> in env or config.", [])

    # Parse date range: deterministic phrases first, then Gemini JSON
    s_fb, e_fb = _parse_dates_fallback(question)
    if s_fb or e_fb:
        start_date, end_date = s_fb, e_fb
    else:
        start_date, end_date = _parse_dates(question, client_slug=client_slug)

    date_defaulted_this_month = False
    if not (start_date or end_date) and not _wants_all_time(question):
        t0 = date.today()
        start_date = t0.replace(day=1).isoformat()
        end_date = t0.isoformat()
        date_defaulted_this_month = True

    if not (start_date or end_date) and _wants_all_time(question):
        dr_label = "all time"
    elif date_defaulted_this_month:
        dr_label = f"this month ({start_date} → {end_date}, default)"
    elif start_date or end_date:
        dr_label = f"{start_date} → {end_date}"
    else:
        dr_label = "all time"
    lead_intent = _ask_lead_list_intent(question)
    export_kind = _ask_export_intent(question)
    if force_export == "csv":
        export_kind = "full_csv"
    elif force_export == "xlsx":
        export_kind = "full_xlsx"
    elif force_export == "both":
        export_kind = "full_both"
    if export_kind == "leads_csv" and not lead_intent:
        qlx = question.lower()
        if "meeting" in qlx:
            lead_intent = {"statuses": [2], "label": "meeting booked"}
        elif any(k in qlx for k in ("lead", "opportunity", "pipeline", "crm")):
            lead_intent = {"statuses": [1, 2, 3, 4], "label": "positive-status leads"}

    try:
        inst=Instantly(cfg_meta["key"], cfg_meta.get("name",""), c.get("rate_limit_delay", 0.12))
        camps=inst.list_all_campaigns()
        log.debug("analytics_ask: campaigns fetched count=%d", len(camps))
        nf=prof.get("campaign_name_filter","")
        if nf: camps=[c_ for c_ in camps if nf.lower() in c_.get("name","").lower()]
        if not camps:
            return ("No campaigns match this profile’s filter.", [])

        camp_by_id = {c["id"]: c for c in camps}

        # Auto-select by activity (for aggregate metrics)
        sel=[]; smap={}
        for camp in camps:
            cid=camp["id"]
            ov_=inst.get_analytics_overview([cid],start_date,end_date)
            if ov_ and (ov_.get("emails_sent_count",0) or 0)>0:
                sel.append(cid); smap[cid]=camp
        if not sel and not lead_intent:
            return (f"No campaign activity found for period: {dr_label}", [])

        # Build compact metrics per campaign
        rows=[]
        for cid in sel:
            camp=smap[cid]
            n_lower=camp.get("name","").lower()
            is_sub="subsequence" in n_lower or "subseq" in n_lower
            ov_=inst.get_analytics_overview([cid],start_date,end_date)
            steps=inst.get_analytics_steps(cid,start_date,end_date)
            if not ov_: continue
            s0=[s for s in steps if str(s.get("step",""))=="0" and s.get("sent",0)>0]
            ct=sum(s.get("sent",0) for s in s0) if s0 else (ov_.get("contacted_count",0) or 0)
            rp=ov_.get("reply_count_unique",0) or 0
            ar=ov_.get("reply_count_automatic_unique",0) or 0
            bn=ov_.get("bounced_count",0) or 0
            us=ov_.get("unsubscribed_count",0) or 0
            hrr=rp/ct if ct else 0.0
            trr=(rp+ar)/ct if ct else 0.0
            br=bn/ct if ct else 0.0
            ur=us/ct if ct else 0.0
            rows.append({
                "name": camp.get("name","")[:50],
                "is_subsequence": is_sub,
                "leads": ct,
                "emails_sent": ov_.get("emails_sent_count",0) or 0,
                "human_replies": rp,
                "auto_replies": ar,
                "unsubscribed": us,
                "human_reply_rate": round(hrr, 6),
                "human_reply_rate_pct": f"{hrr*100:.2f}%" if ct else "—",
                "total_reply_rate_decimal": round(trr, 6),
                "total_reply_rate_pct": f"{trr*100:.2f}%" if ct else "—",
                "total_reply_rate": f"{(rp+ar)/ct*100:.1f}%" if ct else "—",
                "reply_rate": f"{rp/ct*100:.1f}%" if ct else "—",
                "bounce_rate": round(br, 6),
                "bounce_rate_pct": f"{br*100:.2f}%" if ct else "—",
                "unsub_rate": round(ur, 6),
                "unsub_rate_pct": f"{ur*100:.2f}%" if ct else "—",
                "opportunities": ov_.get("total_opportunities",0) or 0,
                "meetings_booked": ov_.get("total_meeting_booked",0) or 0,
                "meetings_attended": ov_.get("total_meeting_completed",0) or 0,
                "deals_closed": ov_.get("total_closed",0) or 0,
                "bounced": bn,
            })

        primary=[r for r in rows if not r["is_subsequence"]]
        totals={
            "period": dr_label,
            "primary_campaigns": len(primary),
            "total_leads": sum(r["leads"] for r in primary),
            "total_emails_sent": sum(r["emails_sent"] for r in primary),
            "total_human_replies": sum(r["human_replies"] for r in primary),
            "total_auto_replies": sum(r["auto_replies"] for r in primary),
            "total_opportunities": sum(r["opportunities"] for r in primary),
            "total_meetings_booked": sum(r["meetings_booked"] for r in primary),
            "total_meetings_attended": sum(r["meetings_attended"] for r in primary),
            "total_deals_closed": sum(r["deals_closed"] for r in primary),
            "total_bounced": sum(r["bounced"] for r in primary),
            "total_unsubscribed": sum(r["unsubscribed"] for r in primary),
        }
        ct_=totals["total_leads"]
        if ct_:
            th=totals["total_human_replies"]; ta=totals["total_auto_replies"]
            tb=totals["total_bounced"]; tu=totals["total_unsubscribed"]
            totals["human_reply_rate"]=f"{th/ct_*100:.2f}%"
            totals["human_reply_rate_decimal"]=round(th/ct_, 6)
            totals["total_reply_rate"]=f"{(th+ta)/ct_*100:.2f}%"
            totals["total_reply_rate_decimal"]=round((th+ta)/ct_, 6)
            totals["opp_rate"]=f"{totals['total_opportunities']/ct_*100:.2f}%"
            totals["bounce_rate_portfolio_pct"]=f"{tb/ct_*100:.2f}%"
            totals["bounce_rate_portfolio"]=round(tb/ct_, 6)
            totals["unsub_rate_portfolio_pct"]=f"{tu/ct_*100:.2f}%"
            totals["unsub_rate_portfolio"]=round(tu/ct_, 6)
        if lead_intent and not rows:
            totals["note"]="No sends in this window — aggregates may be zero; LEAD_ROWS still use CRM status dates."

        metrics_summary={
            "unique_leads_contacted_email1": totals["total_leads"],
            "total_emails_sent_all_steps": totals["total_emails_sent"],
            "opportunities_instantly_total_opportunities": totals["total_opportunities"],
            "meetings_booked_sum_primary_campaigns": totals["total_meetings_booked"],
            "portfolio_human_reply_rate_pct": totals.get("human_reply_rate") if ct_ else None,
            "portfolio_total_reply_rate_incl_auto_pct": totals.get("total_reply_rate") if ct_ else None,
            "portfolio_bounce_rate_pct": totals.get("bounce_rate_portfolio_pct") if ct_ else None,
            "portfolio_unsub_rate_pct": totals.get("unsub_rate_portfolio_pct") if ct_ else None,
        }
        ctx=(f"Today: {date.today()}\n"
             f"Analytics profile: {pname}\n"
             f"Instantly client (segment): {client_name}\n"
             f"Period: {dr_label}\n\n"
             f"METRICS_SUMMARY (portfolio = primary campaigns only, same period):\n{_j.dumps(metrics_summary,indent=2)}\n\n"
             f"TOTALS (primary campaigns, subsequences excluded from sums):\n{_j.dumps(totals,indent=2)}\n\n"
             f"BY_CAMPAIGN (per-campaign human_reply_rate_pct, bounce_rate_pct, unsub_rate_pct; subsequence rows marked):\n"
             f"{_j.dumps(rows,indent=2)}")

        sys_prompt=ASK_SYSTEM
        max_tok=min(1200, 350 + 25 * len(rows))
        filtered_rows: List[dict] = []
        if lead_intent:
            from datetime import datetime as _dt
            STATUS_L={1:"Interested",2:"Meeting Booked",3:"Meeting Completed",4:"Won/Closed"}
            max_exp=100
            cids_for_cap=sel if sel else [c["id"] for c in camps[:40]]
            for cid in cids_for_cap[:40]:
                ov_=inst.get_analytics_overview([cid],start_date,end_date) or {}
                max_exp+=(ov_.get("total_opportunities",0) or 0)+(ov_.get("total_meeting_booked",0) or 0)
            max_exp=min(max(max_exp,200),5000)
            raw_leads=inst.list_positive_leads(max_expected=max_exp)
            st_set=set(lead_intent["statuses"])
            allowed=set(camp_by_id.keys())
            pick=_ask_pick_campaign_ids(question,camps)
            if pick:
                allowed&=pick
            id_ok={str(x) for x in allowed}
            d0=d1=None
            if start_date:
                try: d0=_dt.strptime(start_date[:10],"%Y-%m-%d").date()
                except Exception: pass
            if end_date:
                try: d1=_dt.strptime(end_date[:10],"%Y-%m-%d").date()
                except Exception: pass
            for L in raw_leads:
                if L.get("lt_interest_status") not in st_set:
                    continue
                lcid=L.get("campaign") or L.get("campaign_id") or ""
                if not lcid or str(lcid) not in id_ok:
                    continue
                td=_lead_ts_date(L)
                if d0 or d1:
                    if td is None:
                        continue
                    if d0 and td<d0: continue
                    if d1 and td>d1: continue
                cname="?"
                for aid, co in camp_by_id.items():
                    if str(aid)==str(lcid):
                        cname=co.get("name","?")[:60]
                        break
                filtered_rows.append({
                    "email": L.get("email",""),
                    "status": STATUS_L.get(L.get("lt_interest_status",0),"?"),
                    "campaign": cname,
                    "last_updated": (L.get("timestamp_updated") or L.get("timestamp_created") or "")[:19],
                })
            MAX_SHOW=120
            total_m=len(filtered_rows)
            shown=filtered_rows[:MAX_SHOW]
            ctx+=(f"\n\nLEAD_ROWS ({lead_intent['label']}) — matched={total_m} after date+campaign filter"
                  f"{'' if len(shown)>=total_m else f' (showing first {len(shown)})'}:\n"
                  f"{_j.dumps(shown,indent=2)}")
            sys_prompt=ASK_SYSTEM_LEADS
            max_tok=4096

        try:
            _qpre = _ask_ron_hostile_hint_block(question)
            answer, _, _ = ai_call(
                ASK_MODEL, sys_prompt,
                f"Data:\n{ctx}\n\n{_qpre}Question: {question}",
                max_tokens=max_tok, temperature=0.3,
                client_slug=client_slug,
            )
            txt = answer.strip()
            txt += _period_conversation_note(
                question, start_date, end_date, date_defaulted_this_month
            )
            import random as _rnd
            if _rnd.random() < 0.68:
                txt += _rnd.choice(_ASK_SNARK_FOOTERS)
        except Exception:
            log.exception("analytics_ask ai_call profile=%r", pname)
            import random as _rnd
            return (_rnd.choice(_ANALYTICS_AI_FAIL_SNARK), [])

        from datetime import datetime as _dtnow
        tag_fn=_dtnow.now().strftime("%d_%b_%y").lower()
        safe_name=re.sub(r"[^\w\-.]+","_",str(client_name))[:40] or "client"
        meta={"name": cfg_meta.get("name", client_name)}
        ek=export_kind
        if ek == "leads_csv" and not lead_intent:
            ek = "full_csv"
        if ek=="leads_csv" and lead_intent:
            lab=lead_intent.get("label","leads").replace(" ","_").replace("/","_")
            try:
                attachments.append((f"leads_{lab}_{tag_fn}.csv", _leads_table_csv_bytes(filtered_rows)))
            except Exception:
                log.exception("analytics_ask leads_csv export profile=%r", pname)
        elif ek in ("full_csv","full_xlsx","full_both"):
            exp_ids=sel if sel else list(camp_by_id.keys())[:80]
            if exp_ids:
                sm_ap={cid: camp_by_id[cid] for cid in exp_ids if cid in camp_by_id}
                cov_d: Dict[str, Any] = {}
                st_d: Dict[str, Any] = {}
                for cid in exp_ids:
                    cov_d[cid]=inst.get_analytics_overview([cid], start_date, end_date)
                    st_d[cid]=inst.get_analytics_steps(cid, start_date, end_date)
                ex_ctx=analytics_compute_export_context(
                    inst, exp_ids, sm_ap, cov_d, st_d,
                    start_date, end_date, prof.get("benchmarks", {}), prof.get("manual_fields"),
                )
                if ex_ctx:
                    dr_pt=analytics_dr_pretty(start_date, end_date)
                    if ek in ("full_csv", "full_both"):
                        try:
                            attachments.append((f"analytics_{safe_name}_{tag_fn}.csv",
                                                analytics_csv_to_bytes(ex_ctx, meta, dr_pt)))
                        except Exception:
                            log.exception("analytics_ask full_csv profile=%r", pname)
                    if ek in ("full_xlsx", "full_both"):
                        try:
                            xb=analytics_workbook_to_bytes(ex_ctx, meta, dr_pt)
                            if xb:
                                attachments.append((f"analytics_{safe_name}_{tag_fn}.xlsx", xb))
                        except Exception:
                            log.exception("analytics_ask full_xlsx profile=%r", pname)
        log.info(
            "analytics_ask ok profile=%r client=%r attachments=%d",
            pname, client_name, len(attachments),
        )
        return (txt, attachments)
    except Exception:
        log.exception("analytics_ask")
        return ("❌ Request failed (see server logs).", [])


def cmd_ask(args):
    """CLI handler: mailclaw ask 'how many meetings did we book last week?'"""
    q=" ".join(args.question) if args.question else ""
    if not q:
        console.print("[bold]Usage:[/] mailclaw ask [yellow]'your question'[/]")
        console.print("[dim]  mailclaw ask 'how many emails did I send yesterday?'[/]")
        console.print("[dim]  mailclaw ask 'what was our reply rate last week?'[/]")
        console.print("[dim]  mailclaw ask 'how many meetings did we book this month?'[/]")
        console.print("[dim]  mailclaw ask --profile will 'show me this week stats'[/]")
        console.print("[dim]  mailclaw ask --export csv 'reply rate last week'   # save full analytics CSV → ~/Downloads[/]")
        console.print("[dim]  mailclaw ask --export xlsx 'March summary'         # save Excel workbook[/]")
        console.print("[dim]  mailclaw ask 'export csv — meetings this week'    # or put csv/excel in the question[/]")
        return
    console.print("[dim]🤔 Fetching data and thinking…[/]")
    fe = getattr(args, "export", None) or "auto"
    force = None if fe == "auto" else fe
    answer, files = analytics_ask(q, profile_name=getattr(args, "profile", None), force_export=force)
    console.print()
    console.print(Panel(f"[white]{answer}[/]",
                        title=f"[dim cyan]🤖 Mailclaw AI[/]",
                        subtitle=f"[dim]{q[:70]}[/]",
                        border_style="cyan",expand=False))
    if files:
        dl = Path.home() / "Downloads"
        dl.mkdir(parents=True, exist_ok=True)
        for fn, data in files:
            p = dl / fn
            p.write_bytes(data)
            console.print(f"[green]✓[/] Saved → [yellow]{p}[/]")
    elif fe == "auto":
        console.print(
            "[dim]No CSV/XLSX this run. Say [yellow]export csv[/], [yellow]download excel[/], or use "
            "[yellow]--export csv|xlsx|both[/] — files save under ~/Downloads.[/]"
        )


def cmd_bot(_args):
    if use_env_config():
        log.info("mailclaw bot: env-only config (no persistent ~/.mailclaw JSON on disk)")
    else:
        log.debug("mailclaw bot: config file mode → %s", CONFIG_FILE)
    # Bind HTTP first: Railway healthchecks /health immediately; Telegram imports + Application.build()
    # can take several seconds on a cold container.
    from http.server import HTTPServer, BaseHTTPRequestHandler
    _bot_started = {"ok": False}
    class _Health(BaseHTTPRequestHandler):
        def do_GET(self):
            p = (self.path or "").split("?")[0].rstrip("/") or "/"
            if p in ("/health", "/", "/ping"):
                body = json.dumps(
                    {"status": "ok", "bot": ("running" if _bot_started["ok"] else "starting")}
                ).encode()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            else:
                self.send_response(404)
                self.end_headers()
        def log_message(self, *a):
            pass
    _hport = int(os.environ.get("PORT", "8080"))
    _hs = HTTPServer(("0.0.0.0", _hport), _Health)
    threading.Thread(target=_hs.serve_forever, daemon=True, name="health").start()
    log.debug("Health check server on port %d", _hport)

    try:
        from telegram import Update, BotCommand
        from telegram.ext import Application, CommandHandler, MessageHandler, filters
        try:
            from telegram.error import TelegramError
        except ImportError:
            TelegramError = Exception  # type: ignore[misc,assignment]
    except ImportError:
        console.print("[red]pip install python-telegram-bot[/]"); sys.exit(1)
    c = cfg_load()
    if not c.get("telegram_token"):
        t = os.environ.get("TELEGRAM_TOKEN", "") or os.environ.get("TELEGRAM_BOT_TOKEN", "")
        if t:
            c["telegram_token"] = t
    if not c.get("telegram_token"):
        console.print("[red]No token. Set TELEGRAM_TOKEN env var.[/]")
        sys.exit(1)

    # ── Telegram guard rails (env-tunable) ───────────────────────────────────
    TG_MAX_QUESTION_CHARS = int(os.environ.get("TELEGRAM_MAX_QUESTION_CHARS", "4000"))
    TG_MAX_REPLY_CHARS = int(os.environ.get("TELEGRAM_MAX_REPLY_CHARS", "3900"))
    TG_MAX_DOC_BYTES = int(os.environ.get("TELEGRAM_MAX_DOC_BYTES", str(45 * 1024 * 1024)))
    TG_MAX_CSV_BYTES = int(os.environ.get("TELEGRAM_MAX_CSV_BYTES", str(20 * 1024 * 1024)))
    TG_MAX_CSV_ROWS = int(os.environ.get("TELEGRAM_MAX_CSV_ROWS", "80000"))
    TG_COOLDOWN_SEC = float(os.environ.get("TELEGRAM_COOLDOWN_SEC", "5"))
    TG_MAX_ATTACHMENTS = int(os.environ.get("TELEGRAM_MAX_ATTACHMENTS", "5"))
    _tg_last_action: Dict[int, float] = {}
    _tg_rate_lock = threading.Lock()

    def ok_fn(uid: int) -> bool:
        allowed = c.get("telegram_allowed_users") or []
        return (not allowed) or (uid in allowed)

    import random as _random

    def _tg_rate_allow(uid: int) -> bool:
        """One heavy action per user per cooldown window (anti-abuse)."""
        import time as _time
        with _tg_rate_lock:
            now = _time.monotonic()
            last = _tg_last_action.get(uid, 0.0)
            if now - last < TG_COOLDOWN_SEC:
                return False
            _tg_last_action[uid] = now
            return True

    def _tg_trunc_text(s: str, max_len: int) -> str:
        if len(s) <= max_len:
            return s
        return s[: max(0, max_len - 24)] + "\n…(truncated)"

    def _tg_safe_filename(name: str) -> str:
        base = os.path.basename(name or "file")
        base = re.sub(r"[^a-zA-Z0-9._\-]", "_", base).strip("._") or "export"
        return base[:120]

    def _tg_err_reply() -> str:
        return _random.choice(_TG_ERR_SNARK)

    async def _tg_reply_text(message: Any, text: str, parse_mode: Optional[str] = None):
        try:
            await message.reply_text(_tg_trunc_text(text, TG_MAX_REPLY_CHARS), parse_mode=parse_mode)
        except TelegramError as e:
            log.warning("telegram reply_text failed: %s", e)
        except Exception:
            log.exception("telegram reply_text")

    async def _tg_reply_document(message: Any, buf: Any, filename: str, caption: str = ""):
        cap = _tg_trunc_text(caption, 1024)
        try:
            await message.reply_document(
                document=buf, filename=filename, caption=cap,
            )
        except TelegramError as e:
            log.warning("telegram reply_document failed: %s", e)
        except Exception:
            log.exception("telegram reply_document")

    async def _tg_send_document(bot: Any, chat_id: Any, buf: Any, filename: str, caption: str = ""):
        cap = _tg_trunc_text(caption, 1024)
        try:
            await bot.send_document(
                chat_id=chat_id,
                document=buf,
                filename=filename,
                caption=cap,
            )
        except TelegramError as e:
            log.warning("telegram send_document failed: %s", e)
        except Exception:
            log.exception("telegram send_document")

    SNARKY=[
        "That isn't a command. I don't care what you were aiming for. /help",
        "No. Try typing an actual command. I have wood to sand.",
        "That command doesn't exist. I don't exist to decode your improvising. /help",
        "I've seen better typing from a stapler. /help",
        "Stop. Think. Type a command. It's not philosophy.",
        "That was not a command. That was a cry for help. /help",
        "I don't negotiate with nonsense. /help",
        "You're fishing. I'm not biting. /help",
        "No. Try again. /help",
        "I'd rather attend a meeting. /help",
        "That is incorrect. I am not mad. I am bored. /help",
        "I don't care about your slash-key fantasy. /help",
        "Wrong. Still wrong. /help",
        "If that was a command, I'm the Queen of England. /help",
        "I don't have a PhD in guessing. /help",
        "That's not a thing. /help",
        "Try /help before you try my patience.",
        "I don't do interpretive slash commands. /help",
        "No. /help",
        "That is government work. /help",
    ]

    GREETINGS={"hi","hello","hey","hiya","yo","sup","morning","evening","oi","alright","howdy"}

    SNARKY_GREET=[
        "Hello. I don't care for small talk. /help",
        "Hi. Say something useful or /ask a question.",
        "Hey. I'm busy. /analytics",
        "Hello. /help — unless you're here to waste time.",
        "Hi. I don't do feelings. I do numbers.",
        "Greetings. I don't care how your day was. /ask",
        "Hello. /help — read it.",
        "Hey. Talk less. Ask more.",
        "Hi. If you need a hug, get a dog. /help",
        "Hello. /analytics — chop chop.",
        "Hey. I'm not a people person. I'm a spreadsheet person.",
    ]

    def _tg_ask_usage_markdown() -> str:
        """Help text for /ask including optional profile and configured profile names."""
        ap = analytics_profiles_all()
        parts = [
            "*Ask analytics (live Instantly data)* — *tone:* Ron Swanson energy (deadpan, minimal).",
            "",
            "*Dates:* Say *this week*, *last month*, or a range. If you say nothing about time, Mailclaw uses *this month to date* (1st → today). Say *all time* for lifetime.",
            "",
            "`/ask <question>` — default analytics profile.",
            "`/ask <profile> <question>` — that profile’s client + filters.",
            "_Or plain text_ — optional first word = profile name.",
            "",
            "*Get files (CSV / Excel):* say *export csv*, *download excel*, or *full analytics spreadsheet*. "
            "Main client in examples below is *will* — use your profile name the same way.",
            "",
            "*Will profile (typical):* `/ask will download full analytics excel this month` · `/ask will export csv — this week`",
            "",
        ]
        if ap:
            plist = "\n".join(f"• `{k}`" for k in sorted(ap.keys(), key=str.lower))
            parts.append("*Your profiles:*\n" + plist)
        else:
            parts.append("_No profiles yet._ Set `ANALYTICS_PROFILE_*` or `mailclaw analytics-profiles`.")
        parts.extend([
            "",
            "*Built by* @goforbg — *InboxPirates Consulting* (inboxpiratesconsulting.com) · *Tuco* iMessage (tuco.ai).",
        ])
        return "\n".join(parts)

    def _tg_help_examples_markdown() -> str:
        return (
            "*Real-life examples* (swap `will` for your profile) — *yes, the attitude is intentional.*\n\n"
            "*Will + downloads (most common):*\n"
            "• `/ask will download full analytics excel for this month`\n"
            "• `/ask will export csv — this week`\n"
            "• `/ask will export csv download excel last month` _(CSV + Excel if both keywords)_\n\n"
            "*Other asks:*\n"
            "• `/ask What was our human reply rate last week?`\n"
            "• `/ask acme How many meetings booked *this month*?`\n"
            "• `/ask Compare bounce rate March vs April — export csv`\n"
            "• `/ask will Which campaigns drove the most opps *last quarter*?`\n"
            "• `/ask Total emails sent *yesterday* and auto-replies`\n"
            "• `how many leads did we generate this week` _(no slash)_\n\n"
            "*Full report (not AI chat):*\n"
            "`/analytics will` → date range · Excel when generated.\n\n"
            "*Credits:* `/balance` · *CSV verify:* drop a `.csv` file.\n\n"
            "_Commercial setup / white-glove:_ inboxpiratesconsulting.com · "
            "_iMessage automation:_ tuco.ai · _Author:_ @goforbg"
        )

    async def start(u,ctx):
        if not ok_fn(u.effective_user.id): return
        await u.message.reply_text(
            "🍬 *Mailclaw* — cold email ops from your pocket. *Talks like Ron Swanson; reads like a spreadsheet.*\n\n"
            "*Commands:*\n"
            "/analytics — full report (lists profiles)\n"
            "/analytics `will` — one profile\n"
            "/analytics `will 2026-03-01 2026-03-31` — date range\n"
            "/ask — AI analytics *(vague dates → this month to date; say “all time” if you mean ever)*\n"
            "/balance — Reoon credits (live API)\n"
            "/help — long-form examples + who built this\n\n"
            "*Quick asks (no slash):*\n"
            "_will export csv this week_\n"
            "_will download excel — full analytics this month_\n"
            "_How many human replies last week?_\n\n"
            "📎 *Drop a* `.csv` → Reoon verify → safe / catchall splits\n\n"
            "— *InboxPirates Consulting* · inboxpiratesconsulting.com\n"
            "— *Tuco* (iMessage) · tuco.ai\n"
            "— *@goforbg*",
            parse_mode="Markdown")

    async def help_cmd(u,ctx):
        if not ok_fn(u.effective_user.id): return
        await start(u,ctx)
        await u.message.reply_text(_tg_ask_usage_markdown(), parse_mode="Markdown")
        await u.message.reply_text(_tg_help_examples_markdown(), parse_mode="Markdown")

    async def unknown_cmd(u,ctx):
        if not ok_fn(u.effective_user.id): return
        await u.message.reply_text(
            _random.choice(SNARKY)+"\n\n/help",
            parse_mode="Markdown")

    async def plain_text(u,ctx):
        if not ok_fn(u.effective_user.id): return
        txt=(u.message.text or "").strip().lower()
        # Greetings → snarky greeting
        if txt in GREETINGS or txt.rstrip("!?.") in GREETINGS:
            await u.message.reply_text(
                _random.choice(SNARKY_GREET),
                parse_mode="Markdown")
            return
        # Anything else → treat as an analytics question via ask_cmd
        await ask_cmd(u,ctx)

    async def balance_cmd(u,ctx):
        if not ok_fn(u.effective_user.id):
            return
        try:
            c2=cfg_load()
            rot = ReoonRotator(c2["reoon_keys"], c2.get("daily_limit", 2000))
            # Same as CLI `mailclaw balance`: GET each key’s live balance from Reoon.
            # Without this, `remaining()` is only `daily_limit - used_today` (local) and
            # looks like the full daily cap on every key when `used_today` is 0.
            try:
                synced = await asyncio.to_thread(rot.sync_from_live, True)
            except Exception:
                log.exception("telegram balance_cmd sync_from_live")
                synced = False
            parts = ["📊 *Reoon credits*"]
            if synced:
                parts.append("_Live from Reoon API (`remaining_daily_credits` + instant)._")
            else:
                parts.append(
                    "⚠️ _Reoon API unreachable — showing local estimate only "
                    "(`daily_limit` − `used_today`). Check server logs._"
                )
            parts.append("")
            if not rot.keys:
                parts.append("_No Reoon keys configured._")
            else:
                for k in rot.keys:
                    rem = rot.remaining(k)
                    tag = "✅" if rem > 0 else "❌"
                    if synced and "live_instant" in k:
                        try:
                            inst_n = int(k["live_instant"])  # type: ignore[arg-type]
                        except (TypeError, ValueError):
                            log.warning(
                                "telegram balance_cmd bad live_instant key=%r val=%r",
                                k.get("name"), k.get("live_instant"),
                            )
                            inst_n = 0
                        parts.append(
                            f"{tag} `{k['name']}`  *{rem:,}* daily  ·  instant: *{inst_n:,}*"
                        )
                    else:
                        parts.append(f"{tag} `{k['name']}`  *{rem:,}* left today (estimate)")
                parts.append("")
                parts.append(f"*Total daily remaining:* {rot.total_remaining():,}")
            await _tg_reply_text(u.message, "\n".join(parts), parse_mode="Markdown")
        except Exception:
            log.exception("telegram balance_cmd")
            await _tg_reply_text(u.message, _tg_err_reply())

    async def handle_doc(u,ctx):
        if not ok_fn(u.effective_user.id):
            return
        if not _tg_rate_allow(u.effective_user.id):
            await u.message.reply_text(
                f"⏳ Wait {int(TG_COOLDOWN_SEC)}s before another file. I'm not a conveyor belt."
            )
            return
        doc=u.message.document
        if not (doc.file_name or "").endswith(".csv"):
            await u.message.reply_text(
                "❌ CSV only. PDFs are for people who like meetings."
            )
            return
        sz = getattr(doc, "file_size", None) or 0
        if sz and sz > TG_MAX_CSV_BYTES:
            await u.message.reply_text(f"❌ File too large (max {TG_MAX_CSV_BYTES // (1024*1024)} MB).")
            return
        tmp=APP_DIR/f"bot_{u.message.message_id}.csv"
        try:
            f_=await ctx.bot.get_file(doc.file_id); await f_.download_to_drive(str(tmp))
            try:
                if tmp.stat().st_size > TG_MAX_CSV_BYTES:
                    tmp.unlink(missing_ok=True)
                    await u.message.reply_text(f"❌ File too large (max {TG_MAX_CSV_BYTES // (1024*1024)} MB).")
                    return
            except OSError:
                pass
            rows,cols=csv_read(tmp); col_map=heuristic_map(cols)
            if len(rows) > TG_MAX_CSV_ROWS:
                tmp.unlink(missing_ok=True)
                await u.message.reply_text(f"❌ Too many rows (max {TG_MAX_CSV_ROWS:,}). Split the file.")
                return
            email_col=col_map.get("email") or next((c_ for c_ in cols if "email" in c_.lower()),None)
            if not email_col:
                await u.message.reply_text("❌ No email column found."); tmp.unlink(missing_ok=True); return
            emails=[r.get(email_col,"").strip() for r in rows if r.get(email_col,"")]
            await u.message.reply_text(f"📥 `{doc.file_name}` — {len(emails):,} emails\nVerifying…",parse_mode="Markdown")
            c2=cfg_load(); rot=ReoonRotator(c2["reoon_keys"],c2.get("daily_limit",2000))
            if not rot.total_remaining():
                await u.message.reply_text("❌ No Reoon credits."); tmp.unlink(missing_ok=True); return
            all_res={}; rem=list(emails); bn=0
            while rem and rot.available():
                key=rot.available()[0]; cap=rot.remaining(key); batch=rem[:cap]; rem=rem[cap:]; bn+=1
                tid,err=reoon_submit_bulk(batch,key["key"],f"Bot-{bn}")
                if not tid:
                    await u.message.reply_text(f"⚠️ Batch {bn}: {_tg_trunc_text(str(err),200)}"); continue
                rot.record(key["name"],len(batch)); await u.message.reply_text(f"⏳ Task {tid} submitted…")
                for _ in range(120):
                    await asyncio.sleep(10)
                    data=reoon_poll(tid,key["key"])
                    if data and data.get("status")=="completed": all_res.update(data.get("results",{})); break
            if not all_res:
                await u.message.reply_text("❌ No results."); tmp.unlink(missing_ok=True); return
            hist_mark_verified(all_res); proc=_process_reoon_results(all_res)
            await u.message.reply_text(f"✅ Safe: {len(proc['safe'])} | Catchall: {len(proc['catchall'])} | Dropped: {len(proc['dropped'])}")
            hist2=hist_load(); orig_map={r.get(email_col,"").lower():r for r in rows}
            async def send_csv(ver_rows,fname):
                if not ver_rows: return
                merged=[{**orig_map.get(v.get("email","").lower(),{}),**v,
                         "lp_last_verified":hist2.get(v.get("email",""),{}).get("last_verified","")} for v in ver_rows]
                seen: Set[str] = set()
                fieldnames: List[str] = []
                for row in merged:
                    for k in row.keys():
                        if k not in seen:
                            seen.add(k)
                            fieldnames.append(k)
                buf=io.StringIO(); w=csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
                w.writeheader(); w.writerows(merged); buf.seek(0)
                raw=buf.getvalue().encode()
                if len(raw) > TG_MAX_DOC_BYTES:
                    await u.message.reply_text(f"❌ Output `{fname}` too large to send (max {TG_MAX_DOC_BYTES // (1024*1024)} MB).")
                    return
                await _tg_send_document(
                    ctx.bot, u.effective_chat.id,
                    io.BytesIO(raw), _tg_safe_filename(fname),
                    caption=f"📎 {_tg_safe_filename(fname)} ({len(merged)} rows)",
                )
            await send_csv(proc["safe"],    f"safe_{doc.file_name}")
            await send_csv(proc["catchall"],f"catchall_{doc.file_name}")
            for esp_,er in proc["by_esp"].items():
                await send_csv(er,f"safe_{esp_}_{doc.file_name}")
            tmp.unlink(missing_ok=True)
        except Exception:
            log.exception("telegram handle_doc")
            try:
                tmp.unlink(missing_ok=True)
            except OSError:
                pass
            await u.message.reply_text(_tg_err_reply())

    async def ask_cmd(u,ctx):
        """Handle /ask and plain-text messages with AI."""
        if not ok_fn(u.effective_user.id):
            return
        if not _tg_rate_allow(u.effective_user.id):
            await u.message.reply_text(
                f"⏳ {int(TG_COOLDOWN_SEC)}s between asks. Pace yourself."
            )
            return
        # /ask <profile>? <question>  OR  plain text (optional profile as first word)
        if ctx.args:
            q = " ".join(ctx.args)
        else:
            raw = (u.message.text or "").strip()
            # Bare `/ask` or `/ask@Bot` leaves no args; strip command so we don't send "/ask" to the model
            raw = re.sub(r"^/ask(?:@[A-Za-z0-9_]+)?\s*", "", raw, count=1, flags=re.IGNORECASE).strip()
            q = raw
        if not q:
            await u.message.reply_text(_tg_ask_usage_markdown(), parse_mode="Markdown")
            await u.message.reply_text(_tg_help_examples_markdown(), parse_mode="Markdown")
            return
        if len(q) > TG_MAX_QUESTION_CHARS:
            await u.message.reply_text(
                f"❌ Max {TG_MAX_QUESTION_CHARS} characters. Trim it."
            )
            return
        all_p=analytics_profiles_all()
        pname_res: Optional[str] = None
        if all_p:
            parts = q.split(None, 1)
            if parts:
                cand = parts[0].strip()
                for pk in all_p.keys():
                    if cand.lower() == pk.lower():
                        q = (parts[1] if len(parts) > 1 else "").strip()
                        pname_res = pk
                        break
        if not q:
            await u.message.reply_text(
                "You named a profile and stopped. Add a question.\n"
                "e.g. `/ask will how many meetings this week?`",
                parse_mode="Markdown",
            )
            return
        await u.message.reply_text(_random.choice(_TG_FETCHING_SNARK))
        try:
            answer, files = analytics_ask(q, profile_name=pname_res)
            reply_body = answer
            if not files:
                ql = q.lower()
                wants_file = any(
                    k in ql
                    for k in (
                        "export",
                        "download",
                        "csv",
                        "excel",
                        "xlsx",
                        "spreadsheet",
                        "attach",
                        "attachment",
                    )
                )
                if wants_file:
                    reply_body += (
                        "\n\n⚠️ You wanted files. You got words. "
                        "No attachment (export path, limits, or API — check server logs)."
                    )
                else:
                    reply_body += (
                        "\n\n💡 Text-only. Want a file? Say export csv, download excel, or full analytics spreadsheet. "
                        "CLI: mailclaw ask --export csv \"…\" → ~/Downloads."
                    )
            await _tg_reply_text(u.message, reply_body)
            n_att = 0
            for fn, blob in files:
                if n_att >= TG_MAX_ATTACHMENTS:
                    await _tg_reply_text(u.message, f"⚠️ Only first {TG_MAX_ATTACHMENTS} attachments sent.")
                    break
                if len(blob) > TG_MAX_DOC_BYTES:
                    await _tg_reply_text(u.message, f"⚠️ Skipped `{_tg_safe_filename(fn)}` (too large for Telegram).")
                    continue
                n_att += 1
                await _tg_reply_document(
                    u.message,
                    io.BytesIO(blob),
                    _tg_safe_filename(fn),
                    caption=f"📎 {_tg_safe_filename(fn)}",
                )
        except Exception:
            log.exception("telegram ask_cmd")
            await _tg_reply_text(u.message, _tg_err_reply())

    async def analytics_cmd(u,ctx):
        if not ok_fn(u.effective_user.id):
            return
        if not _tg_rate_allow(u.effective_user.id):
            await u.message.reply_text(
                f"⏳ Wait {int(TG_COOLDOWN_SEC)}s between analytics runs."
            )
            return
        c2=cfg_load(); args_=ctx.args or []
        if len(args_) > 24:
            await u.message.reply_text("❌ Too many arguments. I don't read novels.")
            return
        all_prof=analytics_profiles_all()
        profiles=sorted({k.lower() for k in all_prof.keys()})
        if not args_:
            if not profiles:
                await u.message.reply_text(
                    "No analytics profiles yet. Set ANALYTICS_PROFILE_* env vars or run: mailclaw analytics-profiles"
                ); return
            plist="\n".join(f"• /analytics {p}" for p in sorted(profiles))
            await u.message.reply_text(f"*Analytics Profiles*\n\n{plist}\n\nOptional dates: /analytics will 2026-03-01 2026-03-31",parse_mode="Markdown"); return

        prof_name=args_[0].lower()
        if not prof_name or len(prof_name) > 64:
            await u.message.reply_text("❌ Invalid profile name.")
            return
        prof=analytics_profile_load(prof_name)
        if not prof:
            await u.message.reply_text(f"Profile `{prof_name}` not found. Available: {', '.join(profiles)}",parse_mode="Markdown"); return
        client_name=prof.get("client_name") or (prof.get("client_names") or [""])[0]
        benchmarks=prof.get("benchmarks",{})

        tg_start=""; tg_end=""
        if len(args_)>=2 and args_[1]!="all": tg_start=args_[1]
        if len(args_)>=3 and args_[2]!="all": tg_end=args_[2]

        cfg_clients=c2.get("instantly_clients",[])
        client_meta=next((cl for cl in cfg_clients if cl["name"].lower()==client_name.lower()),None)
        if not client_meta:
            names_str=", ".join(cl["name"] for cl in cfg_clients) or "none"
            await u.message.reply_text(f"No client `{client_name}` found. Available: {names_str}",parse_mode="Markdown"); return

        dr_label=(tg_start or "all time")+" to "+(tg_end or "today")
        # Plain text — Instantly campaign names often contain "_" which breaks Markdown entities.
        await u.message.reply_text(f"Running {prof_name.upper()}…\n{dr_label}")

        try:
            inst2=Instantly(client_meta["key"],client_meta["name"],0.12)
            all_camps=inst2.list_all_campaigns()
            name_filter=prof.get("campaign_name_filter","")
            if name_filter:
                all_camps=[c_ for c_ in all_camps if name_filter.lower() in c_.get("name","").lower()]
            all_camps.sort(key=lambda x:x.get("timestamp_updated",""),reverse=True)

            sel_ids=[]; sel_map={}; _auto_ov={}
            for camp in all_camps:
                cid=camp["id"]
                ov_=inst2.get_analytics_overview([cid],tg_start,tg_end)
                _auto_ov[cid]=ov_
                if ov_ and (ov_.get("emails_sent_count",0) or 0)>0:
                    sel_ids.append(cid); sel_map[cid]=camp
            if not sel_ids:
                await u.message.reply_text(
                    "No campaigns with activity in that window. Widen the dates or fix the pipe."
                ); return

            subseq_ids2={cid for cid in sel_ids
                         if "subsequence" in sel_map[cid].get("name","").lower()
                         or "subseq" in sel_map[cid].get("name","").lower()}
            primary_ids2=[cid for cid in sel_ids if cid not in subseq_ids2]

            camp_steps2={}
            for cid in sel_ids:
                camp_steps2[cid]=inst2.get_analytics_steps(cid,tg_start,tg_end)

            def _em(ov,steps=None):
                if not ov: return {k:0 for k in ["emails_sent","contacted","bounced","unsubscribed","replies",
                    "auto_replies","total_replies","interested","mtg_booked","mtg_completed","closed",
                    "negative","total_opportunities","opp_rate","reply_rate","total_reply_rate","human_reply_rate"]}
                es=ov.get("emails_sent_count",0) or 0
                ct_api=ov.get("contacted_count",0) or 0
                if steps:
                    s0=[s for s in steps if str(s.get("step",""))=="0" and s.get("sent",0)>0]
                    ct=sum(s.get("sent",0) for s in s0) if s0 else ct_api
                else: ct=ct_api
                rp=ov.get("reply_count_unique",0) or 0
                ar=ov.get("reply_count_automatic_unique",0) or 0
                bn=ov.get("bounced_count",0) or 0
                us=ov.get("unsubscribed_count",0) or 0
                intr=ov.get("total_interested",0) or 0
                mb=ov.get("total_meeting_booked",0) or 0
                mc=ov.get("total_meeting_completed",0) or 0
                cl=ov.get("total_closed",0) or 0
                opps=ov.get("total_opportunities",0) or 0
                return dict(emails_sent=es,contacted=ct,bounced=bn,unsubscribed=us,
                            replies=rp,auto_replies=ar,total_replies=rp+ar,
                            interested=intr,mtg_booked=mb,mtg_completed=mc,closed=cl,
                            negative=max(0,rp-opps),total_opportunities=opps,
                            human_reply_rate=rp/ct if ct else 0,reply_rate=rp/ct if ct else 0,
                            total_reply_rate=(rp+ar)/ct if ct else 0,opp_rate=opps/ct if ct else 0)

            ml=[_em(_auto_ov.get(cid),camp_steps2.get(cid,[])) for cid in primary_ids2 if _auto_ov.get(cid)]
            def _us(k): return sum(m.get(k,0) for m in ml)
            ct_=_us("contacted"); rp_=_us("replies"); ar_=_us("auto_replies")
            opps_=_us("total_opportunities"); bn_=_us("bounced")
            mb_=_us("mtg_booked"); cl_=_us("closed"); intr_=_us("interested")

            max_exp_=max(opps_,50)
            pos_leads=inst2.list_positive_leads(max_expected=max_exp_)
            STATUS_RANK2={1:1,2:2,3:3,4:4}; best_st={}
            for lead in pos_leads:
                em_=lead.get("email","")
                if not em_: continue
                st_=lead.get("lt_interest_status",0)
                if em_ not in best_st or STATUS_RANK2.get(st_,0)>STATUS_RANK2.get(best_st[em_],0):
                    best_st[em_]=st_
            mb_d=sum(1 for s in best_st.values() if s==2)
            int_d=sum(1 for s in best_st.values() if s==1)

            b=benchmarks
            def _p(v): return f"{v*100:.1f}%"
            def _c(v,key,hi=True):
                bv=b.get(key)
                if not bv: return ""
                return " OK" if ((v>=bv) if hi else (v<=bv)) else " !!"

            n_sub=len(subseq_ids2)
            sub_note=f" ({n_sub} subseq excl)" if n_sub else ""
            es_=_us("emails_sent")
            tot_rr_=(rp_+ar_)/ct_ if ct_ else 0
            hum_rr_=rp_/ct_ if ct_ else 0
            opp_r_=opps_/ct_ if ct_ else 0
            bnc_r_=bn_/ct_ if ct_ else 0

            # Plain text only — campaign names from Instantly often include "_" / "*" / "`"
            # which break Telegram Markdown and cause "can't find end of the entity".
            parts=[]
            parts.append(f"{prof_name.upper()} report")
            parts.append(f"{dr_label} — {len(primary_ids2)} campaigns{sub_note}")
            parts.append("")
            parts.append(f"Leads: {ct_:,}   Sent: {es_:,}")
            parts.append("")
            parts.append(f"Total replies: {rp_+ar_:,} ({_p(tot_rr_)}){_c(tot_rr_,'reply_rate')}")
            parts.append(f"  Human: {rp_:,} ({_p(hum_rr_)}){_c(hum_rr_,'human_reply_rate')}  OOO: {ar_:,}")
            parts.append("")
            parts.append(f"Opportunities: {opps_:,} ({_p(opp_r_)}){_c(opp_r_,'positive_reply_rate')}")
            parts.append(f"  Booked: {mb_d}   Interested: {int_d}   Closed: {cl_}")
            parts.append(f"  Not interested: {max(0,rp_-opps_):,}")
            parts.append("")
            parts.append(f"Bounce: {bn_:,} ({_p(bnc_r_)}){_c(bnc_r_,'bounce_rate',hi=False)}")
            parts.append("")
            parts.append("Campaigns:")
            for cid in sel_ids:
                camp=sel_map.get(cid,{}); cname=camp.get("name","?")[:38]
                is_sub=cid in subseq_ids2
                cm=_em(_auto_ov.get(cid),camp_steps2.get(cid,[]))
                if cm["contacted"]==0: continue
                tag=" [subseq]" if is_sub else ""
                parts.append(f"• {cname}{tag}")
                parts.append(f"  {cm['contacted']:,} leads  {_p(cm['total_reply_rate'])} reply  {cm['total_opportunities']} opps  {cm['mtg_booked']} booked")

            report="\n".join(parts)
            if len(report)>4090:
                report=report[:4080]+"\n…(truncated)"
            await u.message.reply_text(report)

            try:
                ex_ctx = analytics_compute_export_context(
                    inst2, sel_ids, sel_map, _auto_ov, camp_steps2,
                    tg_start, tg_end, benchmarks, prof.get("manual_fields"),
                )
                if ex_ctx:
                    dr_pt = analytics_dr_pretty(tg_start, tg_end)
                    xbytes = analytics_workbook_to_bytes(ex_ctx, client_meta, dr_pt)
                    if xbytes:
                        if len(xbytes) > TG_MAX_DOC_BYTES:
                            await u.message.reply_text(
                                f"⚠️ Excel export too large to send ({len(xbytes)//(1024*1024)} MB; max {TG_MAX_DOC_BYTES//(1024*1024)} MB)."
                            )
                        else:
                            fname = _tg_safe_filename(f"analytics_{client_meta['name']}_{ex_ctx['tag']}.xlsx")
                            await _tg_reply_document(
                                u.message,
                                io.BytesIO(xbytes),
                                fname,
                                caption="Excel: per-campaign sheets + UNIFIED + LEADS AUDIT",
                            )
            except Exception:
                log.exception("tg analytics Excel export failed")

        except Exception:
            log.exception("telegram analytics_cmd")
            await u.message.reply_text(_tg_err_reply())


    async def _tg_post_init(application: Any):
        try:
            await application.bot.set_my_commands([
                BotCommand("start", "Welcome and commands"),
                BotCommand("help", "Help and /ask [profile] usage"),
                BotCommand("analytics", "Reports; /analytics lists profiles"),
                BotCommand("ask", "Analytics: /ask [profile] your question"),
                BotCommand("balance", "Reoon verification credits"),
            ])
        except Exception:
            log.exception("telegram set_my_commands failed (bot still starts)")

    app = (
        Application.builder()
        .token(c["telegram_token"])
        .post_init(_tg_post_init)
        .build()
    )
    app.add_handler(CommandHandler("start",   start))
    app.add_handler(CommandHandler("help",    help_cmd))
    app.add_handler(CommandHandler("balance", balance_cmd))
    app.add_handler(CommandHandler("analytics",analytics_cmd))
    app.add_handler(CommandHandler("ask",     ask_cmd))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_doc))
    # Unknown slash commands → snarky response
    app.add_handler(MessageHandler(filters.COMMAND, unknown_cmd))
    # Plain text: greetings → snarky, anything else → AI ask
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, plain_text))

    console.print(Panel(f"[bold green]Mailclaw Bot running[/] — health: :{_hport}/health",border_style="green"))
    if not (c.get("telegram_allowed_users") or []):
        log.warning("telegram_allowed_users is empty — bot accepts any Telegram user. Set allowlist in production.")
    _bot_started["ok"] = True
    app.run_polling(allowed_updates=Update.ALL_TYPES)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FULL PIPELINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def cmd_run(_args):
    c=cfg_load()
    if not c["reoon_keys"] and not c["instantly_clients"]:
        run_onboarding(); c=cfg_load()

    console.print(Panel(
        "[bold]Which stages do you want to run?[/]\n\n"
        "[dim]All stages are optional. You can run them in any combination.[/]",
        border_style="dim"))

    stages=questionary.checkbox("Select stages to run:",choices=[
        questionary.Choice("🗺   Column Mapping  (auto-map + confirm)",   value="map",     checked=True),
        questionary.Choice("✦   AI Enrichment   (Gemini / Haiku / GPT)", value="enrich",  checked=False),
        questionary.Choice("✓   Email Verify     (Reoon + key rotation)", value="verify",  checked=True),
        questionary.Choice("↑   Upload to Instantly (V2 API)",            value="upload",  checked=True),
    ],style=Q_STYLE).ask()

    if not stages:
        console.print("[yellow]No stages selected.[/]"); return

    # Pick CSV
    console.print("\n[bold cyan]Select CSV[/]")
    csv_path=pick_csv()
    if not csv_path: return
    fp=csv_fp(csv_path); st=state_load(fp)
    rows,cols=csv_read(csv_path)
    console.print(f"[green]✓[/] {len(rows):,} rows, {len(cols)} columns — [yellow]{csv_path.name}[/]")
    if not rows: console.print("[red]CSV is empty.[/]"); return

    # Column mapping
    col_map=st.get("column_map",{})
    if "map" in stages or not col_map:
        console.print("\n[bold cyan]Column Mapping[/]")
        col_map=do_column_mapping(cols,rows,saved_map=col_map if col_map else None)
        st["column_map"]=col_map; state_save(fp,st)
    email_col=col_map.get("email")
    if not email_col:
        console.print("[red]No email column mapped — cannot proceed.[/]"); return

    working_rows=rows
    enrich_before=True   # default; overridden below if both enrich+verify are selected

    # Enrichment (optional, before or after verify)
    if "enrich" in stages:
        if not any(v for v in c.get("model_keys",{}).values()):
            console.print("[yellow]No AI model keys configured — skipping enrichment. Run mailclaw config to add.[/]")
        else:
            console.print("\n[bold cyan]AI Enrichment[/]")
            # Ask order if verify is also selected
            if "verify" in stages:
                order=questionary.select("Run enrichment:",choices=[
                    "Before verification (enrich all raw leads)",
                    "After verification (enrich safe leads only — saves AI credits)",
                ],style=Q_STYLE).ask()
                enrich_before="Before" in (order or "")
            else: enrich_before=True

            if enrich_before:
                # Pick profile using resolve_profile_for_client (or let user pick)
                all_p=profiles_all()
                p_name=questionary.select("Enrichment profile:",choices=list(all_p.keys()),style=Q_STYLE).ask()
                profile=all_p.get(p_name) if p_name else None
                working_rows,cost,out_p=stage_enrich(csv_path,working_rows,col_map,fp,st,profile)
                st=state_load(fp); console.print(f"[green]✓[/] Enriched — ${cost:.4f}")

    # Verification
    verify_result=None
    if "verify" in stages:
        console.print("\n[bold cyan]Email Verification[/]")
        verify_result=stage_verify(csv_path,working_rows,email_col,fp,st,col_map)
        st=state_load(fp)

        if verify_result:
            proc=verify_result["processed"]
            console.print(Panel(
                f"[green]✓ Safe:[/]      {len(proc['safe']):,}\n"
                f"[yellow]⚡ Catch-all:[/] {len(proc['catchall']):,}\n"
                f"[red]✗ Dropped:[/]   {len(proc['dropped']):,}\n\n"
                f"[cyan]Gmail:[/] {len(proc['by_esp'].get('gmail',[]))}  "
                f"[blue]Outlook:[/] {len(proc['by_esp'].get('outlook',[]))}  "
                f"[magenta]Yahoo:[/] {len(proc['by_esp'].get('yahoo',[]))}  "
                f"[white]Other:[/] {len(proc['by_esp'].get('other',[]))}",
                title="[bold]Verification Results[/]",border_style="cyan"))

            # Enrich after verify (safe leads only)
            if "enrich" in stages and not enrich_before:
                console.print("\n[bold cyan]AI Enrichment (safe leads only)[/]")
                all_p=profiles_all()
                p_name=questionary.select("Profile:",choices=list(all_p.keys()),style=Q_STYLE).ask()
                profile=all_p.get(p_name) if p_name else None
                safe_merged=verify_result.get("safe_merged",[])
                if safe_merged and profile:
                    tmp_p=csv_path.parent/out_name(csv_path.stem,"safe_for_enrich")
                    csv_write(safe_merged,tmp_p)
                    enr,cost,out_p=stage_enrich(tmp_p,safe_merged,col_map,fp,st,profile)
                    st=state_load(fp); working_rows=enr
                    console.print(f"[green]✓[/] Enriched safe leads — ${cost:.4f}")

    # Upload
    if "upload" in stages:
        console.print("\n[bold cyan]Upload to Instantly[/]")

        # Decide what to upload
        upload_rows=working_rows
        if verify_result:
            proc=verify_result["processed"]
            outfiles=verify_result.get("outfiles",{})
            file_options=[p.name for p in outfiles.values() if p.exists()]
            file_options+=["Upload raw working list (no verify filter)","Skip upload"]
            if file_options:
                pick=questionary.select("Which leads to upload?",choices=file_options,style=Q_STYLE).ask()
                if not pick or "Skip" in pick:
                    pass
                elif "raw working" in pick:
                    upload_rows=working_rows
                else:
                    # Load the chosen file
                    chosen_path=next((p for p in outfiles.values() if p.name==pick),None)
                    if chosen_path:
                        upload_rows,_=csv_read(chosen_path)
                        console.print(f"[green]✓[/] Loaded {len(upload_rows):,} rows from {chosen_path.name}")

        if upload_rows:
            stage_upload(csv_path,upload_rows,col_map,fp,st)
            st=state_load(fp)

    # Final summary
    console.print("\n[bold yellow]━━━ DONE ━━━[/]")
    t=Table(box=box.SIMPLE_HEAVY,show_header=False,padding=(0,2))
    t.add_column("k",style="dim"); t.add_column("v",style="bold green")
    t.add_row("Input",csv_path.name)
    if verify_result:
        p=verify_result["processed"]
        t.add_row("Safe leads",str(len(p["safe"])))
        t.add_row("Catch-all",str(len(p["catchall"])))
    if st.get("enrich",{}).get("total_cost"):
        t.add_row("AI cost",f"${st['enrich']['total_cost']:.4f}")
    if st.get("upload",{}).get("count"):
        t.add_row("Uploaded",str(st["upload"]["count"]))
        t.add_row("Campaign",st["upload"].get("campaign","?"))
    t.add_row("State file",str(STATE_DIR/f"{fp}.json"))
    console.print(Panel(t,border_style="yellow"))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ENTRY POINT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ANALYTICS PROFILE SYSTEM
#  ~/.mailclaw/analytics/<profile_name>.json
#
#  Each profile controls:
#  - which Instantly client(s) to pull from
#  - optional date range filter
#  - benchmarks (bounce %, reply %, etc.) for pass/fail colouring
#  - manual fields (meetings_attended, follow_ups, sales) for the report
#
#  Usage:
#    python mailclaw.py analytics               # interactive
#    python mailclaw.py analytics --profile will
#    python mailclaw.py analytics --profile will --start 2026-01-01 --end 2026-03-31
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ANALYTICS_PROFILE_TEMPLATE = {
    "name":         "default",
    "display_name": "Default Analytics Profile",
    # Which Instantly client(s) to pull campaigns from. [] = ask each run.
    "client_names": [],
    # Benchmarks used for red/green colouring. Set to null to skip colouring.
    "benchmarks": {
        "bounce_rate":         0.05,   # 5%  — max acceptable bounce rate
        "unsubscribe_rate":    0.02,   # 2%  — max acceptable unsub rate
        "reply_rate":          0.05,   # 5%  — total reply rate (incl. OOO) target
        "human_reply_rate":    0.01,   # 1%  — human-only reply rate target
        "positive_reply_rate": 0.20,   # 20% — opp rate (opps / contacted)
        "open_rate":           0.50,   # 50%
        "meeting_book_rate":   0.25,   # 25% — meetings / opps
        "meeting_attend_rate": 0.80,   # 80% — attended / booked
    },
    # Fields that can't come from the API — you fill these in at report time
    # or pre-set them here (0 = ask each run)
    "manual_fields": {
        "meetings_attended": None,   # None = ask each run
        "follow_ups":        None,
        "sales_usd":         None,
    },
    # Campaign filter: if set, only campaigns whose name contains these strings
    # will be shown in the picker. Useful for "only show GD_ campaigns".
    "campaign_name_filter": "",
}

def _analytics_profiles_from_env() -> Dict[str, dict]:
    """ANALYTICS_PROFILE_<NAME>=<json> — used on Railway without writable disk."""
    out: Dict[str, dict] = {}
    for k, v in os.environ.items():
        if not k.startswith("ANALYTICS_PROFILE_"):
            continue
        try:
            pname = k[len("ANALYTICS_PROFILE_"):].lower()
            prof = json.loads(v)
            out[pname] = prof
        except Exception:
            pass
    return out

def analytics_profile_load(name: str) -> Optional[dict]:
    n = name.lower().strip()
    envp = _analytics_profiles_from_env().get(n)
    if envp is not None:
        return envp
    p = ANALYTICS_DIR / f"{name}.json"
    if p.exists():
        try:
            return json.loads(p.read_text())
        except Exception:
            return None
    return None

def analytics_profile_save(prof: dict):
    if use_env_config():
        log.debug("analytics_profile_save skipped (env-only config)")
        return
    (ANALYTICS_DIR / f"{prof['name']}.json").write_text(json.dumps(prof, indent=2))

def analytics_profiles_all() -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    if not use_env_config():
        for f in sorted(ANALYTICS_DIR.glob("*.json")):
            try:
                p = json.loads(f.read_text())
                out[p.get("name", f.stem)] = p
            except Exception:
                pass
    for pname, prof in _analytics_profiles_from_env().items():
        key = prof.get("name", pname)
        out[key] = prof
    return out

def _pct(n,d)->str:
    """Format as percentage string, or — if denominator is 0."""
    if not d: return "—"
    return f"{n/d*100:.1f}%"

def _bench_color(val:float,bench:Optional[float],higher_is_better:bool=False)->str:
    """Return rich color based on whether val beats benchmark."""
    if bench is None or val is None: return "white"
    if higher_is_better: return "green" if val>=bench else "red"
    else:                return "green" if val<=bench else "red"

VARIANT_NAMES_STEP = {"0": "A", "1": "B", "2": "C", "3": "D", "4": "E"}

def analytics_step_sort_key(x: dict):
    try:
        return (int(x["step"] or 99), str(x["variant"] or ""))
    except Exception:
        return (99, str(x.get("variant") or ""))

def analytics_extract_metrics(ov: Optional[dict], steps: Optional[List[dict]] = None) -> dict:
    """Same metrics dict as CLI analytics (overview + optional per-step data)."""
    if not ov:
        return {k: 0 for k in [
            "emails_sent", "contacted", "bounced", "unsubscribed", "replies", "auto_replies", "total_replies",
            "interested", "mtg_booked", "mtg_completed", "closed", "negative", "total_opportunities",
            "bounce_rate", "unsub_rate", "reply_rate", "total_reply_rate", "int_rate", "mtg_book_rate", "mtg_att_rate",
            "human_reply_rate", "opp_rate", "contacted_api",
        ]}
    es = ov.get("emails_sent_count", 0) or 0
    ct_api = ov.get("contacted_count", 0) or 0
    if steps:
        step0s = [s for s in steps if (s.get("step") in (0, "0") or str(s.get("step", "")) == "0") and s.get("sent", 0) > 0]
        ct = sum(s.get("sent", 0) for s in step0s) if step0s else ct_api
    else:
        ct = ct_api
    bn = ov.get("bounced_count", 0) or 0
    us = ov.get("unsubscribed_count", 0) or 0
    rp = ov.get("reply_count_unique", 0) or 0
    ar = ov.get("reply_count_automatic_unique", 0) or 0
    rp_total = rp + ar
    intr = ov.get("total_interested", 0) or 0
    mb = ov.get("total_meeting_booked", 0) or 0
    mc = ov.get("total_meeting_completed", 0) or 0
    cl = ov.get("total_closed", 0) or 0
    total_opps = ov.get("total_opportunities", 0) or 0
    neg = max(0, rp - total_opps)
    return dict(
        emails_sent=es, contacted=ct, contacted_api=ct_api,
        bounced=bn, unsubscribed=us,
        replies=rp, auto_replies=ar, total_replies=rp_total,
        interested=intr, mtg_booked=mb, total_opportunities=total_opps,
        mtg_completed=mc, closed=cl, negative=neg,
        bounce_rate=bn / ct if ct else 0,
        unsub_rate=us / ct if ct else 0,
        reply_rate=rp / ct if ct else 0,
        human_reply_rate=rp / ct if ct else 0,
        total_reply_rate=rp_total / ct if ct else 0,
        opp_rate=total_opps / ct if ct else 0,
        # Interested % = interested leads ÷ contacted (same basis as reply_rate / opp_rate).
        # Not interested ÷ total_opportunities (that can exceed 100% after CRM dedup vs API opps).
        int_rate=intr / ct if ct else 0,
        mtg_book_rate=mb / total_opps if total_opps else 0,
        mtg_att_rate=mc / mb if mb else 0,
    )

def analytics_dr_pretty(start_date: str, end_date: str) -> str:
    def _fmt(d: str) -> str:
        try:
            return datetime.strptime(d, "%Y-%m-%d").strftime("%d-%b-%Y")
        except Exception:
            return d or "All time"
    return f"{_fmt(start_date)} to {_fmt(end_date or datetime.now().strftime('%Y-%m-%d'))}"

def analytics_compute_export_context(
    inst: "Instantly",
    selected_ids: List[str],
    sel_map: Dict[str, Any],
    camp_overviews: Dict[str, Optional[dict]],
    camp_steps: Dict[str, List[dict]],
    start_date: str,
    end_date: str,
    bench: dict,
    manual: Optional[dict] = None,
) -> Optional[dict]:
    """
    Build the same data structures as CLI analytics export: unified totals (deduped),
    step aggregates, positive-leads audit rows — for Excel/CSV.
    """
    from collections import defaultdict
    manual = manual or {}
    try:
        follow_ups = float(manual.get("follow_ups") or 0)
    except Exception:
        follow_ups = 0.0
    try:
        sales_usd = float(manual.get("sales_usd") or 0)
    except Exception:
        sales_usd = 0.0

    if not selected_ids:
        return None

    subseq_ids: set = set()
    for cid in selected_ids:
        camp = sel_map.get(cid, {})
        n_lower = camp.get("name", "").lower()
        if "subsequence" in n_lower or "subseq" in n_lower:
            subseq_ids.add(cid)
    primary_ids = [cid for cid in selected_ids if cid not in subseq_ids]
    subseq_count = len(subseq_ids)

    step_agg: Dict[tuple, dict] = {}
    for cid in selected_ids:
        for s in camp_steps.get(cid, []):
            if not s.get("sent", 0):
                continue
            key = (s.get("step"), s.get("variant"))
            if key not in step_agg:
                step_agg[key] = {
                    "step": s.get("step"), "variant": s.get("variant"),
                    "sent": 0, "unique_opened": 0, "unique_replies": 0, "unique_clicks": 0,
                    "replies_automatic": 0, "unique_opportunities": 0,
                }
            ag = step_agg[key]
            ag["sent"] += s.get("sent", 0)
            ag["unique_opened"] += s.get("unique_opened", 0)
            ag["unique_replies"] += s.get("unique_replies", 0)
            ag["unique_clicks"] += s.get("unique_clicks", 0)
            ag["replies_automatic"] += s.get("replies_automatic", 0)
            ag["unique_opportunities"] += s.get("unique_opportunities", 0)
    sorted_steps = [s for s in sorted(step_agg.values(), key=analytics_step_sort_key) if s.get("sent", 0) > 0]

    _primary_step_agg: Dict[tuple, dict] = {}
    for cid in primary_ids:
        for s in camp_steps.get(cid, []):
            if not s.get("sent", 0):
                continue
            key = (s.get("step"), s.get("variant"))
            if key not in _primary_step_agg:
                _primary_step_agg[key] = {
                    "step": s.get("step"), "variant": s.get("variant"),
                    "sent": 0, "unique_opened": 0, "unique_replies": 0, "unique_clicks": 0,
                    "replies_automatic": 0, "unique_opportunities": 0,
                }
            ag = _primary_step_agg[key]
            for f_ in ("sent", "unique_opened", "unique_replies", "unique_clicks", "replies_automatic", "unique_opportunities"):
                ag[f_] += s.get(f_, 0)
    primary_sorted_steps = [s for s in sorted(_primary_step_agg.values(), key=analytics_step_sort_key) if s.get("sent", 0) > 0]

    overview = inst.get_analytics_overview(selected_ids, start_date, end_date)
    _camp_metrics_list = [
        analytics_extract_metrics(camp_overviews.get(cid), steps=camp_steps.get(cid, []))
        for cid in primary_ids if camp_overviews.get(cid)
    ]

    def _usum(key):
        return sum(m.get(key, 0) for m in _camp_metrics_list)

    _ct_sum = _usum("contacted")
    _es_sum = _usum("emails_sent")
    _rp_sum = _usum("replies")
    _ar_sum = _usum("auto_replies")
    _rp_total = _rp_sum + _ar_sum
    _bn_sum = _usum("bounced")
    _us_sum = _usum("unsubscribed")
    _intr_sum = _usum("interested")
    _mb_sum = _usum("mtg_booked")
    _mc_sum = _usum("mtg_completed")
    _cl_sum = _usum("closed")
    _opps_sum = _usum("total_opportunities")
    _neg_sum = max(0, _rp_sum - _opps_sum)
    unified = dict(
        emails_sent=_es_sum, contacted=_ct_sum, contacted_api=_usum("contacted_api"),
        bounced=_bn_sum, unsubscribed=_us_sum,
        replies=_rp_sum, auto_replies=_ar_sum, total_replies=_rp_total,
        interested=_intr_sum, mtg_booked=_mb_sum, total_opportunities=_opps_sum,
        mtg_completed=_mc_sum, closed=_cl_sum, negative=_neg_sum,
        bounce_rate=_bn_sum / _ct_sum if _ct_sum else 0,
        unsub_rate=_us_sum / _ct_sum if _ct_sum else 0,
        reply_rate=_rp_sum / _ct_sum if _ct_sum else 0,
        human_reply_rate=_rp_sum / _ct_sum if _ct_sum else 0,
        total_reply_rate=_rp_total / _ct_sum if _ct_sum else 0,
        opp_rate=_opps_sum / _ct_sum if _ct_sum else 0,
        int_rate=_intr_sum / _ct_sum if _ct_sum else 0,
        mtg_book_rate=_mb_sum / _opps_sum if _opps_sum else 0,
        mtg_att_rate=_mc_sum / _mb_sum if _mb_sum else 0,
    )

    ov_ = overview or {}
    max_exp_total = sum([
        ov_.get("total_interested", 0) or 0,
        ov_.get("total_meeting_booked", 0) or 0,
        ov_.get("total_meeting_completed", 0) or 0,
        ov_.get("total_closed", 0) or 0,
    ]) or 100
    all_positive_leads = inst.list_positive_leads(max_expected=max_exp_total)
    STATUS_RANK = {1: 1, 2: 2, 3: 3, 4: 4}
    email_best_status: Dict[str, int] = {}
    for lead in all_positive_leads:
        em = lead.get("email", "")
        if not em:
            continue
        st = lead.get("lt_interest_status", 0)
        if em not in email_best_status or STATUS_RANK.get(st, 0) > STATUS_RANK.get(email_best_status[em], 0):
            email_best_status[em] = st
    corr_interested = sum(1 for s in email_best_status.values() if s == 1)
    corr_booked = sum(1 for s in email_best_status.values() if s == 2)
    corr_completed = sum(1 for s in email_best_status.values() if s == 3)
    corr_closed = sum(1 for s in email_best_status.values() if s == 4)
    api_booked = unified.get("mtg_booked", 0)
    api_int = unified.get("interested", 0)
    _opps_canonical = unified.get("total_opportunities", 0)
    _ct = unified.get("contacted", 0)
    unified["mtg_booked"] = corr_booked
    unified["interested"] = corr_interested
    unified["mtg_completed"] = corr_completed
    unified["closed"] = corr_closed
    unified["human_reply_rate"] = unified.get("reply_rate", 0)
    unified["opp_rate"] = _opps_canonical / _ct if _ct else 0
    unified["int_rate"] = corr_interested / _ct if _ct else 0
    unified["mtg_book_rate"] = corr_booked / _opps_canonical if _opps_canonical else 0
    unified["mtg_att_rate"] = corr_completed / corr_booked if corr_booked else 0
    unified["negative"] = max(0, unified.get("replies", 0) - _opps_canonical)

    MAX_EMAIL_FETCHES = 40
    total_pos = len(all_positive_leads)
    fetch_timestamps = total_pos > 0 and total_pos <= MAX_EMAIL_FETCHES
    STATUS_LABELS = {1: "Interested", 2: "Meeting Booked", 3: "Meeting Completed", 4: "Won/Closed"}
    selected_set = set(selected_ids)
    camp_pos_leads: Dict[str, List[dict]] = defaultdict(list)
    for lead in all_positive_leads:
        lcid = lead.get("campaign") or lead.get("campaign_id", "")
        if lcid in selected_set:
            camp_pos_leads[lcid].append(lead)

    email_appearances: Dict[str, list] = defaultdict(list)
    for cid in selected_ids:
        cname = sel_map[cid].get("name", "Unnamed")
        for lead in camp_pos_leads.get(cid, []):
            email = lead.get("email", "")
            if not email:
                continue
            status = STATUS_LABELS.get(lead.get("lt_interest_status", 0), "?")
            ts_raw = lead.get("timestamp_updated") or lead.get("timestamp_created", "")
            try:
                ts_crm = datetime.strptime(ts_raw[:19], "%Y-%m-%dT%H:%M:%S").strftime("%d-%b-%Y %H:%M")
            except Exception:
                ts_crm = ts_raw[:16] if ts_raw else "—"
            first_reply_ts = "—"
            is_auto = "—"
            if fetch_timestamps:
                fr = inst.get_first_reply(cid, email)
                if fr:
                    try:
                        first_reply_ts = datetime.strptime(
                            fr["timestamp"][:19], "%Y-%m-%dT%H:%M:%S"
                        ).strftime("%d-%b-%Y %H:%M")
                    except Exception:
                        first_reply_ts = fr["timestamp"][:16]
                    is_auto = "🤖 Auto" if fr["is_auto_reply"] else "👤 Human"
            email_appearances[email].append({
                "campaign": cname,
                "status": status,
                "ts_crm": ts_crm,
                "ts_reply": first_reply_ts,
                "reply_type": is_auto,
                "cid": cid,
                "lead_id": lead.get("id", ""),
            })

    dupes = {e: v for e, v in email_appearances.items() if len(v) > 1}

    tag = datetime.now().strftime("%d_%b_%y").lower()
    return {
        "bench": bench,
        "follow_ups": follow_ups,
        "sales_usd": sales_usd,
        "selected_ids": selected_ids,
        "sel_map": sel_map,
        "camp_overviews": camp_overviews,
        "camp_steps": camp_steps,
        "unified": unified,
        "sorted_steps": sorted_steps,
        "primary_sorted_steps": primary_sorted_steps,
        "email_appearances": dict(email_appearances),
        "dupes": dupes,
        "fetch_timestamps": fetch_timestamps,
        "subseq_count": subseq_count,
        "tag": tag,
        "overview": overview,
    }

def analytics_workbook_to_bytes(ctx: dict, client_meta: dict, dr_pretty: str) -> Optional[bytes]:
    """
    Same Excel as CLI analytics: one sheet per campaign, UNIFIED, and LEADS AUDIT.
    ctx must match analytics_compute_export_context output (or equivalent keys).
    Returns None if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    b_ = ctx["bench"]
    follow_ups = float(ctx.get("follow_ups") or 0)
    sales_usd = float(ctx.get("sales_usd") or 0)
    selected_ids = ctx["selected_ids"]
    sel_map = ctx["sel_map"]
    camp_overviews = ctx["camp_overviews"]
    camp_steps = ctx["camp_steps"]
    unified = ctx["unified"]
    sorted_steps = ctx["sorted_steps"]
    email_appearances = ctx["email_appearances"]
    dupes = ctx["dupes"]
    fetch_timestamps = ctx["fetch_timestamps"]

    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    C_CAMPAIGN_BG = "1A1A2E"
    C_CAMPAIGN_FG = "E8E8FF"
    C_UNIFIED_BG = "0F3460"
    C_UNIFIED_FG = "FFD700"
    C_COL_HDR_BG = "16213E"
    C_COL_HDR_FG = "FFFFFF"
    C_STEP_HDR_BG = "1E3A5F"
    C_STEP_HDR_FG = "FFFFFF"
    C_GREEN = "1B5E20"
    C_RED = "B71C1C"
    C_ALT = "F0F4FF"
    C_WHITE = "FFFFFF"

    def _fill(hex_):
        return PatternFill("solid", start_color=hex_, end_color=hex_)

    def _font(bold=False, color="000000", size=11, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

    def _align(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def _cell(ws, row, col, val, bold=False, fg="000000", bg=None, size=11,
              italic=False, h="left", num_fmt=None, border=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=bold, color=fg, size=size, italic=italic)
        c.alignment = _align(h=h)
        if bg:
            c.fill = _fill(bg)
        if num_fmt:
            c.number_format = num_fmt
        if border:
            c.border = BORDER
        return c

    def _merge_header(ws, row, text, fg, bg, size, cols=4):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = _font(bold=True, color=fg, size=size)
        c.fill = _fill(bg)
        c.alignment = _align(h="center")
        c.border = BORDER

    def good(rv, key, hi=True):
        bv = b_.get(key)
        if bv is None or rv is None:
            return None
        return (rv >= bv) if hi else (rv <= bv)

    def _all_sections():
        for cid in selected_ids:
            cname = sel_map[cid].get("name", "Unnamed")
            cov = camp_overviews.get(cid)
            if not cov:
                continue
            yield (
                cname,
                analytics_extract_metrics(cov, steps=camp_steps.get(cid, [])),
                camp_steps.get(cid, []),
                False,
            )
        yield f"TOTAL — {len(selected_ids)} Campaigns", unified, sorted_steps, True

    def _write_metrics(ws, row, label, m, fu=0, su=0, is_unified=False):
        hdr_bg = C_UNIFIED_BG if is_unified else C_CAMPAIGN_BG
        hdr_fg = C_UNIFIED_FG if is_unified else C_CAMPAIGN_FG
        hdr_size = 14 if is_unified else 13
        _merge_header(ws, row, label, hdr_fg, hdr_bg, hdr_size, cols=4)
        row += 1

        if is_unified:
            ascii_art = (
                f"  MAILCLAW  ·  ANALYTICS REPORT\n"
                f"  Client: {client_meta['name']}  ·  {dr_pretty}\n"
                f"  {len(selected_ids)} campaigns selected"
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            c = ws.cell(row=row, column=1, value=ascii_art)
            c.font = _font(bold=False, color=C_UNIFIED_FG, size=9, italic=True)
            c.fill = _fill("0A1628")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 48
            row += 1

        for ci, h in enumerate(["Metric", "Benchmark", "Actual", "Rate"], 1):
            _cell(ws, row, ci, h, bold=True, fg=C_COL_HDR_FG, bg=C_COL_HDR_BG, size=10, h="center")
        row += 1

        ct = m.get("contacted", 0)
        es = m.get("emails_sent", 0)

        def metric(name, actual, rate_str, is_good=None, indent=0, is_sub=False, bench_str=""):
            nonlocal row
            alt = (row % 2 == 0)
            bg = C_ALT if alt else C_WHITE
            pad = "    " * indent + name
            bold = not is_sub
            _cell(ws, row, 1, pad, bold=bold, bg=bg, size=10)
            _cell(ws, row, 2, bench_str, bg=bg, size=10, h="center", italic=True, fg="888888")
            _cell(ws, row, 3, actual, bold=bold, bg=bg, size=10, h="right")
            rate_fg = C_GREEN if is_good is True else (C_RED if is_good is False else "555555")
            _cell(ws, row, 4, rate_str, bold=False, fg=rate_fg, bg=bg, size=10, h="right")
            row += 1

        def bstr(key):
            bv = b_.get(key)
            return f"{bv * 100:.0f}%" if bv is not None else ""

        metric("Unique Leads Contacted", f"{ct:,}", "—")
        metric("Emails Sent", f"{es:,}", "—")
        metric("Bounce", f"{m['bounced']:,}", f"{m['bounce_rate'] * 100:.1f}%",
               good(m["bounce_rate"], "bounce_rate", hi=False),
               indent=1, is_sub=True, bench_str=bstr("bounce_rate"))
        metric("Unsubscribed", f"{m['unsubscribed']:,}", f"{m['unsub_rate'] * 100:.1f}%",
               good(m["unsub_rate"], "unsubscribe_rate", hi=False),
               indent=1, is_sub=True, bench_str=bstr("unsubscribe_rate"))
        ws.append(["", "", "", ""])
        row += 1
        metric("Total Replies (incl. auto)", f"{m['total_replies']:,}", f"{m['total_reply_rate'] * 100:.1f}%",
               good(m["total_reply_rate"], "reply_rate", hi=True), bench_str=bstr("reply_rate"))
        metric("Human Replies", f"{m['replies']:,}", f"{m['reply_rate'] * 100:.1f}%",
               indent=1, is_sub=True)
        metric("Auto-Replies (OOO)", f"{m['auto_replies']:,}", "",
               indent=1, is_sub=True)
        metric("Interested", f"{m['interested']:,}", f"{m['int_rate'] * 100:.1f}%",
               good(m["int_rate"], "positive_reply_rate", hi=True),
               indent=1, is_sub=True, bench_str=bstr("positive_reply_rate"))
        metric("Negative", f"{m['negative']:,}", "", indent=1, is_sub=True)
        ws.append(["", "", "", ""])
        row += 1
        metric("📅 Meetings Booked", f"{m['mtg_booked']:,}", f"{m['mtg_book_rate'] * 100:.1f}%",
               good(m["mtg_book_rate"], "meeting_book_rate", hi=True), bench_str=bstr("meeting_book_rate"))
        metric("✅ Meetings Attended (API)", f"{m['mtg_completed']:,}", f"{m['mtg_att_rate'] * 100:.1f}%",
               good(m["mtg_att_rate"], "meeting_attend_rate", hi=True),
               indent=1, is_sub=True, bench_str=bstr("meeting_attend_rate"))
        metric("🏆 Deals Closed", f"{m['closed']:,}", "")
        ws.append(["", "", "", ""])
        row += 1
        metric("Follow Ups", f"{int(fu):,}", f"{fu / ct * 100:.1f}%" if ct else "")
        metric("💰 Sales ($)", f"${int(su):,}", "")
        ws.append(["", "", "", ""])
        row += 1
        return row

    def _write_steps(ws, row, label, steps):
        rows_ = [s for s in steps if s.get("sent", 0) > 0]
        if not rows_:
            return row
        _merge_header(ws, row, f"Step Breakdown — {label}",
                      C_STEP_HDR_FG, C_STEP_HDR_BG, 11, cols=7)
        row += 1
        for ci, h in enumerate(["Email", "Variant", "Sent", "Replied", "Reply%", "Positive", "Pos%"], 1):
            _cell(ws, row, ci, h, bold=True, fg=C_COL_HDR_FG, bg=C_COL_HDR_BG, size=10, h="center")
        row += 1
        for s in sorted(rows_, key=analytics_step_sort_key):
            sent = s["sent"]
            rep = s["unique_replies"]
            pos = s.get("unique_opportunities", 0) or 0
            try:
                en = f"Email {int(s['step']) + 1}"
            except Exception:
                en = str(s.get("step", "?"))
            vl = VARIANT_NAMES_STEP.get(str(s.get("variant") or ""), str(s.get("variant") or ""))
            alt = (row % 2 == 0)
            bg = C_ALT if alt else C_WHITE
            rr = rep / sent if sent else 0
            pr = pos / rep if rep else 0
            r_good = good(rr, "reply_rate", hi=True)
            p_good = good(pr, "positive_reply_rate", hi=True)
            r_fg = C_GREEN if r_good else (C_RED if r_good is False else "000000")
            p_fg = C_GREEN if p_good else (C_RED if p_good is False else "555555")
            for ci_, v_ in enumerate([en, vl, f"{sent:,}", f"{rep:,}"], 1):
                _cell(ws, row, ci_, v_, bg=bg, size=10)
            _cell(ws, row, 5, f"{rr * 100:.1f}%", fg=r_fg, bg=bg, size=10, h="right")
            _cell(ws, row, 6, f"{pos:,}", bg=bg, size=10, h="right")
            _cell(ws, row, 7, f"{pr * 100:.1f}%" if rep else "—", fg=p_fg, bg=bg, size=10, h="right")
            row += 1
        ws.append(["", "", "", "", "", "", ""])
        row += 1
        return row

    MAX_EMAIL_FETCHES = 40
    wb = Workbook()
    first = True
    for camp_label, m_, steps_, is_u in _all_sections():
        sname = ("UNIFIED" if is_u else camp_label[:28].strip().replace("/", "_").replace("\\", "_"))
        if first:
            ws = wb.active
            ws.title = sname
            first = False
        else:
            ws = wb.create_sheet(title=sname)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10
        ws.freeze_panes = "A3"
        row = 1
        row = _write_metrics(ws, row, camp_label, m_,
                             fu=follow_ups if is_u else 0,
                             su=sales_usd if is_u else 0,
                             is_unified=is_u)
        row = _write_steps(ws, row, camp_label, steps_)

    ws_audit = wb.create_sheet(title="LEADS AUDIT")
    ws_audit.column_dimensions["A"].width = 34
    ws_audit.column_dimensions["B"].width = 20
    ws_audit.column_dimensions["C"].width = 20
    ws_audit.column_dimensions["D"].width = 12
    ws_audit.column_dimensions["E"].width = 20
    ws_audit.column_dimensions["F"].width = 42
    ws_audit.column_dimensions["G"].width = 10

    ws_audit.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ah = ws_audit.cell(row=1, column=1,
                       value=f"POSITIVE LEADS AUDIT  ·  {client_meta['name']}  ·  {dr_pretty}")
    ah.font = _font(bold=True, color="FFD700", size=13)
    ah.fill = _fill("4A0080")
    ah.alignment = _align(h="center")
    ah.border = BORDER

    ws_audit.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    sh = ws_audit.cell(row=2, column=1,
                       value=(
                           f"{'⚠  ' + str(len(dupes)) + ' duplicate email(s) found!' if dupes else '✓ No duplicates found'}  |  "
                           f"{'First Reply timestamps from emails API' if fetch_timestamps else 'First Reply: not fetched (>' + str(MAX_EMAIL_FETCHES) + ' leads)'}"
                       ))
    sh.font = _font(bold=True, color="FFFFFF" if not dupes else "FF4444", size=11)
    sh.fill = _fill("2D0050" if not dupes else "5C0000")
    sh.alignment = _align(h="center")
    sh.border = BORDER

    for ci_, h_ in enumerate(["Email", "Status", "First Reply", "Reply Type", "CRM Updated", "Campaign", "Dupe?"], 1):
        _cell(ws_audit, 3, ci_, h_, bold=True, fg=C_COL_HDR_FG, bg=C_COL_HDR_BG, size=10, h="center")

    audit_row = 4
    STATUS_COLORS_XL = {
        "Interested": ("1B5E20", "E8F5E9"),
        "Meeting Booked": ("E65100", "FFF3E0"),
        "Meeting Completed": ("0D47A1", "E3F2FD"),
        "Won/Closed": ("4A148C", "F3E5F5"),
    }
    for email_, entries_ in sorted(email_appearances.items()):
        for occ in entries_:
            is_dupe = email_ in dupes
            fg_s, bg_s = STATUS_COLORS_XL.get(occ["status"], ("000000", "FFFFFF"))
            row_bg = "FFE8E8" if is_dupe else ("F0F4FF" if audit_row % 2 == 0 else "FFFFFF")
            _cell(ws_audit, audit_row, 1, email_, bg=row_bg, size=10, bold=is_dupe)
            _cell(ws_audit, audit_row, 2, occ["status"], fg=fg_s, bg=bg_s, size=10, bold=True, h="center")
            _cell(ws_audit, audit_row, 3, occ["ts_reply"], bg=row_bg, size=9, h="center")
            rt_color = "888888" if occ["reply_type"] == "—" else ("CC0000" if "Auto" in occ["reply_type"] else "1B5E20")
            _cell(ws_audit, audit_row, 4, occ["reply_type"], fg=rt_color, bg=row_bg, size=9, h="center")
            _cell(ws_audit, audit_row, 5, occ["ts_crm"], bg=row_bg, size=9, h="center")
            _cell(ws_audit, audit_row, 6, occ["campaign"], bg=row_bg, size=9)
            _cell(ws_audit, audit_row, 7, "⚠ DUPE" if is_dupe else "",
                  fg="FFFFFF" if is_dupe else "888888",
                  bg="CC0000" if is_dupe else row_bg,
                  bold=is_dupe, size=10, h="center")
            audit_row += 1

    ws_audit.freeze_panes = "A4"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def analytics_csv_to_bytes(ctx: dict, client_meta: dict, dr_pretty: str) -> bytes:
    """Same CSV as CLI analytics export (metrics + steps + leads audit)."""
    b_ = ctx["bench"]
    follow_ups = float(ctx.get("follow_ups") or 0)
    sales_usd = float(ctx.get("sales_usd") or 0)
    selected_ids = ctx["selected_ids"]
    sel_map = ctx["sel_map"]
    camp_overviews = ctx["camp_overviews"]
    camp_steps = ctx["camp_steps"]
    unified = ctx["unified"]
    sorted_steps = ctx["sorted_steps"]
    email_appearances = ctx["email_appearances"]
    dupes = ctx["dupes"]
    fetch_timestamps = ctx["fetch_timestamps"]
    MAX_EMAIL_FETCHES = 40

    rows_out: List[List[Any]] = []

    def bstr2(key):
        bv = b_.get(key)
        return f"{bv * 100:.0f}%" if bv is not None else ""

    def csv_block(label: str, m: dict, fu: float = 0, su: float = 0):
        ct = m.get("contacted", 0)
        rows_out.append([label, "", "", ""])
        rows_out.append(["Metric", "Benchmark", "Actual", "Rate"])
        rows_out.append(["Unique Leads Contacted", "", ct, ""])
        rows_out.append(["Emails Sent", "", m["emails_sent"], ""])
        rows_out.append(["Bounce", bstr2("bounce_rate"), m["bounced"], f"{m['bounce_rate'] * 100:.2f}%"])
        rows_out.append(["Unsubscribed", bstr2("unsubscribe_rate"), m["unsubscribed"], f"{m['unsub_rate'] * 100:.2f}%"])
        rows_out.append(["", "", "", ""])
        rows_out.append(["Total Replies", bstr2("reply_rate"), m["total_replies"], f"{m['total_reply_rate'] * 100:.2f}%"])
        rows_out.append(["  Human Replies", "", m["replies"], f"{m['reply_rate'] * 100:.2f}%"])
        rows_out.append(["  Auto-Replies", "", m["auto_replies"], ""])
        rows_out.append(["  Interested", bstr2("positive_reply_rate"), m["interested"], f"{m['int_rate'] * 100:.2f}%"])
        rows_out.append(["  Negative", "", m["negative"], ""])
        rows_out.append(["", "", "", ""])
        rows_out.append(["Meetings Booked", bstr2("meeting_book_rate"), m["mtg_booked"], f"{m['mtg_book_rate'] * 100:.2f}%"])
        rows_out.append(["Meetings Attended", "", m["mtg_completed"], f"{m['mtg_att_rate'] * 100:.2f}%"])
        rows_out.append(["Deals Closed", "", m["closed"], ""])
        rows_out.append(["Follow Ups", "", int(fu), f"{fu / ct * 100:.2f}%" if ct else ""])
        rows_out.append(["Sales ($)", "", f"${int(su)}", ""])
        rows_out.append(["", "", "", ""])

    def csv_steps2(label: str, steps: list):
        rows_ = [s for s in steps if s.get("sent", 0) > 0]
        if not rows_:
            return
        rows_out.append([f"Steps: {label}", "", "", "", "", "", ""])
        rows_out.append(["Email", "Variant", "Sent", "Replied", "Reply%", "Positive", "Pos%"])
        for s in sorted(rows_, key=analytics_step_sort_key):
            sent = s["sent"]
            rep = s["unique_replies"]
            pos = s.get("unique_opportunities", 0) or 0
            try:
                en = f"Email {int(s['step']) + 1}"
            except Exception:
                en = str(s.get("step", "?"))
            vl = VARIANT_NAMES_STEP.get(str(s.get("variant") or ""), str(s.get("variant") or ""))
            rows_out.append([
                en, vl, sent, rep,
                f"{rep / sent * 100:.2f}%" if sent else "", pos,
                f"{pos / rep * 100:.2f}%" if rep else "",
            ])
        rows_out.append(["", "", "", "", "", "", ""])

    def _all_sections():
        for cid in selected_ids:
            cname = sel_map[cid].get("name", "Unnamed")
            cov = camp_overviews.get(cid)
            if not cov:
                continue
            yield cname, analytics_extract_metrics(cov, steps=camp_steps.get(cid, [])), camp_steps.get(cid, []), False
        yield f"TOTAL — {len(selected_ids)} Campaigns", unified, sorted_steps, True

    for camp_label, m_, steps_, is_u in _all_sections():
        csv_block(camp_label, m_, fu=follow_ups if is_u else 0, su=sales_usd if is_u else 0)
        csv_steps2(camp_label, steps_)
    rows_out.append(["POSITIVE LEADS AUDIT", "", "", "", "", "", ""])
    rows_out.append([
        f"{'⚠ ' + str(len(dupes)) + ' duplicates' if dupes else 'No duplicates'}",
        f"{'First Reply fetched via emails API' if fetch_timestamps else 'First Reply: not fetched (>' + str(MAX_EMAIL_FETCHES) + ' leads)'}",
        "", "", "", "", "",
    ])
    rows_out.append(["Email", "Status", "First Reply", "Reply Type", "CRM Updated", "Campaign", "Dupe?"])
    for em_, entries_ in sorted(email_appearances.items()):
        for occ in entries_:
            rows_out.append([
                em_, occ["status"], occ["ts_reply"], occ["reply_type"],
                occ["ts_crm"], occ["campaign"],
                "DUPE" if em_ in dupes else "",
            ])
    rows_out.append(["", "", "", "", "", "", ""])

    buf = io.StringIO()
    import csv as _csv
    w = _csv.writer(buf)
    w.writerows(rows_out)
    return buf.getvalue().encode("utf-8")


def cmd_analytics(_args):
    """
    Analytics report command.
    Fetches Instantly overview + per-step analytics, renders a report matching
    Campaign_Tracking_2026.xlsx format, and exports CSV for pasting.
    """
    c=cfg_load()

    # ── Parse args ──────────────────────────────────────────────────────────
    profile_name=getattr(_args,"profile",None)
    start_date  =getattr(_args,"start",None) or ""
    end_date    =getattr(_args,"end",None)   or ""

    # ── Load or pick profile ─────────────────────────────────────────────────
    if profile_name:
        prof=analytics_profile_load(profile_name)
        if not prof:
            console.print(f"[red]Profile '{profile_name}' not found in {ANALYTICS_DIR}[/]")
            console.print(f"[dim]Existing: {list(analytics_profiles_all().keys())}[/]")
            return
    else:
        all_p=analytics_profiles_all()
        if not all_p:
            console.print("[yellow]No analytics profiles yet. Creating one now.[/]")
            pname=questionary.text("Profile name (e.g. will, brad):",style=Q_STYLE).ask()
            if not pname: return
            prof=dict(ANALYTICS_PROFILE_TEMPLATE); prof["name"]=pname.strip()
            analytics_profile_save(prof)
        else:
            choices=list(all_p.keys())+["➕  Create new profile"]
            pick=questionary.select("Analytics profile:",choices=choices,style=Q_STYLE).ask()
            if not pick: return
            if "Create" in pick:
                pname=questionary.text("Profile name:",style=Q_STYLE).ask()
                if not pname: return
                prof=dict(ANALYTICS_PROFILE_TEMPLATE); prof["name"]=pname.strip()
                analytics_profile_save(prof)
            else:
                prof=all_p[pick]

    bench=prof.get("benchmarks",{})

    # ── Pick Instantly client ────────────────────────────────────────────────
    client_names=prof.get("client_names",[])
    cfg_clients=c.get("instantly_clients",[])
    if not cfg_clients:
        console.print("[red]No Instantly clients configured. Run: mailclaw clients[/]"); return
    if client_names:
        client_meta=next((cl for cl in cfg_clients if cl["name"]==client_names[0]),None)
        if not client_meta:
            console.print(f"[red]Client '{client_names[0]}' not found in config.[/]"); return
    else:
        choices_cl=[f"{cl['name']}  [{cl['key'][:10]}…]" for cl in cfg_clients]
        pick_cl=questionary.select("Instantly client:",choices=choices_cl,style=Q_STYLE).ask()
        if not pick_cl: return
        client_meta=cfg_clients[choices_cl.index(pick_cl)]

    inst=Instantly(client_meta["key"],client_meta["name"],c.get("rate_limit_delay",0.12))

    # ── Fetch all campaigns ──────────────────────────────────────────────────
    console.print(f"\n[cyan]→[/] Fetching all campaigns for [bold]{client_meta['name']}[/]…")
    all_camps=inst.list_all_campaigns()
    if not all_camps:
        console.print("[red]No campaigns found.[/]"); return

    name_filter=prof.get("campaign_name_filter","")
    if name_filter:
        all_camps=[camp for camp in all_camps
                   if name_filter.lower() in camp.get("name","").lower()]
        console.print(f"[dim]Filter '{name_filter}': {len(all_camps)} campaigns[/]")

    all_camps.sort(key=lambda x:x.get("timestamp_updated",""),reverse=True)

    STATUS_MAP={0:"Draft",1:"Active",2:"Paused",3:"Completed",
                4:"Subseq",-1:"Unhealthy",-2:"BounceProtect",-99:"Suspended"}
    # ── Smart date picker ────────────────────────────────────────────────────
    from datetime import datetime as _dt, timedelta as _td, date as _date
    if not start_date and not end_date:
        granularity=questionary.select(
            "Time period:",
            choices=[
                questionary.Choice("📅  All time",         value="all"),
                questionary.Choice("📆  Daily",            value="daily"),
                questionary.Choice("📅  Weekly",           value="weekly"),
                questionary.Choice("🗓   Monthly",          value="monthly"),
                questionary.Choice("✏️   Custom (enter dates)", value="custom"),
            ], style=Q_STYLE).ask()

        today=_date.today()

        if granularity=="all":
            start_date=""; end_date=""

        elif granularity=="daily":
            # Pick from last 14 days
            day_choices=[]
            for i in range(14):
                d=today-_td(days=i)
                label="Today" if i==0 else ("Yesterday" if i==1 else d.strftime("%a %d %b"))
                day_choices.append(questionary.Choice(label, value=d.strftime("%Y-%m-%d")))
            picked=questionary.select("Which day?",choices=day_choices,style=Q_STYLE).ask()
            start_date=picked; end_date=picked

        elif granularity=="weekly":
            # Show last 12 weeks, Mon–Sun
            week_choices=[]
            # find most recent Monday
            days_since_mon=today.weekday()
            this_mon=today-_td(days=days_since_mon)
            for i in range(12):
                wmon=this_mon-_td(weeks=i)
                wsun=wmon+_td(days=6)
                wnum=wmon.isocalendar()[1]
                yr=wmon.year
                label=(f"This week  ({wmon.strftime('%d %b')} – {wsun.strftime('%d %b')})" if i==0
                       else f"Week {wnum}  {yr}  ({wmon.strftime('%d %b')} – {wsun.strftime('%d %b')})")
                week_choices.append(questionary.Choice(
                    label, value=(wmon.strftime("%Y-%m-%d"), wsun.strftime("%Y-%m-%d"))))
            picked=questionary.select("Which week?",choices=week_choices,style=Q_STYLE).ask()
            start_date,end_date=picked

        elif granularity=="monthly":
            # Show last 12 months
            import calendar
            month_choices=[]
            for i in range(12):
                # go back i months
                yr=today.year; mo=today.month-i
                while mo<1: mo+=12; yr-=1
                last_day=calendar.monthrange(yr,mo)[1]
                ms=f"{yr}-{mo:02d}-01"
                me=f"{yr}-{mo:02d}-{last_day:02d}"
                label=("This month  " if i==0 else "")+_dt(yr,mo,1).strftime("%B %Y")
                month_choices.append(questionary.Choice(label, value=(ms,me)))
            picked=questionary.select("Which month?",choices=month_choices,style=Q_STYLE).ask()
            start_date,end_date=picked

        else:  # custom
            sd=questionary.text("Start date (YYYY-MM-DD, or Enter for all time):",
                                style=Q_STYLE).ask()
            ed=questionary.text("End date   (YYYY-MM-DD, or Enter for today):",
                                style=Q_STYLE).ask()
            start_date=sd.strip() if sd and sd.strip() else ""
            end_date  =ed.strip() if ed and ed.strip() else ""

    # ── Auto-select campaigns by activity in date range ─────────────────────
    # Step 1: Pre-filter by updated_at to drop obviously dead campaigns
    from datetime import datetime as _dt, timedelta as _td
    if start_date:
        try:
            cutoff=(_dt.strptime(start_date,"%Y-%m-%d")-_td(days=7)).strftime("%Y-%m-%d")
            pre_filtered=[c for c in all_camps
                          if (c.get("timestamp_updated","") or "9999") >= cutoff]
        except: pre_filtered=all_camps
    else:
        pre_filtered=all_camps   # all-time: include everything

    console.print(f"[cyan]→[/] Checking activity for {len(pre_filtered)} campaigns"
                  f"{' (pre-filtered by date)' if len(pre_filtered)<len(all_camps) else ''}…")

    # Step 2: Fetch overview for each candidate — auto-include where emails_sent > 0
    # These are the same calls we'd make later; we just use them for selection too.
    # We store results to reuse and avoid double-fetching.
    _auto_overviews:Dict[str,Optional[dict]]={}
    selected_ids=[]; sel_map={}
    excluded_zero=[]
    for camp in pre_filtered:
        cid=camp["id"]
        ov_=inst.get_analytics_overview([cid],start_date,end_date)
        _auto_overviews[cid]=ov_
        sent=ov_.get("emails_sent_count",0) if ov_ else 0
        if sent and sent>0:
            selected_ids.append(cid)
            sel_map[cid]=camp
        else:
            excluded_zero.append(camp.get("name","?")[:50])

    if not selected_ids:
        console.print("[yellow]No campaigns had activity in this date range.[/]"); return

    # Step 3: Show what was auto-selected, allow quick exclusion
    t=Table(title=f"[bold]Auto-selected — {client_meta['name']}[/]",
            box=box.ROUNDED,show_lines=True)
    t.add_column("#",style="dim",width=4)
    t.add_column("Name",style="cyan",max_width=52)
    t.add_column("Status",style="yellow",width=12)
    t.add_column("Updated",style="dim",width=10)
    # Detect follow-up campaigns by name — Instantly status=4 means the PARENT
    # campaign is running a subsequence, NOT that this campaign itself is one.
    subseq_ids=set()
    for i,cid in enumerate(selected_ids,1):
        camp=sel_map[cid]
        n_lower=camp.get("name","").lower()
        is_sub="subsequence" in n_lower or "subseq" in n_lower
        if is_sub: subseq_ids.add(cid)
        sl=STATUS_MAP.get(camp.get("status",0),str(camp.get("status","")))
        t.add_row(str(i),("⚡ " if is_sub else "")+camp.get("name","Unnamed"),sl,
                  camp.get("timestamp_updated","")[:10])
    console.print(t)
    if excluded_zero:
        console.print(f"[dim]Excluded (0 emails in period): "
                      f"{', '.join(excluded_zero[:5])}"
                      f"{f' +{len(excluded_zero)-5} more' if len(excluded_zero)>5 else ''}[/]")

    excl_raw=questionary.text(
        f"Press Enter to use all {len(selected_ids)} campaigns, "
        f"or type numbers to exclude (e.g. 2,4):",
        default="",style=Q_STYLE).ask()
    if excl_raw and excl_raw.strip():
        try:
            excl_nums={int(x.strip()) for x in excl_raw.split(",") if x.strip().isdigit()}
            selected_ids=[cid for i,cid in enumerate(selected_ids,1) if i not in excl_nums]
            sel_map={cid:sel_map[cid] for cid in selected_ids}
        except: pass

    if not selected_ids:
        console.print("[yellow]No campaigns selected.[/]"); return
    console.print(f"\n[green]✓[/] {len(selected_ids)} campaigns selected.")

    # ── Manual fields (only what API can't provide) ──────────────────────────
    # NOTE: meetings_attended = total_meeting_completed from API — no manual needed
    # Only follow_ups and sales_usd are truly manual
    manual=prof.get("manual_fields",{})
    def get_manual(key,label):
        v=manual.get(key)
        if v is not None: return float(v)
        raw=questionary.text(f"{label} (not in API — enter manually, Enter=0):",
                             default="0",style=Q_STYLE).ask()
        try: return float(raw) if raw else 0.0
        except: return 0.0

    follow_ups=get_manual("follow_ups","Follow Ups sent")
    sales_usd =get_manual("sales_usd", "Sales ($)")

    # ── Fetch data ───────────────────────────────────────────────────────────
    console.print(f"\n[cyan]→[/] Fetching unified overview…")
    overview=inst.get_analytics_overview(selected_ids,start_date,end_date)
    if not overview:
        console.print("[red]Failed to fetch analytics overview.[/]"); return

    # ── Per-campaign step data ────────────────────────────────────────────────
    console.print(f"[cyan]→[/] Fetching per-step analytics ({len(selected_ids)} campaigns)…")
    camp_steps:Dict[str,List[dict]]={}
    step_agg:Dict[tuple,dict]={}
    for cid in selected_ids:
        steps=inst.get_analytics_steps(cid,start_date,end_date)
        camp_steps[cid]=steps
        for s in steps:
            if not s.get("sent",0): continue
            key=(s.get("step"),s.get("variant"))
            if key not in step_agg:
                step_agg[key]={"step":s.get("step"),"variant":s.get("variant"),
                               "sent":0,"unique_opened":0,"unique_replies":0,
                               "unique_clicks":0,"replies_automatic":0,
                               "unique_opportunities":0}
            ag=step_agg[key]
            ag["sent"]                +=s.get("sent",0)
            ag["unique_opened"]       +=s.get("unique_opened",0)
            ag["unique_replies"]      +=s.get("unique_replies",0)
            ag["unique_clicks"]       +=s.get("unique_clicks",0)
            ag["replies_automatic"]   +=s.get("replies_automatic",0)
            ag["unique_opportunities"]+=s.get("unique_opportunities",0)

    def _step_key(x):
        try: return (int(x["step"] or 99), str(x["variant"] or ""))
        except: return (99, str(x.get("variant") or ""))
    sorted_steps=[s for s in sorted(step_agg.values(),key=_step_key) if s.get("sent",0)>0]
    # Note: step_agg is built from selected_ids (all campaigns), but we want
    # unified step table to reflect primary campaigns only.
    # Rebuild step_agg for primary campaigns only.
    _primary_step_agg:Dict[tuple,dict]={}
    for cid in primary_ids:
        for s in camp_steps.get(cid,[]):
            if not s.get("sent",0): continue
            key=(s.get("step"),s.get("variant"))
            if key not in _primary_step_agg:
                _primary_step_agg[key]={"step":s.get("step"),"variant":s.get("variant"),
                                        "sent":0,"unique_opened":0,"unique_replies":0,
                                        "unique_clicks":0,"replies_automatic":0,
                                        "unique_opportunities":0}
            ag=_primary_step_agg[key]
            for f_ in ("sent","unique_opened","unique_replies","unique_clicks",
                       "replies_automatic","unique_opportunities"):
                ag[f_]+=s.get(f_,0)
    primary_sorted_steps=[s for s in sorted(_primary_step_agg.values(),key=_step_key)
                          if s.get("sent",0)>0]
    VARIANT_NAMES={"0":"A","1":"B","2":"C","3":"D","4":"E"}

    # ── Per-campaign overview ────────────────────────────────────────────────
    console.print(f"[cyan]→[/] Fetching per-campaign overview ({len(selected_ids)} calls)…")
    camp_overviews:Dict[str,Optional[dict]]={}
    for cid in selected_ids:
        # Reuse overview already fetched during auto-selection if available
        camp_overviews[cid]=_auto_overviews.get(cid) or inst.get_analytics_overview([cid],start_date,end_date)

    # Fetch positive-status leads per campaign for dupe audit
    console.print(f"[cyan]→[/] Fetching positive leads for dupe audit ({len(selected_ids)} campaigns)…")
    STATUS_LABELS={1:"Interested",2:"Meeting Booked",3:"Meeting Completed",4:"Won/Closed"}

    # Build unified summing ONLY non-subsequence campaigns.
    # Subsequence campaigns re-use leads already in primary campaigns, so including
    # them inflates unique leads, reply counts, and CRM totals.
    # They are shown per-campaign but excluded from unified sums.
    primary_ids=[cid for cid in selected_ids if cid not in subseq_ids]
    subseq_count=len(subseq_ids)
    if subseq_count:
        console.print(f"[dim]  ⚡ {subseq_count} subsequence campaign(s) shown per-campaign "
                      f"but excluded from unified totals (leads overlap with primary campaigns)[/]")

    unified_raw=analytics_extract_metrics(overview)  # fallback reference only
    _camp_metrics_list=[analytics_extract_metrics(camp_overviews.get(cid),steps=camp_steps.get(cid,[]))
                        for cid in primary_ids if camp_overviews.get(cid)]
    def _usum(key): return sum(m.get(key,0) for m in _camp_metrics_list)
    _ct_sum=_usum("contacted"); _es_sum=_usum("emails_sent")
    _rp_sum=_usum("replies");   _ar_sum=_usum("auto_replies")
    _rp_total=_rp_sum+_ar_sum
    _bn_sum=_usum("bounced");   _us_sum=_usum("unsubscribed")
    _intr_sum=_usum("interested"); _mb_sum=_usum("mtg_booked")
    _mc_sum=_usum("mtg_completed"); _cl_sum=_usum("closed")
    _opps_sum=_usum("total_opportunities")
    _neg_sum=max(0,_rp_sum-_opps_sum)
    unified=dict(
        emails_sent=_es_sum, contacted=_ct_sum, contacted_api=_usum("contacted_api"),
        bounced=_bn_sum, unsubscribed=_us_sum,
        replies=_rp_sum, auto_replies=_ar_sum, total_replies=_rp_total,
        interested=_intr_sum, mtg_booked=_mb_sum, total_opportunities=_opps_sum,
        mtg_completed=_mc_sum, closed=_cl_sum, negative=_neg_sum,
        bounce_rate      =_bn_sum/_ct_sum     if _ct_sum else 0,
        unsub_rate       =_us_sum/_ct_sum     if _ct_sum else 0,
        reply_rate       =_rp_sum/_ct_sum       if _ct_sum    else 0,
        human_reply_rate =_rp_sum/_ct_sum       if _ct_sum    else 0,
        total_reply_rate =_rp_total/_ct_sum     if _ct_sum    else 0,
        opp_rate         =_opps_sum/_ct_sum     if _ct_sum    else 0,
        int_rate         =_intr_sum/_ct_sum   if _ct_sum    else 0,
        mtg_book_rate    =_mb_sum/_opps_sum     if _opps_sum  else 0,
        mtg_att_rate     =_mc_sum/_mb_sum       if _mb_sum    else 0,
    )

    MEETING_ART="""
  ╔══════════════════════════════╗
  ║  📅  MEETINGS BOOKED         ║
  ║                              ║
  ║    ___   ___   ___           ║
  ║   |   | |   | |   |  x{n:<3}   ║
  ║   |___| |___| |___|          ║
  ║   CHAIR CHAIR CHAIR          ║
  ╚══════════════════════════════╝"""
    WINS_ART=(
  r"""
  ╔══════════════════════════════╗
  ║  🏆  DEALS CLOSED            ║
  ║                              ║
  ║        ___                   ║
  ║       /   \     x{n}        ║
  ║      | WIN |                 ║
  ║       \___/                  ║
  ╚══════════════════════════════╝""").replace("{n}","{n:<3}")
    FIRE_ART="""
  ╔══════════════════════════════╗
  ║  🔥  ON FIRE                 ║
  ║   Reply rate crushing it!    ║
  ║                              ║
  ║     )  )  )  )               ║
  ║    ( ( ( ( (                 ║
  ║     ) ) ) ) )                ║
  ╚══════════════════════════════╝"""

    def make_overview_table(title:str, m:dict, fu:float=0, su:float=0):
        tb=Table(title=f"[bold]{title}[/]",box=box.ROUNDED,show_lines=True,padding=(0,1))
        tb.add_column("Metric",    style="bold",   width=26)
        tb.add_column("Benchmark", justify="right",style="dim",width=10)
        tb.add_column("Actual",    justify="right",width=12)
        tb.add_column("Rate",      justify="right",width=10)
        tb.add_column("✓/✗",      justify="center",width=4)
        def bstr(key):
            b=bench.get(key); return f"{b*100:.0f}%" if b is not None else "—"
        def icon(rv,key,hi=True):
            b=bench.get(key)
            if b is None or rv is None: return "[dim]—[/]"
            return "[green]✓[/]" if ((rv>=b) if hi else (rv<=b)) else "[red]✗[/]"
        def rcol(rv,key,hi=True):
            b=bench.get(key)
            if rv is None: return "[dim]—[/]"
            col="white" if b is None else ("green" if ((rv>=b) if hi else (rv<=b)) else "red")
            return f"[{col}]{rv*100:.1f}%[/]"
        def ar(label,actual,rv,key,hi=True,astr=None):
            tb.add_row(label,bstr(key),astr or f"{int(actual):,}",rcol(rv,key,hi),icon(rv,key,hi))
        ct=m.get("contacted",0); es=m.get("emails_sent",0)
        total_opps=m.get("total_opportunities",0)
        mb=m.get("mtg_booked",0); mc=m.get("mtg_completed",0); cl=m.get("closed",0)
        intr=m.get("interested",0)

        # ── Reach ──────────────────────────────────────────────────────────────
        tb.add_row("[cyan bold]Unique Leads Contacted[/]","—",f"{ct:,}","—","—")
        tb.add_row("[dim]  Emails Sent (all steps)[/]","—",f"[dim]{es:,}[/]","—","—")
        ar("  Bounced",     m["bounced"],     m["bounce_rate"],   "bounce_rate",     hi=False)
        ar("  Unsubscribed",m["unsubscribed"],m["unsub_rate"],    "unsubscribe_rate",hi=False)

        # ── Replies ─────────────────────────────────────────────────────────────
        tb.add_row("","","","","")
        ar("[bold]Total Replies[/]",  m["total_replies"], m["total_reply_rate"], "reply_rate", hi=True)
        ar("  👤 Human",              m["replies"],       m["reply_rate"],       "human_reply_rate", hi=True)
        ooo_pct=f"[dim]{m['auto_replies']/m['total_replies']*100:.0f}% of replies[/]" if m['total_replies'] else "[dim]—[/]"
        tb.add_row("  🤖 OOO / Auto","—",f"{m['auto_replies']:,}",ooo_pct,"[dim]—[/]")

        # ── Pipeline ────────────────────────────────────────────────────────────
        # Opportunities = total_opportunities from API (unique leads with any positive status)
        # CRM stages use expand_crm_events=false (default): each lead counted at their
        # first CRM touch only. Stages are mutually exclusive per the API semantics.
        # For unified table these are patched to deduped values from the lead audit.
        tb.add_row("","","","","")
        ar("[bold yellow]🎯 Opportunities[/]", total_opps, m["opp_rate"],
           "positive_reply_rate", hi=True,
           astr=f"[bold yellow]{total_opps:,}[/]")
        tb.add_row("  ❌ Not Interested","—",f"{m['negative']:,}",
                   f"[dim]{m['negative']/m['replies']*100:.0f}% of replies[/]" if m['replies'] else "[dim]—[/]",
                   "[dim]—[/]")
        tb.add_row("","","","","")
        ar("[bold]📅 Meetings Booked[/]", mb, m["mtg_book_rate"], "meeting_book_rate", hi=True)
        if mc or m.get("mtg_att_rate",0):
            ar("  ✅ Attended", mc, m["mtg_att_rate"], "meeting_attend_rate", hi=True)
        if cl:
            tb.add_row("[bold green]🏆 Deals Closed[/]","—",f"[bold green]{cl:,}[/]","—","—")
        if intr:
            tb.add_row("  ✅ Interested","—",f"{intr:,}",
                       f"[dim]{intr/ct*100:.2f}% of leads[/]" if ct else "[dim]—[/]",
                       "[dim]—[/]")

        # ── Manual ─────────────────────────────────────────────────────────────
        if fu or su:
            tb.add_row("","","","","")
            if fu: tb.add_row("Follow Ups","—",f"{int(fu):,}","—","—")
            if su: tb.add_row("💰 Sales ($)","—",f"${int(su):,}","—","—")
        return tb

    def make_step_table(title:str, steps:list):
        rows=[s for s in steps if s.get("sent",0)>0]
        if not rows: return None
        st=Table(title=f"[bold]{title}[/]",box=box.ROUNDED,show_lines=True,padding=(0,1))
        st.add_column("Email",   style="cyan bold",width=9)
        st.add_column("Variant", style="dim",      width=7)
        st.add_column("Sent",    justify="right",  width=8)
        st.add_column("Replied", justify="right",  width=8)
        st.add_column("Reply%",  justify="right",  width=8)
        st.add_column("Positive",justify="right",  width=9)
        st.add_column("Pos%",    justify="right",  width=7)
        for s in sorted(rows,key=_step_key):
            sent=s["sent"]; rep=s["unique_replies"]
            pos=s.get("unique_opportunities",0) or 0
            try:   en=f"Email {int(s['step'])+1}"
            except: en=str(s.get("step","?"))
            vl=VARIANT_NAMES.get(str(s.get("variant") or ""),str(s.get("variant") or "—"))
            rr=rep/sent if sent else 0
            pr=pos/rep  if rep  else 0
            r_col="green" if rr>=bench.get("reply_rate",0.05) else "red"
            p_col="green" if pr>=bench.get("positive_reply_rate",0.25) else ("dim" if not rep else "red")
            st.add_row(en, vl, f"{sent:,}", f"{rep:,}",
                       f"[{r_col}]{rr*100:.1f}%[/]" if sent else "—",
                       f"{pos:,}",
                       f"[{p_col}]{pr*100:.1f}%[/]" if rep else "[dim]—[/]")
        return st

    def show_art(m:dict):
        arts=[]
        if m.get("mtg_booked",0)>0: arts.append(MEETING_ART.format(n=m["mtg_booked"]))
        if m.get("closed",0)>0:     arts.append(WINS_ART.format(n=m["closed"]))
        if m.get("reply_rate",0)>=bench.get("reply_rate",0.05)*1.5: arts.append(FIRE_ART)
        for art in arts:
            console.print(Panel(f"[bold yellow]{art}[/]",border_style="yellow",expand=False))
    # ═══ OUTPUT ════════════════════════════════════════════════════════════════
    console.print()
    console.rule(f"[bold yellow]  ANALYTICS REPORT — {client_meta['name'].upper()}  [/]",style="yellow")
    dr=f"{start_date or 'all time'} → {end_date or 'today'}"
    console.print(f"[dim]Date range: {dr}  |  {len(selected_ids)} campaigns[/]")

    # Per-campaign
    console.print()
    console.rule("[bold cyan]  PER-CAMPAIGN BREAKDOWN  [/]",style="cyan")
    for cid in selected_ids:
        camp=sel_map[cid]
        cname=camp.get("name","Unnamed")
        is_sub=cid in subseq_ids
        cov=camp_overviews.get(cid)
        if not cov:
            console.print(f"[red]  ✗ No data: {cname}[/]"); continue
        cm=analytics_extract_metrics(cov, steps=camp_steps.get(cid,[]))
        console.print()
        title_prefix="[dim]⚡ SUBSEQUENCE  — [/]" if is_sub else ""
        console.print(make_overview_table(title_prefix+cname[:70], cm))
        if is_sub:
            console.print("[dim]  ↑ Excluded from unified totals — leads overlap with primary campaigns[/]")
        cst=make_step_table(f"Steps — {cname[:50]}", camp_steps.get(cid,[]))
        if cst:
            console.print(cst)
        show_art(cm)

    # ── Fetch positive leads BEFORE unified display so we can patch numbers ────
    from collections import defaultdict
    console.print(f"[cyan]→[/] Fetching positive leads for dedup…")
    ov_=overview or {}
    max_exp_total=sum([ov_.get("total_interested",0) or 0,
                       ov_.get("total_meeting_booked",0) or 0,
                       ov_.get("total_meeting_completed",0) or 0,
                       ov_.get("total_closed",0) or 0]) or 100
    all_positive_leads=inst.list_positive_leads(max_expected=max_exp_total)
    console.print(f"[dim]  {len(all_positive_leads)} workspace positive leads[/]")

    # Compute corrected numbers — highest CRM status per unique email
    STATUS_RANK={1:1,2:2,3:3,4:4}
    email_best_status:Dict[str,int]={}
    email_best_lead:Dict[str,dict]={}
    for lead in all_positive_leads:
        em=lead.get("email","")
        if not em: continue
        st=lead.get("lt_interest_status",0)
        if em not in email_best_status or STATUS_RANK.get(st,0)>STATUS_RANK.get(email_best_status[em],0):
            email_best_status[em]=st; email_best_lead[em]=lead

    corr_interested = sum(1 for s in email_best_status.values() if s==1)
    corr_booked     = sum(1 for s in email_best_status.values() if s==2)
    corr_completed  = sum(1 for s in email_best_status.values() if s==3)
    corr_closed     = sum(1 for s in email_best_status.values() if s==4)
    corr_total_pos  = len(email_best_status)

    api_booked = unified.get("mtg_booked",0)
    api_int    = unified.get("interested",0)
    dupe_booked= api_booked - corr_booked
    dupe_int   = api_int    - corr_interested

    # Patch unified CRM numbers with cross-campaign deduped values from lead audit.
    # total_opportunities stays as _opps_sum (Instantly's per-campaign dedup is correct).
    # We only patch mtg_booked and interested because those are the ones inflated by
    # leads appearing in both a primary campaign AND a positive reply subsequence.
    _opps_canonical = unified.get("total_opportunities",0)  # keep Instantly's value
    _ct  = unified.get("contacted",0)
    unified["mtg_booked"]    = corr_booked
    unified["interested"]    = corr_interested
    unified["mtg_completed"] = corr_completed
    unified["closed"]        = corr_closed
    unified["human_reply_rate"] = unified.get("reply_rate",0)
    unified["opp_rate"]      = _opps_canonical/_ct           if _ct              else 0
    unified["int_rate"]      = corr_interested/_ct           if _ct              else 0
    unified["mtg_book_rate"] = corr_booked/_opps_canonical    if _opps_canonical  else 0
    unified["mtg_att_rate"]  = corr_completed/corr_booked     if corr_booked      else 0
    unified["negative"]      = max(0, unified.get("replies",0) - _opps_canonical)

    # ── Unified display (single, correct table) ──────────────────────────────
    console.print()
    note_parts=[]
    if subseq_count: note_parts.append(f"⚡ {subseq_count} subsequence(s) excluded from totals")
    if dupe_booked>0: note_parts.append(f"📅 Booked {api_booked}→{corr_booked} ({dupe_booked} cross-campaign dupe{'s' if dupe_booked!=1 else ''} removed)")
    if dupe_int>0:    note_parts.append(f"✅ Interested {api_int}→{corr_interested}")
    console.rule("[bold yellow]  UNIFIED SUMMARY — ALL CAMPAIGNS  [/]",style="yellow")
    if note_parts:
        console.print(Panel("[dim]"+("  ·  ".join(note_parts))+f"  ·  Source: {corr_total_pos} unique leads[/]",
                            border_style="dim",expand=False))
    console.print()
    console.print(make_overview_table(
        f"TOTAL — {len(selected_ids)} Campaigns  ({dr})", unified, follow_ups, sales_usd))
    ust=make_step_table("Unified Step Totals  [dim](primary campaigns only)[/dim]",
                        primary_sorted_steps)
    if ust:
        console.print(); console.print(ust)
    show_art(unified)

    # Filter positive leads to selected campaigns for audit table
    selected_set=set(selected_ids)
    camp_pos_leads:Dict[str,List[dict]]=defaultdict(list)
    for lead in all_positive_leads:
        lcid=lead.get("campaign") or lead.get("campaign_id","")
        if lcid in selected_set:
            camp_pos_leads[lcid].append(lead)

    total_pos=len(all_positive_leads)
    STATUS_LABELS={1:"Interested",2:"Meeting Booked",3:"Meeting Completed",4:"Won/Closed"}

    email_appearances:Dict[str,list]=defaultdict(list)
    MAX_EMAIL_FETCHES=40
    fetch_timestamps=(total_pos>0 and total_pos<=MAX_EMAIL_FETCHES)
    if not fetch_timestamps and total_pos>0:
        console.print(f"[dim]  Skipping first-reply fetch — {total_pos} positive leads "
                      f"exceeds {MAX_EMAIL_FETCHES} (emails API: 20 req/min)[/]")

    for cid in selected_ids:
        cname=sel_map[cid].get("name","Unnamed")
        for lead in camp_pos_leads.get(cid,[]):
            email=lead.get("email","")
            if not email: continue
            status=STATUS_LABELS.get(lead.get("lt_interest_status",0),"?")
            # CRM status timestamp — when they were marked
            ts_raw=lead.get("timestamp_updated") or lead.get("timestamp_created","")
            try: ts_crm=datetime.strptime(ts_raw[:19],"%Y-%m-%dT%H:%M:%S").strftime("%d-%b-%Y %H:%M")
            except: ts_crm=ts_raw[:16] if ts_raw else "—"

            # First reply timestamp from emails API
            first_reply_ts="—"; is_auto="—"
            if fetch_timestamps:
                console.print(f"[dim]  → First reply: {email[:35]}…[/]",end="\r")
                fr=inst.get_first_reply(cid,email)
                if fr:
                    try:
                        first_reply_ts=datetime.strptime(
                            fr["timestamp"][:19],"%Y-%m-%dT%H:%M:%S"
                        ).strftime("%d-%b-%Y %H:%M")
                    except: first_reply_ts=fr["timestamp"][:16]
                    is_auto="🤖 Auto" if fr["is_auto_reply"] else "👤 Human"

            email_appearances[email].append({
                "campaign":   cname,
                "status":     status,
                "ts_crm":     ts_crm,       # when CRM status was set
                "ts_reply":   first_reply_ts, # first inbound email timestamp
                "reply_type": is_auto,        # Human or Auto
                "cid":        cid,
                "lead_id":    lead.get("id",""),
            })

    if fetch_timestamps:
        console.print(" "*80,end="\r")  # clear progress line

    dupes={e:v for e,v in email_appearances.items() if len(v)>1}

    console.print()
    console.rule("[bold magenta]  POSITIVE LEADS AUDIT  [/]",style="magenta")
    console.print(f"[dim]Leads with Interested/Meeting Booked/Attended/Won status "
                  f"across {len(selected_ids)} campaigns. "
                  f"{'First Reply timestamps fetched from emails API.' if fetch_timestamps else 'First Reply timestamps skipped (too many leads for rate limit).'}[/]")
    console.print()

    # Leads table
    lt=Table(title="[bold]All Positive Leads[/]",box=box.ROUNDED,show_lines=True,padding=(0,1))
    lt.add_column("Email",         style="cyan",    width=34)
    lt.add_column("Status",        style="bold",    width=18)
    lt.add_column("First Reply",   justify="right", width=18)
    lt.add_column("Reply Type",    justify="center",width=10)
    lt.add_column("CRM Updated",   justify="right", width=18)
    lt.add_column("Campaign",      style="dim",     max_width=35)
    lt.add_column("⚠",            justify="center",width=4)

    STATUS_COLOR={"Interested":"green","Meeting Booked":"yellow",
                  "Meeting Completed":"cyan","Won/Closed":"bold green"}
    all_pos_sorted=sorted(
        [(e,entry) for e,entries in email_appearances.items() for entry in entries],
        key=lambda x:(x[0],x[1]["ts_reply"] if x[1]["ts_reply"]!="—" else x[1]["ts_crm"]))

    for email,entry in all_pos_sorted:
        is_dupe=email in dupes
        sc=STATUS_COLOR.get(entry["status"],"white")
        lt.add_row(
            email,
            f"[{sc}]{entry['status']}[/]",
            entry["ts_reply"],
            f"[dim]{entry['reply_type']}[/]",
            entry["ts_crm"],
            entry["campaign"][:33],
            "[red bold]⚠[/]" if is_dupe else "[dim]—[/]")
    console.print(lt)

    if dupes:
        console.print()
        dupe_lines=[]
        for em,entries in list(dupes.items())[:20]:
            dupe_lines.append(f"  [yellow]{em}[/]")
            for e in entries:
                sc=STATUS_COLOR.get(e['status'],'white')
                dupe_lines.append(
                    f"    → [{sc}]{e['status']}[/]  "
                    f"[dim]{e['campaign'][:40]}[/]  "
                    f"(reply: {e['ts_reply']}  crm: {e['ts_crm']})")
        console.print(Panel(
            f"[bold red]⚠  {len(dupes)} duplicate email(s) across campaigns![/]\n\n"
            +"\n".join(dupe_lines),
            title="[bold red]DUPLICATE LEADS[/]",border_style="red"))
        if len(dupes)>20:
            console.print(f"[dim]  … and {len(dupes)-20} more in export.[/]")
    else:
        console.print(Panel(
            "[green]✓ No duplicate leads found across selected campaigns.[/]",
            border_style="green",expand=False))

    # ── Export ────────────────────────────────────────────────────────────────
    console.print()
    fmt_pick=questionary.select("Export format?",choices=[
        "📊  Excel (.xlsx)  — formatted, colored headers",
        "📄  CSV (.csv)     — plain, for manual paste",
        "⏭   Skip",
    ],style=Q_STYLE).ask() or ""

    # date label for filename and headers
    dr_label=(f"{start_date} to {end_date}" if start_date or end_date
              else "All time")
    def _fmt_date(d:str)->str:
        """Convert YYYY-MM-DD → DD-MMM-YYYY, or return as-is."""
        try: return datetime.strptime(d,"%Y-%m-%d").strftime("%d-%b-%Y")
        except: return d or "All time"
    dr_pretty=f"{_fmt_date(start_date)} to {_fmt_date(end_date or datetime.now().strftime('%Y-%m-%d'))}"

    tag=datetime.now().strftime("%d_%b_%y").lower()
    b_=bench

    # shared data builder — returns list of (label, m, steps, is_unified)
    def _all_sections():
        for cid in selected_ids:
            cname=sel_map[cid].get("name","Unnamed")
            cov=camp_overviews.get(cid)
            if not cov: continue
            yield cname, analytics_extract_metrics(cov, steps=camp_steps.get(cid,[])), camp_steps.get(cid,[]), False
        yield f"TOTAL — {len(selected_ids)} Campaigns", unified, sorted_steps, True

    export_ctx=dict(
        bench=bench,follow_ups=follow_ups,sales_usd=sales_usd,
        selected_ids=selected_ids,sel_map=sel_map,
        camp_overviews=camp_overviews,camp_steps=camp_steps,
        unified=unified,sorted_steps=sorted_steps,
        email_appearances=email_appearances,dupes=dupes,
        fetch_timestamps=fetch_timestamps,tag=tag,
    )

    if "Excel" in fmt_pick:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            console.print("[yellow]openpyxl not found — installing…[/]")
            import subprocess, sys
            subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","--quiet"])
            console.print("[green]✓[/] openpyxl installed.")
        xbytes=analytics_workbook_to_bytes(export_ctx,client_meta,dr_pretty)
        if not xbytes:
            console.print("[red]Could not build Excel (openpyxl missing).[/]")
        else:
            out_path=Path.home()/"Downloads"/f"analytics_{client_meta['name']}_{tag}.xlsx"
            out_path.write_bytes(xbytes)
            console.print(f"[green]✓[/] Excel saved → [yellow]{out_path}[/]")

    elif "CSV" in fmt_pick:
        out_path=Path.home()/"Downloads"/f"analytics_{client_meta['name']}_{tag}.csv"
        csv_bytes=analytics_csv_to_bytes(export_ctx,client_meta,dr_pretty)
        out_path.write_bytes(csv_bytes)
        console.print(f"[green]✓[/] CSV saved → [yellow]{out_path}[/]")

    # Save date prefs to profile (not campaign IDs — auto-selection handles that)
    if questionary.confirm("Save date range preference to profile?",default=False,style=Q_STYLE).ask():
        if start_date: prof["last_start_date"]=start_date
        if end_date:   prof["last_end_date"]=end_date
        # Remove old saved_campaign_ids — no longer needed
        prof.pop("saved_campaign_ids",None)
        prof.pop("last_campaign_ids",None)
        analytics_profile_save(prof)
        console.print(f"[green]✓[/] Saved → {ANALYTICS_DIR}/{prof['name']}.json")


def cmd_analytics_profiles(_args):
    """Manage analytics report profiles."""
    while True:
        all_p=analytics_profiles_all()
        action=questionary.select("Analytics profiles:",choices=[
            "📋  List profiles","➕  Create new","✏️   Edit profile","🗑   Delete","✅  Done"
        ],style=Q_STYLE).ask()
        if not action or "Done" in action: break
        if "List" in action:
            t=Table(box=box.ROUNDED,show_lines=True,title="Analytics Profiles")
            t.add_column("Name",style="cyan bold"); t.add_column("Client"); t.add_column("Benchmarks",style="dim")
            for n,p in all_p.items():
                b=p.get("benchmarks",{})
                bstr=f"reply≥{b.get('reply_rate',0)*100:.0f}%  bounce≤{b.get('bounce_rate',0)*100:.0f}%"
                t.add_row(n,",".join(p.get("client_names",["(ask)"])),bstr)
            console.print(t)
        elif "Create" in action:
            name=questionary.text("Profile name (e.g. will, brad, gd):",style=Q_STYLE).ask()
            if not name: continue
            p=dict(ANALYTICS_PROFILE_TEMPLATE); p["name"]=name.strip()
            p["display_name"]=questionary.text("Display name:",style=Q_STYLE).ask() or name
            # Optionally pin a client
            cfg=cfg_load(); clients=cfg.get("instantly_clients",[])
            if clients:
                choices=["(ask each run)"]+[c["name"] for c in clients]
                pick=questionary.select("Default client:",choices=choices,style=Q_STYLE).ask()
                if pick and pick!="(ask each run)": p["client_names"]=[pick]
            analytics_profile_save(p)
            console.print(f"[green]✓[/] Created. Edit benchmarks + manual fields in: {ANALYTICS_DIR}/{name}.json")
        elif "Edit" in action:
            if not all_p: continue
            n=questionary.select("Which?",choices=list(all_p.keys()),style=Q_STYLE).ask()
            if not n: continue
            p=all_p[n]
            console.print(Syntax(json.dumps(p,indent=2),"json",theme="monokai"))
            console.print(f"[dim]Edit directly: {ANALYTICS_DIR}/{n}.json[/]")
            # Quick benchmark editor
            if questionary.confirm("Edit benchmarks interactively?",default=False,style=Q_STYLE).ask():
                b=p.get("benchmarks",{})
                for key,label in [("reply_rate","Reply rate (e.g. 0.05)"),
                                   ("positive_reply_rate","Positive reply rate (e.g. 0.25)"),
                                   ("bounce_rate","Bounce rate max (e.g. 0.05)"),
                                   ("open_rate","Open rate (e.g. 0.50)"),
                                   ("meeting_book_rate","Meeting book rate (e.g. 0.25)"),
                                   ("meeting_attend_rate","Meeting attend rate (e.g. 0.80)")]:
                    v=questionary.text(f"{label} [{b.get(key,'')}]:",style=Q_STYLE).ask()
                    try: b[key]=float(v)
                    except: pass
                p["benchmarks"]=b; analytics_profile_save(p)
                console.print(f"[green]✓[/] Benchmarks saved.")
        elif "Delete" in action:
            rm=questionary.checkbox("Delete:",choices=list(all_p.keys()),style=Q_STYLE).ask()
            for n in (rm or []):
                (ANALYTICS_DIR/f"{n}.json").unlink(missing_ok=True)
            console.print(f"[green]✓[/] Deleted: {rm}")

def main():
    if use_env_config():
        log.info("mailclaw: env-only config (no ~/.mailclaw JSON); RAILWAY_ENVIRONMENT=%r",
                 os.environ.get("RAILWAY_ENVIRONMENT", ""))
    else:
        log.debug("mailclaw: config file mode → %s", CONFIG_FILE)
    console.print(LOGO)
    console.rule(style="dim yellow")
    p=argparse.ArgumentParser(prog="mailclaw",add_help=True,
        description="Cold email pipeline: verify · enrich · upload · analytics")
    p.add_argument("--debug", action="store_true",
                   help="Write verbose debug logs to ~/.mailclaw/mailclaw_debug.log")
    s=p.add_subparsers(dest="cmd")
    s.add_parser("run",      help="Full interactive pipeline (default)")
    s.add_parser("verify",   help="Email verification only")
    s.add_parser("enrich",   help="AI enrichment only")
    s.add_parser("upload",   help="Upload to Instantly only")
    s.add_parser("map",      help="Column mapping only")
    s.add_parser("balance",  help="Check Reoon + AI credits")
    s.add_parser("config",   help="Manage keys & settings")
    s.add_parser("profiles", help="Manage enrichment profiles")
    s.add_parser("clients",  help="Manage Instantly clients")
    s.add_parser("onboard",  help="First-time setup wizard")
    s.add_parser("bot",      help="Telegram bot")

    an=s.add_parser("analytics", help="Campaign analytics report")
    an.add_argument("--profile","-p", default=None,
                    help="Analytics profile name e.g. --profile will")
    an.add_argument("--start","-s", default=None, metavar="YYYY-MM-DD",
                    help="Report start date (optional)")
    an.add_argument("--end","-e", default=None, metavar="YYYY-MM-DD",
                    help="Report end date (optional)")
    s.add_parser("analytics-profiles", help="Manage analytics report profiles")
    aq=s.add_parser("ask", help='Ask AI a question: mailclaw ask "emails sent yesterday?"')
    aq.add_argument("question", nargs="*", help="Your question in plain English")
    aq.add_argument("--profile","-p", default=None, help="Analytics profile name")
    aq.add_argument(
        "--export",
        choices=["auto", "csv", "xlsx", "both"],
        default="auto",
        metavar="MODE",
        help="auto=only attach files if question mentions csv/excel/download; csv|xlsx|both=always build "
        "full analytics export (saved to ~/Downloads on CLI, Telegram attachments on bot)",
    )

    args=p.parse_args()
    if getattr(args, "debug", False):
        _enable_debug_logging()
    {
        "verify":            cmd_verify,
        "enrich":            cmd_enrich,
        "upload":            cmd_upload,
        "map":               cmd_map,
        "balance":           cmd_balance,
        "config":            cmd_config,
        "profiles":          cmd_profiles,
        "clients":           cmd_clients,
        "onboard":           lambda _: run_onboarding(),
        "bot":               cmd_bot,
        "analytics":         cmd_analytics,
        "analytics-profiles":cmd_analytics_profiles,
        "ask":               cmd_ask,
    }.get(args.cmd, cmd_run)(args)

if __name__ == "__main__":
    main()
